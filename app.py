from flask import Flask, render_template, request, redirect, url_for, session, flash
from flask_pymongo import PyMongo
from flask import send_from_directory 
from flask import make_response
from flask import jsonify, request, session, send_file, render_template, redirect, url_for, send_from_directory
from flask import Flask, render_template, request, abort
from flask import jsonify
from flask import render_template, session, redirect, url_for, flash
from werkzeug.security import generate_password_hash, check_password_hash
import smtplib, random, os
from dotenv import load_dotenv
from email.mime.text import MIMEText
import re
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
from bson import ObjectId
import json
from pytz import timezone
import requests
import time
import docx
import mimetypes
from utils.text_extractor import extract_text
from utils.similarity import calculate_similarity,  calculate_cosine_similarity, highlight_matches
from head_pose_detector import head_pose_detector
import numpy as np
import cv2
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import calendar
from dateutil.relativedelta import relativedelta
import threading
from datetime import datetime, date
from datetime import datetime, timedelta
from bson import ObjectId
import os
import csv
import traceback
import docx
from datetime import datetime, timedelta
from collections import defaultdict
from datetime import datetime, timedelta
import pytz

app = Flask(__name__)
app.secret_key = 'supersecretkey'  #  will be used for securely signing the session

# MongoDB connection
app.config['MONGO_URI'] = 'mongodb://localhost:27017/exam_system'
mongo = PyMongo(app)

# Check if it is working well, Head pose detector is already initialized in head_pose_detector.py
if head_pose_detector is not None:
    print("Head Pose Detector loaded successfully")
else:
    print("Head Pose Detector not available")


# Home route → redirect to login
@app.route('/')
def home():
    return redirect(url_for('login'))

# ============================
# Login Route
# ============================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user_id = request.form.get('user_id')
        password = request.form.get('password')

        user = mongo.db.users.find_one({'user_id': user_id})
        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['user_id']
            session['name'] = user['name']
            session['role'] = user['role']

            # Role-based redirection
            if user['role'] == 'student':
                return redirect(url_for('student_dashboard'))
            elif user['role'] == 'lecturer':
                return redirect(url_for('lec_dashboard'))
            else:
                flash('Unknown role assigned to user.', 'danger')
                return redirect(url_for('login'))
        else:
            flash('Invalid UserID or Password', 'danger')

    return render_template('login.html')

# ============================
# Register
# ============================
@app.route('/verify-otp', methods=['POST'])
def verify_otp():
        data = request.get_json()
        email = data.get('email')
        otp_input = data.get('otp')

        record = mongo.db.otps.find_one({'email': email})
        if not record or record['otp'] != otp_input:
            return jsonify(success=False, error="Invalid or expired OTP code")
        return jsonify(success=True)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'GET':
        return render_template('register.html')

    # POST logic
    user_id = request.form.get('user_id')
    email = request.form.get('email')
    password = request.form.get('password')
    name = request.form.get('name')
    role = request.form.get('role')
    phone = request.form.get('phone')
    gender = request.form.get('gender')
    otp_input = request.form.get('otp')

    # PASSWORD VALIDATION
    if not re.match(r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[^A-Za-z0-9])[A-Za-z\d\W]{8,}$', password):
        flash('Password must be at least 8 characters with uppercase, lowercase, number, and special character.', 'danger')
        return redirect(url_for('register'))

    # EMAIL VALIDATION
    if not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email):
        flash('Please enter a valid email address.', 'danger')
        return redirect(url_for('register'))

    hashed_pw = generate_password_hash(password)

    # Check duplicate user ID
    if mongo.db.users.find_one({'user_id': user_id}):
        flash('User ID already exists', 'danger')
        return redirect(url_for('register'))

    # Check duplicate email
    if mongo.db.users.find_one({'email': email}):
        flash('Email already exists', 'danger')
        return redirect(url_for('register'))
    
    # Save user
    mongo.db.users.insert_one({
        'user_id': user_id,
        'name': name,
        'email': email,
        'password': hashed_pw,
        'role': role,
        'phone': phone,
        'gender': gender,
        'photo_url': "/static/img/default-profile.jpg",
        'created_at': datetime.now()
    })

    mongo.db.otps.delete_one({'email': email})
    flash('Registration successful! Please login.', 'success')
    return redirect(url_for('login'))

load_dotenv()  # Load .env file

@app.route('/send-otp', methods=['POST'])
def send_otp():
    data = request.get_json()
    email = data.get('email')
    if not email:
        return jsonify(success=False, error="Email is required"), 400

    otp = str(random.randint(100000, 999999))

    # Try to send email FIRST
    try:
        subject = "Email Verification - Online Examination System"
        body = f"""
        <html>
        <head>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    background-color: #f4f4f4;
                    margin: 0;
                    padding: 20px;
                }}
                .container {{
                    max-width: 600px;
                    margin: 0 auto;
                    background-color: white;
                    border-radius: 10px;
                    overflow: hidden;
                    box-shadow: 0 5px 15px rgba(0,0,0,0.1);
                }}
                .header {{
                    background: #004080;
                    color: white;
                    padding: 30px;
                    text-align: center;
                }}
                .content {{
                    padding: 30px;
                }}
                .otp-box {{
                    background: #f8f9fa;
                    border: 2px dashed #667eea;
                    border-radius: 8px;
                    padding: 20px;
                    text-align: center;
                    margin: 20px 0;
                }}
                .otp-code {{
                    font-size: 32px;
                    font-weight: bold;
                    color: #667eea;
                    letter-spacing: 4px;
                    margin: 10px 0;
                }}
                .footer {{
                    background: #f8f9fa;
                    padding: 20px;
                    text-align: center;
                    color: #6c757d;
                    font-size: 12px;
                }}
                .warning {{
                    background: #fff3cd;
                    border: 1px solid #ffeaa7;
                    border-radius: 5px;
                    padding: 15px;
                    margin: 20px 0;
                    color: #856404;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>Email Verification</h1>
                    <p>Online Examination System</p>
                </div>
                <div class="content">
                    <h2>Welcome to Online Examination System!</h2>
                    <p>Thank you for registering. Please use the verification code below to complete your registration:</p>
                    
                    <div class="otp-box">
                        <p>Your verification code is:</p>
                        <div class="otp-code">{otp}</div>
                    </div>
                    
                    <div class="warning">
                        <strong>Security Notice:</strong>
                        <ul>
                            <li>Never share this code with anyone</li>
                            <li>If you didn't request this, please ignore this email</li>
                        </ul>
                    </div>
                    
                    <p>If you have any questions, please contact our support team.</p>
                    <p>Best regards,<br>ProctorTrack System Team</p>
                </div>
                <div class="footer">
                    <p>This is an automated email. Please do not reply to this message.</p>
                    <p>&copy; 2025 ProctorTrack Sdn. Bhd. All rights reserved.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        msg = MIMEText(body, 'html')
        msg['Subject'] = subject
        msg['From'] = os.getenv("EMAIL_USER")
        msg['To'] = email
        
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(os.getenv("EMAIL_USER"), os.getenv("EMAIL_PASS"))
            server.send_message(msg)
        
        # ONLY save to database if email was sent successfully
        mongo.db.otps.update_one(
            {'email': email},
            {'$set': {'otp': otp, 'created_at': datetime.utcnow()}},
            upsert=True
        )
        
        return jsonify(success=True, message="OTP sent successfully")
        
    except Exception as e:
        print(f"Email sending failed: {e}")
        # Don't save OTP if email failed
        return jsonify(success=False, error="Failed to send OTP. Please try again.")
    
@app.route('/check-user-id', methods=['POST'])
def check_user_id():
    """Check if user ID already exists"""
    data = request.get_json()
    user_id = data.get('user_id')
    
    if not user_id:
        return jsonify(exists=False, error="User ID is required"), 400
    
    existing_user = mongo.db.users.find_one({'user_id': user_id})
    return jsonify(exists=bool(existing_user))

@app.route('/check-email-duplicate', methods=['POST'])
def check_email_duplicate():
    """Check if email already exists"""
    data = request.get_json()
    email = data.get('email')
    
    if not email:
        return jsonify(exists=False, error="Email is required"), 400
    
    # Email format validation
    if not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email):
        return jsonify(exists=False, error="Invalid email format"), 400
    
    existing_user = mongo.db.users.find_one({'email': email})
    return jsonify(exists=bool(existing_user))

# ============================
# Forgot Password
# ============================
@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'GET':
        return render_template('forgot_password.html')
    
    # Handle POST request (fallback for non-AJAX requests)
    if request.method == 'POST':
        email = request.form.get('email')
        if email:
            # Redirect to the main forgot password page with email pre-filled
            return redirect(url_for('forgot_password') + f'?email={email}')
    
    return render_template('forgot_password.html')

@app.route('/forgot-password-otp', methods=['POST'])
def forgot_password_otp():
    """Send OTP for password reset"""
    data = request.get_json()
    email = data.get('email')
    
    if not email:
        return jsonify(success=False, error="Email is required"), 400
    
    # Check if user exists with this email
    user = mongo.db.users.find_one({'email': email})
    if not user:
        return jsonify(success=False, error="Email not found"), 404
    
    # Generate OTP
    otp = str(random.randint(100000, 999999))
    
    # Try to send email first
    try:
        subject = "Password Reset - Online Examination System"
        body = f"""
        <html>
        <head>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    background-color: #f4f4f4;
                    margin: 0;
                    padding: 20px;
                }}
                .container {{
                    max-width: 600px;
                    margin: 0 auto;
                    background-color: white;
                    border-radius: 10px;
                    overflow: hidden;
                    box-shadow: 0 5px 15px rgba(0,0,0,0.1);
                }}
                .header {{
                    background: #004080;
                    color: white;
                    padding: 30px;
                    text-align: center;
                }}
                .content {{
                    padding: 30px;
                }}
                .otp-box {{
                    background: #f8f9fa;
                    border: 2px dashed #667eea;
                    border-radius: 8px;
                    padding: 20px;
                    text-align: center;
                    margin: 20px 0;
                }}
                .otp-code {{
                    font-size: 32px;
                    font-weight: bold;
                    color: #667eea;
                    letter-spacing: 4px;
                    margin: 10px 0;
                }}
                .footer {{
                    background: #f8f9fa;
                    padding: 20px;
                    text-align: center;
                    color: #6c757d;
                    font-size: 12px;
                }}
                .warning {{
                    background: #fff3cd;
                    border: 1px solid #ffeaa7;
                    border-radius: 5px;
                    padding: 15px;
                    margin: 20px 0;
                    color: #856404;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>Password Reset Request</h1>
                    <p>Online Examination System</p>
                </div>
                <div class="content">
                    <h2>Hello {user['name']},</h2>
                    <p>We received a request to reset your password. Use the verification code below to proceed:</p>
                    
                    <div class="otp-box">
                        <p>Your verification code is:</p>
                        <div class="otp-code">{otp}</div>
                        <p><strong>Valid for 1 minute</strong></p>
                    </div>
                    
                    <div class="warning">
                        <strong>Security Notice:</strong>
                        <ul>
                            <li>This code will expire in 1 minute</li>
                            <li>Never share this code with anyone</li>
                            <li>If you didn't request this, please ignore this email</li>
                        </ul>
                    </div>
                    
                    <p>If you have any questions, please contact our support team.</p>
                    <p>Best regards,<br>ProctorTrack System Team</p>
                </div>
                <div class="footer">
                    <p>This is an automated email. Please do not reply to this message.</p>
                    <p>&copy; 2025 ProctorTrack Sdn. Bhd. All rights reserved.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        msg = MIMEText(body, 'html')
        msg['Subject'] = subject
        msg['From'] = os.getenv("EMAIL_USER")
        msg['To'] = email
        
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(os.getenv("EMAIL_USER"), os.getenv("EMAIL_PASS"))
            server.send_message(msg)
        
        # Save OTP to database with expiration time
        expiry_time = datetime.utcnow() + timedelta(minutes=1)
        
        # Use a separate collection for password reset OTPs
        mongo.db.reset_otps.update_one(
            {'email': email},
            {
                '$set': {
                    'otp': otp,
                    'created_at': datetime.utcnow(),
                    'expires_at': expiry_time,
                    'attempts': 0,
                    'verified': False
                }
            },
            upsert=True
        )
        
        return jsonify(success=True, message="OTP sent successfully")
        
    except Exception as e:
        print(f"Email sending failed: {e}")
        return jsonify(success=False, error="Failed to send OTP. Please try again.")

@app.route('/verify-reset-otp', methods=['POST'])
def verify_reset_otp():
    """Verify OTP for password reset"""
    data = request.get_json()
    email = data.get('email')
    otp_input = data.get('otp')
    
    if not email or not otp_input:
        return jsonify(success=False, error="Email and OTP are required"), 400
    
    # Find OTP record
    otp_record = mongo.db.reset_otps.find_one({'email': email})
    
    if not otp_record:
        return jsonify(success=False, error="OTP not found or expired"), 400
    
    # Check if expired
    if datetime.utcnow() > otp_record['expires_at']:
        mongo.db.reset_otps.delete_one({'email': email})
        return jsonify(success=False, error="OTP has expired. Please request a new one."), 400
    
    # Check attempt limit
    if otp_record.get('attempts', 0) >= 5:
        mongo.db.reset_otps.delete_one({'email': email})
        return jsonify(success=False, error="Too many attempts. Please request a new OTP."), 400
    
    # Verify OTP
    if otp_record['otp'] != otp_input:
        # Increment attempts
        mongo.db.reset_otps.update_one(
            {'email': email},
            {'$inc': {'attempts': 1}}
        )
        return jsonify(success=False, error="Invalid OTP. Please try again."), 400

    # Mark as verified AND extend expiry for password reset (10 minutes from now)
    mongo.db.reset_otps.update_one(
        {'email': email},
        {
            '$set': {
                'verified': True,
                'expires_at': datetime.utcnow() + timedelta(minutes=10)  # Extended to 10 minutes
            }
        }
    )
    
    return jsonify(success=True, message="OTP verified successfully")

@app.route('/reset-password-final', methods=['POST'])
def reset_password_final():
    """Final step - reset the password"""
    data = request.get_json()
    email = data.get('email')
    new_password = data.get('newPassword')
    
    if not email or not new_password:
        return jsonify(success=False, error="Email and new password are required"), 400
    
    # Check if OTP was verified
    otp_record = mongo.db.reset_otps.find_one({'email': email})
    if not otp_record or not otp_record.get('verified'):
        return jsonify(success=False, error="OTP verification required"), 400
    
    # Check if OTP is still valid (not expired)
    if datetime.utcnow() > otp_record['expires_at']:
        mongo.db.reset_otps.delete_one({'email': email})
        return jsonify(success=False, error="OTP has expired. Please start over."), 400
    
    # Validate password strength
    if not re.match(r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[^A-Za-z0-9])[A-Za-z\d\W]{8,}$', new_password):
        return jsonify(success=False, error="Password must be at least 8 characters with uppercase, lowercase, number, and special character"), 400
    
    # Check if user exists
    user = mongo.db.users.find_one({'email': email})
    if not user:
        return jsonify(success=False, error="User not found"), 404
    
    try:
        # Hash the new password
        hashed_password = generate_password_hash(new_password)
        
        # Update user password
        result = mongo.db.users.update_one(
            {'email': email},
            {
                '$set': {
                    'password': hashed_password,
                    'password_updated_at': datetime.utcnow()
                }
            }
        )
        
        if result.modified_count > 0:
            # Clean up OTP record
            mongo.db.reset_otps.delete_one({'email': email})
            
            # Send confirmation email
            send_password_reset_confirmation(email, user['name'])
            
            return jsonify(success=True, message="Password reset successfully")
        else:
            return jsonify(success=False, error="Failed to update password"), 500
            
    except Exception as e:
        print(f"Password update error: {e}")
        return jsonify(success=False, error="An error occurred while updating password"), 500

def send_password_reset_confirmation(email, user_name):
    """Send confirmation email after password reset"""
    try:
        subject = "Password Reset Confirmation - Online Examination System"
        body = f"""
        <html>
        <head>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    background-color: #f4f4f4;
                    margin: 0;
                    padding: 20px;
                }}
                .container {{
                    max-width: 600px;
                    margin: 0 auto;
                    background-color: white;
                    border-radius: 10px;
                    overflow: hidden;
                    box-shadow: 0 5px 15px rgba(0,0,0,0.1);
                }}
                .header {{
                    background: #28a745;
                    color: white;
                    padding: 30px;
                    text-align: center;
                }}
                .content {{
                    padding: 30px;
                }}
                .success-box {{
                    background: #d4edda;
                    border: 1px solid #c3e6cb;
                    border-radius: 8px;
                    padding: 20px;
                    text-align: center;
                    margin: 20px 0;
                    color: #155724;
                }}
                .footer {{
                    background: #f8f9fa;
                    padding: 20px;
                    text-align: center;
                    color: #6c757d;
                    font-size: 12px;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>Password Reset Successful</h1>
                    <p>Online Examination System</p>
                </div>
                <div class="content">
                    <h2>Hello {user_name},</h2>
                    <div class="success-box">
                        <h3>✅ Password Updated Successfully!</h3>
                        <p>Your password has been reset successfully.</p>
                        <p><strong>Reset Time:</strong> {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC</p>
                    </div>
                    
                    <p>You can now log in to your account using your new password.</p>
                    
                    <p>If you did not make this change, please contact our support team immediately.</p>
                    
                    <p>Best regards,<br>ProctorTrack System Team</p>
                </div>
                <div class="footer">
                    <p>This is an automated email. Please do not reply to this message.</p>
                    <p>&copy; 2025 ProctorTrack Sdn. Bhd. All rights reserved.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        msg = MIMEText(body, 'html')
        msg['Subject'] = subject
        msg['From'] = os.getenv("EMAIL_USER")
        msg['To'] = email
        
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(os.getenv("EMAIL_USER"), os.getenv("EMAIL_PASS"))
            server.send_message(msg)
            
    except Exception as e:
        print(f"Confirmation email sending failed: {e}")
        # Don't fail the whole process if confirmation email fails

# ============================
# Cleanup Functions
# ============================
def cleanup_expired_otps():
    """Clean up expired OTPs from both collections"""
    try:
        current_time = datetime.utcnow()
        
        # Clean up registration OTPs
        mongo.db.otps.delete_many({'created_at': {'$lt': current_time - timedelta(minutes=10)}})
        
        # Clean up reset OTPs
        mongo.db.reset_otps.delete_many({'expires_at': {'$lt': current_time}})
    
    except Exception as e:
        print(f"OTP cleanup error: {e}")

# Run cleanup every 5 minutes
def background_cleanup():
    """Background task to clean up expired OTPs"""
    while True:
        cleanup_expired_otps()
        time.sleep(300)  # Sleep for 5 minutes

# Start the background cleanup thread
cleanup_thread = threading.Thread(target=background_cleanup)
cleanup_thread.daemon = True
cleanup_thread.start()

# ============================
# Additional Helper Functions
# ============================
@app.route('/check-email-exists', methods=['POST'])
def check_email_exists():
    """Check if email exists in the system (for client-side validation)"""
    data = request.get_json()
    email = data.get('email')
    
    if not email:
        return jsonify(exists=False, error="Email is required"), 400
    
    user = mongo.db.users.find_one({'email': email})
    return jsonify(exists=bool(user))


# ============================
# Student Dashboard
# ============================
@app.route('/student/dashboard')
def student_dashboard():
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))
    
    user_id = session['user_id']
    user = mongo.db.users.find_one({'user_id': user_id})
    
    # ============================
    # Get Student's Enrolled Courses
    # ============================
    enrolled_courses = list(mongo.db.courses.find({"students": user_id}))
    courses_count = len(enrolled_courses)
    
    # ============================
    # Get Student's Assessments
    # ============================
    student_assessments = list(mongo.db.assessments.find({"students": user_id}))
    
    # ============================
    # Get Upcoming Exams (with timetable)
    # ============================
    upcoming_exams = []
    
    for assessment in student_assessments:
        # Find timetable for this assessment
        timetable = mongo.db.exam_timetables.find_one({"assessment_code": assessment["assessment_code"]})
        if timetable:
            exam_date_str = timetable["exam_date"]
            try:
                exam_date = datetime.strptime(exam_date_str, "%Y-%m-%d").date()
                # Only include future exams
                if exam_date >= date.today():
                    exam_info = {
                        "assessment_code": assessment["assessment_code"],
                        "title": assessment["title"],
                        "exam_date": exam_date,
                        "start_time": timetable["start_time"],
                        "end_time": timetable["end_time"],
                        "date_formatted": exam_date.strftime("%b %d"),
                        "time_info": f"{timetable['start_time']} - {timetable['end_time']}"
                    }
                    upcoming_exams.append(exam_info)
            except ValueError:
                # Handle invalid date format
                pass
    
    # Sort by exam date
    upcoming_exams.sort(key=lambda x: x["exam_date"])
    upcoming_exams_count = len(upcoming_exams)
    
    # ============================
    # Get Student Submissions
    # ============================
    recent_submissions = []
    
    # Get student's submissions (latest 3)
    submissions = list(mongo.db.submissions.find({"student_id": user_id}).sort("submitted_at", -1).limit(3))
    
    for submission in submissions:
        # Get assessment title
        assessment = mongo.db.assessments.find_one({"assessment_code": submission["assessment_code"]})
        
        # Check if there's a resubmission request for this submission
        resubmission_request = mongo.db.resubmission_requests.find_one({
            "student_id": user_id,
            "assessment_code": submission["assessment_code"]
        })
        
        submission_info = {
            "title": assessment["title"] if assessment else f"Assignment {submission['assessment_code']}",
            "submission_date": submission["submitted_at"],
            "has_resubmission_request": resubmission_request is not None,
            "resubmission_status": resubmission_request.get("status", "") if resubmission_request else ""
        }
        recent_submissions.append(submission_info)
    
    # Calculate submission statistics
    submissions_count = mongo.db.submissions.count_documents({"student_id": user_id})
    
    # Count pending resubmission requests (status != "Approved" and status != "Rejected")
    pending_requests = mongo.db.resubmission_requests.count_documents({
        "student_id": user_id, 
        "status": {"$nin": ["Approved", "Rejected"]}
    })
    
    # If no explicit status field or status is empty, also count those as pending
    pending_requests += mongo.db.resubmission_requests.count_documents({
        "student_id": user_id,
        "$or": [
            {"status": {"$exists": False}},
            {"status": ""}
        ]
    })
    
    # ============================
    # Get Latest Announcements
    # ============================
    try:
        # Get active announcements (latest 3 for dashboard)
        announcements = list(mongo.db.announcements.find({
            'is_active': True
        }).sort('created_at', -1).limit(3))
        
        # Format announcement dates and creator info
        for announcement in announcements:
            if announcement.get('created_at'):
                announcement['formatted_date'] = announcement['created_at'].strftime('%b %d, %Y')
            else:
                announcement['formatted_date'] = 'Recent'
            
            # Get creator name
            creator = mongo.db.users.find_one({'user_id': announcement.get('created_by')})
            announcement['creator_name'] = creator.get('name', 'Administration') if creator else 'Administration'
    
    except Exception as e:
        print(f"Error loading announcements: {e}")
        announcements = []
    
    # ============================
    # Prepare Dashboard Data
    # ============================
    dashboard_data = {
        'user': user,
        'enrolled_courses': enrolled_courses,
        'upcoming_exams': upcoming_exams,
        'recent_submissions': recent_submissions,
        'announcements': announcements,
        'stats': {
            'courses_count': courses_count,
            'upcoming_exams_count': upcoming_exams_count,
            'submissions_count': submissions_count,
            'pending_requests': pending_requests
        }
    }
    
    return render_template('student_dashboard.html', **dashboard_data)

# ============================
# Student Announcements
# ============================
@app.route('/student/announcements')
def student_announcements():
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))
    
    try:
        # Get all active announcements, sorted by latest first
        announcements = list(mongo.db.announcements.find({
            'is_active': True
        }).sort('created_at', -1))
        
        # Format dates for display
        for announcement in announcements:
            if announcement.get('created_at'):
                announcement['formatted_date'] = announcement['created_at'].strftime('%b %d, %Y at %I:%M %p')
            else:
                announcement['formatted_date'] = 'Recent'
            
            # Get creator name
            creator = mongo.db.users.find_one({'user_id': announcement.get('created_by')})
            announcement['creator_name'] = creator.get('name', 'Administration') if creator else 'Administration'
        
        return render_template('student_announcements.html', announcements=announcements)
    
    except Exception as e:
        flash(f"Error loading announcements: {str(e)}", "danger")
        return redirect(url_for('student_dashboard'))
                         
@app.context_processor
def inject_user():
    user = mongo.db.users.find_one({'user_id': session.get('user_id')}) if 'user_id' in session else None
    return dict(user=user)

@app.context_processor
def inject_common_data():
    return dict(current_year=datetime.now().year)   

# ============================
# Student Profile
# ============================
@app.route('/profile')
def profile():
    if 'user_id' not in session or session.get('role') != 'student':
        flash('Please log in first.', 'danger')
        return redirect(url_for('login'))

    user = mongo.db.users.find_one({'user_id': session['user_id']})
    return render_template('profile.html', user=user)


@app.route('/edit/profile', methods=['GET', 'POST'])
def edit_profile():
    if 'user_id' not in session or session.get('role') != 'student':
        flash('Please log in first.', 'danger')
        return redirect(url_for('login'))

    user = mongo.db.users.find_one({'user_id': session['user_id']})
    errors = {}

    if request.method == 'POST':
        # Step 1: Get form data
        name = request.form.get('name')
        email = request.form.get('email')
        phone = request.form.get('phone')
        gender = request.form.get('gender')
        photo_file = request.files.get('photo')

        # Validate email format
        if not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email):
            errors['email'] = "Please enter a valid email address."

        # Check if email is already used by another user
        existing_user = mongo.db.users.find_one({'email': email, 'user_id': {'$ne': session['user_id']}})
        if existing_user:
            errors['email'] = "Email already exists."

        # If validation fails
        if errors:
            return render_template('edit_profile.html', user=user, errors=errors)

        update_data = {
            'name': name,
            'email': email,
            'phone': phone,
            'gender': gender
        }

        # Step 2: Handle photo upload
        if photo_file and photo_file.filename != '':
            uploads_folder = os.path.join('static', 'uploads')
            os.makedirs(uploads_folder, exist_ok=True)

            # Get old photo path
            old_photo = user.get('photo_url')
            
            # Define protected/default images that should never be deleted
            PROTECTED_IMAGES = [
                '/static/img/default-profile.jpg'
            ]
            
            # Save new photo first
            filename = secure_filename(photo_file.filename)
            new_filename = f"{ObjectId()}{os.path.splitext(filename)[1]}"
            save_path = os.path.join(uploads_folder, new_filename)
            photo_file.save(save_path)

            # Update the photo_url in update_data
            update_data['photo_url'] = f"/static/uploads/{new_filename}"

            # Delete old photo ONLY if it's not a protected/default image
            if old_photo:
                # Check if it's a protected image
                is_protected = any(protected in old_photo for protected in PROTECTED_IMAGES)
                
                if not is_protected:
                    try:
                        # Remove leading '/' and check if file exists
                        old_file_path = old_photo[1:] if old_photo.startswith('/') else old_photo
                        if os.path.exists(old_file_path):
                            os.remove(old_file_path)
                            print(f"Deleted old profile picture: {old_file_path}")
                    except Exception as e:
                        print(f"Error deleting old profile picture: {e}")
                else:
                    print(f"Skipped deletion of protected image: {old_photo}")

        # Step 3: Update MongoDB
        mongo.db.users.update_one(
            {'user_id': session['user_id']},
            {'$set': update_data}
        )

        # Step 4: Update session name so navbar shows correct value
        session['name'] = name

        flash('Profile updated successfully!', 'success_redirect')
        return redirect(url_for('profile'))

    return render_template('edit_profile.html', user=user, errors={})


@app.route('/reset/password', methods=['GET', 'POST'])
def reset_password():
    if 'user_id' not in session or session.get('role') != 'student':
        flash("You must be logged in to reset your password.", "danger")
        return redirect(url_for('login'))

    errors = {}

    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']

        user = get_user_from_session(session['user_id'])

        if not user:
            errors['current_password'] = "User not found."
        elif not check_password(user['password'], current_password):
            errors['current_password'] = "Current password is incorrect."

        # PASSWORD VALIDATION
        if not re.match(r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[^A-Za-z0-9])[A-Za-z\d\W]{8,}$', new_password):
            errors['new_password'] = "Password must be at least 8 characters with uppercase, lowercase, number, and special character."

        if new_password != confirm_password:
            errors['confirm_password'] = "Passwords do not match."

        if not errors:
            update_user_password(session['user_id'], new_password)
            flash("Password updated successfully!", "success_redirect")
            return redirect(url_for('profile'))

    return render_template('reset_password.html', errors=errors)

# --- UTILITIES ---
def get_user_from_session(user_id):
    return mongo.db.users.find_one({"user_id": user_id})

def check_password(stored_hashed, entered_plain):
    return check_password_hash(stored_hashed, entered_plain)

def update_user_password(user_id, new_password):
    hashed_pw = generate_password_hash(new_password)
    mongo.db.users.update_one({'user_id': user_id}, {'$set': {'password': hashed_pw}})

# ============================
# Lecturer Dashboard 
# ============================
@app.route('/lec/dashboard')
def lec_dashboard():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    user = mongo.db.users.find_one({'user_id': session['user_id']})
    session['name'] = user.get('name')  # Ensure session.name is set for navbar
    
    # Calculate dynamic statistics
    try:
        # 1. Total Students - Count users where role = "student"
        total_students = mongo.db.users.count_documents({'role': 'student'})
        
        # 2. Active Courses - Count all courses
        active_courses = mongo.db.courses.count_documents({})
        
        # 3. Total Exams - Count all assessments
        total_exams = mongo.db.assessments.count_documents({})
        
        # 4. Flagged Activities - Count all exam violations
        flagged_activities = mongo.db.exam_violations.count_documents({})
        
        # 5. Plagiarism Cases - Count submissions with plagiarism_score + quotes_score > 20%
        plagiarism_cases = mongo.db.submissions.count_documents({
            '$expr': {
                '$gt': [
                    {'$add': ['$plagiarism_score', '$quotes_score']},
                    20.0
                ]
            }
        })
        
        # Additional data for dashboard sections
        
        # Recent Submissions (last 10)
        recent_submissions = list(mongo.db.submissions.find({}).sort('submitted_at', -1).limit(10))
        for submission in recent_submissions:
            # Get student name
            student = mongo.db.users.find_one({'user_id': submission.get('student_id')})
            submission['student_name'] = student.get('name', 'Unknown') if student else 'Unknown'
            
            # Get assessment title
            assessment = mongo.db.assessments.find_one({'assessment_code': submission.get('assessment_code')})
            submission['assessment_title'] = assessment.get('title', 'Unknown Assessment') if assessment else 'Unknown Assessment'

            # Determine if plagiarism is already checked
            total_score = (submission.get('plagiarism_score') or 0) + (submission.get('quotes_score') or 0)
            submission['is_checked'] = total_score > 0
            submission['is_flagged'] = total_score > 20
        
        # Recent Resubmission Requests (last 10)
        recent_requests = list(mongo.db.resubmission_requests.find({}).sort('requested_at', -1).limit(10))
        for request in recent_requests:
            # Get student name
            student = mongo.db.users.find_one({'user_id': request.get('student_id')})
            request['student_name'] = student.get('name', 'Unknown') if student else 'Unknown'
            
            # Get assessment title
            assessment = mongo.db.assessments.find_one({'assessment_code': request.get('assessment_code')})
            request['assessment_title'] = assessment.get('title', 'Unknown Assessment') if assessment else 'Unknown Assessment'
        
        # Upcoming Exams (next 5)
        from datetime import datetime
        today = datetime.now().strftime('%Y-%m-%d')
        upcoming_exams = list(mongo.db.exam_timetables.find({
            'exam_date': {'$gte': today}
        }).sort('exam_date', 1).limit(5))
        
        for exam in upcoming_exams:
            # Get assessment details
            assessment = mongo.db.assessments.find_one({'assessment_code': exam.get('assessment_code')})
            if assessment:
                exam['title'] = assessment.get('title', 'Unknown Exam')
                exam['student_count'] = len(assessment.get('students', []))
            else:
                exam['title'] = 'Unknown Exam'
                exam['student_count'] = 0
        
        # Course Performance Data
        course_performance = []
        courses = list(mongo.db.courses.find({}))
        for course in courses:
            student_count = len(course.get('students', []))
            
            # Count flagged activities for this course's students
            flagged_count = mongo.db.exam_violations.count_documents({
                'student_id': {'$in': course.get('students', [])}
            })
            
            course_performance.append({
                'course_name': course.get('course_name', 'Unknown Course'),
                'student_count': student_count,
                'flagged_count': flagged_count
            })
        
        # Recent Exam Evidence (last 10 violations)
        recent_evidence = list(mongo.db.exam_evidence.find({}).sort('timestamp', -1).limit(10))
        for evidence in recent_evidence:
            # Get student name
            student = mongo.db.users.find_one({'user_id': evidence.get('student_id')})
            evidence['student_name'] = student.get('name', 'Unknown') if student else 'Unknown'
        
        # Chart Data
        
        # 1. Activity Trends Chart Data (last 7 months)
        
        activity_trends = {
            'labels': [],
            'exams_data': [],
            'flagged_data': [],
            'plagiarism_data': []
        }

        current_date = datetime.now()

        # Last 7 full months, including current
        for i in range(6, -1, -1):
            # Always get the first day of each month
            month_date = current_date.replace(day=1) - relativedelta(months=i)
            month_start = month_date.replace(hour=0, minute=0, second=0, microsecond=0)

            # Calculate next month start for range end
            if month_date.month == 12:
                month_end = month_start.replace(year=month_date.year + 1, month=1)
            else:
                month_end = month_start.replace(month=month_date.month + 1)

            # Add label (e.g., "Jan", "Feb", etc.)
            month_name = calendar.month_abbr[month_date.month]
            activity_trends['labels'].append(month_name)

            # 1. Exams Taken (based on exam_date string comparison)
            month_start_str = month_start.strftime('%Y-%m-%d')
            month_end_str = month_end.strftime('%Y-%m-%d')
            monthly_exams = mongo.db.exam_timetables.count_documents({
                'exam_date': {'$gte': month_start_str, '$lt': month_end_str}
            })

            # 2. Flagged Activities (only 'reviewed' and datetime range)
            monthly_flagged = mongo.db.exam_violations.count_documents({
                'timestamp': {'$gte': month_start, '$lt': month_end},
                'status': 'reviewed'
            })


            # 3. Plagiarism (sum of scores > 20 from submissions)
            monthly_plagiarism = mongo.db.submissions.count_documents({
                'submitted_at': {'$gte': month_start, '$lt': month_end},
                '$expr': {
                    '$gt': [
                        {'$add': ['$plagiarism_score', '$quotes_score']},
                        20.0
                    ]
                }
            })

            # Append to chart data
            activity_trends['exams_data'].append(monthly_exams)
            activity_trends['flagged_data'].append(monthly_flagged)
            activity_trends['plagiarism_data'].append(monthly_plagiarism)
        
        # 2. Submissions by Course Chart Data
        submissions_by_course = {
            'labels': [],
            'data': []
        }
        
        courses_for_chart = list(mongo.db.courses.find({}).limit(5))  # Top 5 courses
        for course in courses_for_chart:
            course_name = course.get('course_name', 'Unknown')
            
            # Count submissions for this course (through assessments)
            course_assessments = list(mongo.db.assessments.find({
                'course_codes': course.get('course_code')
            }))
            
            submission_count = 0
            for assessment in course_assessments:
                count = mongo.db.submissions.count_documents({
                    'assessment_code': assessment.get('assessment_code')
                })
                submission_count += count
            
            submissions_by_course['labels'].append(course_name)
            submissions_by_course['data'].append(submission_count)
        
        # 3. Cheating Detection Types Chart Data
        detection_types = {
            'labels': ['Head Movement', 'Tab Switching', 'Forbidden Shortcut', 'Others'],
            'data': []
        }
        
        # Count different violation types
        head_movement_count = mongo.db.exam_violations.count_documents({
            'violation_type': {'$regex': 'looking', '$options': 'i'}
        })
        
        tab_switch_count = mongo.db.exam_violations.count_documents({
            'violation_type': {'$regex': 'tab', '$options': 'i'}
        })

        forbidden_shortcut_count = mongo.db.exam_violations.count_documents({
            'violation_type': {'$regex': 'shortcut', '$options': 'i'}
        })

        other_violations = mongo.db.exam_violations.count_documents({}) - (head_movement_count + tab_switch_count + forbidden_shortcut_count)

        detection_types['data'] = [head_movement_count, tab_switch_count, forbidden_shortcut_count, max(0, other_violations)]

        # Get Recent Announcements (last 4 active announcements)
        recent_announcements = list(mongo.db.announcements.find({
            'is_active': True
        }).sort('created_at', -1).limit(4))
        
        # Format announcement dates
        for announcement in recent_announcements:
            if announcement.get('created_at'):
                announcement['formatted_date'] = announcement['created_at'].strftime('%b %d, %Y')
            else:
                announcement['formatted_date'] = 'Recent'
        
        # SINGLE comprehensive data dictionary
        dashboard_data = {
            'stats': {
                'total_students': total_students,
                'active_courses': active_courses,
                'total_exams': total_exams,
                'flagged_activities': flagged_activities,
                'plagiarism_cases': plagiarism_cases
            },
            'recent_submissions': recent_submissions,
            'recent_requests': recent_requests,
            'upcoming_exams': upcoming_exams,
            'course_performance': course_performance[:4],  # Limit to 4 for display
            'recent_evidence': recent_evidence,
            'announcements': recent_announcements,
            'chart_data': {  # Chart data
                'activity_trends': activity_trends,
                'submissions_by_course': submissions_by_course,
                'detection_types': detection_types
            }
        }
        
    except Exception as e:
        print(f"Error calculating dashboard data: {e}")

    return render_template('lec_dashboard.html', user=user, **dashboard_data)


# ============================
# Announcements Management
# ============================
@app.route('/lec/announcements')
def lec_announcements():
    """View all announcements"""
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))
    
    try:
        # Get all announcements, sorted by latest first
        announcements = list(mongo.db.announcements.find({}).sort('created_at', -1))
        
        # Format dates for display
        for announcement in announcements:
            if announcement.get('created_at'):
                announcement['formatted_date'] = announcement['created_at'].strftime('%b %d, %Y at %I:%M %p')
            
            # Get creator name
            creator = mongo.db.users.find_one({'user_id': announcement.get('created_by')})
            announcement['creator_name'] = creator.get('name', 'Unknown') if creator else 'System'
        
        return render_template('lec_announcements.html', announcements=announcements)
    
    except Exception as e:
        flash(f"Error loading announcements: {str(e)}", "danger")
        return redirect(url_for('lec_dashboard'))

@app.route('/lec/announcements/create', methods=['GET', 'POST'])
def lec_announcement_create():
    """Create new announcement"""
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        try:
            title = request.form.get('title', '').strip()
            content = request.form.get('content', '').strip()
            priority = request.form.get('priority', 'normal')
            
            # Validation
            if not title or not content:
                flash("Title and content are required.", "danger")
                return render_template('lec_announcement_create.html')
            
            # Create announcement
            announcement_data = {
                'title': title,
                'content': content,
                'priority': priority,
                'created_at': datetime.now(),
                'created_by': session['user_id'],
                'is_active': True
            }
            
            result = mongo.db.announcements.insert_one(announcement_data)
            
            if result.inserted_id:
                flash("Announcement created successfully!", "success")
                return redirect(url_for('lec_announcements'))
            else:
                flash("Failed to create announcement.", "danger")
        
        except Exception as e:
            flash(f"Error creating announcement: {str(e)}", "danger")
    
    return render_template('lec_announcement_create.html')

@app.route('/lec/announcements/edit/<announcement_id>', methods=['GET', 'POST'])
def lec_announcement_edit(announcement_id):
    """Edit existing announcement"""
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))
    
    try:
        # Get announcement
        announcement = mongo.db.announcements.find_one({'_id': ObjectId(announcement_id)})
        if not announcement:
            flash("Announcement not found.", "danger")
            return redirect(url_for('lec_announcements'))
        
        if request.method == 'POST':
            title = request.form.get('title', '').strip()
            content = request.form.get('content', '').strip()
            priority = request.form.get('priority', 'normal')
            
            # Validation
            if not title or not content:
                flash("Title and content are required.", "danger")
                return render_template('lec_announcement_edit.html', announcement=announcement)
            
            # Update announcement
            update_data = {
                'title': title,
                'content': content,
                'priority': priority,
                'updated_at': datetime.now(),
                'updated_by': session['user_id']
            }
            
            result = mongo.db.announcements.update_one(
                {'_id': ObjectId(announcement_id)},
                {'$set': update_data}
            )
            
            if result.modified_count > 0:
                flash("Announcement updated successfully!", "success")
                return redirect(url_for('lec_announcements'))
            else:
                flash("No changes made.", "info")
        
        return render_template('lec_announcement_edit.html', announcement=announcement)
    
    except Exception as e:
        flash(f"Error editing announcement: {str(e)}", "danger")
        return redirect(url_for('lec_announcements'))

@app.route('/lec/announcements/delete/<announcement_id>', methods=['POST'])
def lec_announcement_delete(announcement_id):
    """Delete announcement"""
    if 'user_id' not in session or session.get('role') != 'lecturer':
        if request.is_json or request.headers.get('Content-Type') == 'application/json':
            return jsonify({'message': 'Unauthorized access.'}), 403
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))
    
    try:
        result = mongo.db.announcements.delete_one({'_id': ObjectId(announcement_id)})
        
        if result.deleted_count > 0:
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'message': 'Announcement deleted successfully!'})
            flash("Announcement deleted successfully!", "success")
        else:
            if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'message': 'Announcement not found.'}), 404
            flash("Announcement not found.", "danger")
    
    except Exception as e:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'message': f'Error deleting announcement: {str(e)}'}), 500
        flash(f"Error deleting announcement: {str(e)}", "danger")
    
    return redirect(url_for('lec_announcements'))

@app.route('/lec/announcements/toggle/<announcement_id>', methods=['POST'])
def lec_announcement_toggle(announcement_id):
    """Toggle announcement active status"""
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))
    
    try:
        announcement = mongo.db.announcements.find_one({'_id': ObjectId(announcement_id)})
        if not announcement:
            flash("Announcement not found.", "danger")
            return redirect(url_for('lec_announcements'))
        
        new_status = not announcement.get('is_active', True)
        
        result = mongo.db.announcements.update_one(
            {'_id': ObjectId(announcement_id)},
            {'$set': {'is_active': new_status, 'updated_at': datetime.now()}}
        )
        
        status_text = "activated" if new_status else "deactivated"
        flash(f"Announcement {status_text} successfully!", "success")
    
    except Exception as e:
        flash(f"Error toggling announcement: {str(e)}", "danger")
    
    return redirect(url_for('lec_announcements'))

# ============================
# Lecturer Profile 
# ============================
@app.route('/lec/profile')
def lec_profile():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash('Please log in first.', 'danger')
        return redirect(url_for('login'))

    user = mongo.db.users.find_one({'user_id': session['user_id']})
    return render_template('lec_profile.html', user=user)

@app.route('/lec/edit/profile', methods=['GET', 'POST'])
def lec_edit_profile():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    user = mongo.db.users.find_one({'user_id': session['user_id']})
    errors = {}

    if request.method == 'POST':
        name = request.form.get('name')
        email = request.form.get('email')
        phone = request.form.get('phone')
        gender = request.form.get('gender')
        photo_file = request.files.get('photo')

        # Validate email format
        if not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email):
            errors['email'] = "Please enter a valid email address."

        # Check if email is already used by another user
        existing_user = mongo.db.users.find_one({'email': email, 'user_id': {'$ne': session['user_id']}})
        if existing_user:
            errors['email'] = "Email already exists."

        # If validation fails
        if errors:
            return render_template('lec_edit_profile.html', user=user, errors=errors)

        update_data = {
            'name': name,
            'email': email,
            'phone': phone,
            'gender': gender
        }

        # Handle profile photo
        if photo_file and photo_file.filename != '':
            uploads_folder = os.path.join('static', 'uploads')
            os.makedirs(uploads_folder, exist_ok=True)

            # Delete old photo if exists
            old_photo = user.get('photo_url')
            if old_photo and os.path.exists(old_photo[1:]):
                os.remove(old_photo[1:])

            filename = secure_filename(photo_file.filename)
            new_filename = f"{ObjectId()}{os.path.splitext(filename)[1]}"
            save_path = os.path.join(uploads_folder, new_filename)
            photo_file.save(save_path)

            update_data['photo_url'] = f"/static/uploads/{new_filename}"

        # Update user in database
        mongo.db.users.update_one(
            {'user_id': session['user_id']},
            {'$set': update_data}
        )

        session['name'] = name
        flash('Profile updated successfully!', 'success_redirect')
        return redirect(url_for('lec_profile'))

    return render_template('lec_edit_profile.html', user=user, errors={})


@app.route('/lec/reset/password', methods=['GET', 'POST'])
def lec_reset_password():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("You must be logged in to reset your password.", "danger")
        return redirect(url_for('login'))

    errors = {}

    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']

        user = lec_get_user_from_session(session['user_id'])

        if not user:
            errors['current_password'] = "User not found."
        elif not lec_check_password(user['password'], current_password):
            errors['current_password'] = "Current password is incorrect."

        if not re.match(r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[^A-Za-z0-9])[A-Za-z\d\W].{8,}$', new_password):
            errors['new_password'] = "Password must be at least 8 characters and include uppercase, lowercase, number, and special character."

        if new_password != confirm_password:
            errors['confirm_password'] = "Passwords do not match."

        if not errors:
            lec_update_user_password(session['user_id'], new_password)
            flash("Password updated successfully!", "success_redirect")
            return redirect(url_for('lec_profile'))

    return render_template('lec_reset_password.html', errors=errors)

# --- UTILITIES ---
def lec_get_user_from_session(user_id):
    return mongo.db.users.find_one({"user_id": user_id})

def lec_check_password(stored_hashed, entered_plain):
    return check_password_hash(stored_hashed, entered_plain)

def lec_update_user_password(user_id, new_password):
    hashed_pw = generate_password_hash(new_password)
    mongo.db.users.update_one({'user_id': user_id}, {'$set': {'password': hashed_pw}})

@app.route('/lec/register', methods=['GET', 'POST'])
def lec_register():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    errors = {}

    if request.method == 'POST':
        user_id = request.form['user_id']
        name = request.form['name']
        email = request.form['email']
        phone = request.form['phone']
        gender = request.form['gender']
        role = request.form['role']
        password = request.form['password']
        confirm_password = request.form['confirm_password']

        # Check if user ID already exists
        existing_user = mongo.db.users.find_one({'user_id': user_id})
        if existing_user:
            errors['user_id'] = "User ID already exists."

        # STANDARDIZED EMAIL VALIDATION
        if not re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email):
            errors['email'] = "Please enter a valid email address."

        # Check if email already exists
        existing_user = mongo.db.users.find_one({'email': email})
        if existing_user:
            errors['email'] = "Email already exists."

        # Password match
        if password != confirm_password:
            errors['confirm_password'] = "Passwords do not match."

        # PASSWORD VALIDATION
        if not re.match(r'^(?=.*[a-z])(?=.*[A-Z])(?=.*\d)(?=.*[^A-Za-z0-9])[A-Za-z\d\W]{8,}$', password):
            errors['password'] = "Password must be at least 8 characters with uppercase, lowercase, number, and special character."

        # If errors, re-render with error messages
        if errors:
            return render_template('lec_register.html', errors=errors)

        # Save new user
        hashed_pw = generate_password_hash(password)
        new_user = {
            'user_id': user_id,
            'name': name,
            'email': email,
            'phone': phone,
            'gender': gender,
            'role': role,
            'password': hashed_pw,
            'photo_url': "/static/img/default-profile.jpg",
            'created_at': datetime.now()
        }

        mongo.db.users.insert_one(new_user)
        flash("New user registered successfully!", "success")
        return redirect(url_for('lec_register'))

    return render_template('lec_register.html', errors={})

# ============================
# Student List 
# ============================
@app.route('/lec/student-list')
def student_list():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    gender_filter = request.args.get('gender')
    query = {'role': 'student'}
    
    if gender_filter:
        query['gender'] = gender_filter

    students = mongo.db.users.find(query)
    return render_template('student_list.html', students=students, selected_gender=gender_filter)

@app.route('/lec/student/edit/<student_id>', methods=['GET', 'POST'])
def edit_student(student_id):
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    student = mongo.db.users.find_one({'_id': ObjectId(student_id)})
    
    if not student:
        flash("Student not found.", "danger")
        return redirect(url_for('student_list'))

    if request.method == 'POST':
        updated_data = {
            'user_id': request.form['user_id'],
            'name': request.form['name'],
            'email': request.form['email'],
            'phone': request.form['phone'],
            'gender': request.form['gender']
        }
        
        errors = {}
        
        # Check if user ID already exists (excluding current student)
        existing_user = mongo.db.users.find_one({
            'user_id': request.form['user_id'],
            '_id': {'$ne': ObjectId(student_id)}
        }) 
        if existing_user:
            errors['user_id'] = "User ID already exists."

        # Check if email already exists (excluding current student)
        existing_user = mongo.db.users.find_one({
            'email': request.form['email'],
            '_id': {'$ne': ObjectId(student_id)}
        })
        if existing_user:
            errors['email'] = "Email already exists."
        
        # If there are errors, return to form with errors
        if errors:
            return render_template('edit_student.html', student=student, errors=errors)

        
        # If no errors, proceed with update
        mongo.db.users.update_one({'_id': ObjectId(student_id)}, {'$set': updated_data})
        flash("Student updated successfully!", "success")
        return redirect(url_for('student_list'))

    return render_template('edit_student.html', student=student, errors={})

@app.route('/lec/student/delete/<student_id>', methods=['DELETE'])
def delete_student(student_id):
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify(success=False, message="Unauthorized access."), 403

    try:
        student = mongo.db.users.find_one({'_id': ObjectId(student_id), 'role': 'student'})
        if not student:
            return jsonify(success=False, message="Student not found."), 404

        user_id = student.get('user_id')

        # Check for linked records
        user_id = student.get('user_id')

        # Check if student is referenced anywhere else
        linked_courses = mongo.db.courses.find_one({'students': user_id})
        linked_assessments = mongo.db.assessments.find_one({'students': user_id})
        linked_submissions = mongo.db.submissions.find_one({'student_id': user_id})
        linked_violations = mongo.db.exam_violations.find_one({'student_id': user_id})
        linked_evidence = mongo.db.exam_evidence.find_one({'student_id': user_id})

        if any([linked_courses, linked_assessments, linked_submissions, linked_violations, linked_evidence]):
            return jsonify(success=False, message="Cannot delete. Student is linked to other records."), 400

        # Safe to delete
        mongo.db.users.delete_one({'_id': ObjectId(student_id)})
        return jsonify(success=True, message="Student deleted successfully.")
    
    except Exception as e:
        print(f"Error deleting student: {e}")
        return jsonify(success=False, message="Internal server error."), 500

# ============================
# Export Student List to Excel
# ============================
@app.route('/lec/student-list/export')
def export_student_list():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    try:
        # Get filter parameters (same as in student_list route)
        gender_filter = request.args.get('gender')
        query = {'role': 'student'}
        
        if gender_filter:
            query['gender'] = gender_filter

        # Fetch students data
        students = list(mongo.db.users.find(query))
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Student List"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['No.', 'User ID', 'Full Name', 'Email', 'Phone', 'Gender', 'Registration Date']
        
        # Add headers to worksheet
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Add student data
        for row, student in enumerate(students, 2):
            # No.
            ws.cell(row=row, column=1, value=row-1).border = border
            
            # User ID
            ws.cell(row=row, column=2, value=student.get('user_id', '')).border = border
            
            # Full Name
            ws.cell(row=row, column=3, value=student.get('name', '')).border = border
            
            # Email
            ws.cell(row=row, column=4, value=student.get('email', '')).border = border
            
            # Phone
            ws.cell(row=row, column=5, value=student.get('phone', '')).border = border
            
            # Gender
            ws.cell(row=row, column=6, value=student.get('gender', '')).border = border
            
            # Registration Date
            created_at = student.get('created_at')
            if created_at:
                # Format date as DD/MM/YYYY
                reg_date = created_at.strftime('%d/%m/%Y')
            else:
                reg_date = 'N/A'
            ws.cell(row=row, column=7, value=reg_date).border = border
        
        # Auto-adjust column widths
        column_widths = [8, 15, 25, 30, 15, 10, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # Add summary information
        summary_row = len(students) + 3
        ws.cell(row=summary_row, column=1, value="Total Students:").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=len(students)).font = Font(bold=True)
        
        if gender_filter:
            ws.cell(row=summary_row + 1, column=1, value=f"Filter Applied:").font = Font(bold=True)
            ws.cell(row=summary_row + 1, column=2, value=f"Gender = {gender_filter}").font = Font(bold=True)
        
        # Export timestamp
        ws.cell(row=summary_row + 2, column=1, value="Exported on:").font = Font(bold=True)
        ws.cell(row=summary_row + 2, column=2, value=datetime.now().strftime('%d/%m/%Y %H:%M:%S')).font = Font(bold=True)
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create filename
        filter_suffix = f"_{gender_filter}" if gender_filter else ""
        filename = f"Student_List{filter_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        flash(f"Student list exported successfully! ({len(students)} students)", "success")
        return response
        
    except Exception as e:
        print(f"❌ Error exporting student list: {e}")
        flash("Failed to export student list. Please try again.", "danger")
        return redirect(url_for('student_list'))
    
# ============================
# Lecturer List 
# ============================
@app.route('/lec/lecturer-list')
def lec_list():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    gender_filter = request.args.get('gender')
    query = {'role': 'lecturer'}
    
    if gender_filter:
        query['gender'] = gender_filter

    lecturers = mongo.db.users.find(query)
    return render_template('lec_list.html', lecturers= lecturers, selected_gender=gender_filter)


@app.route('/lec/lecturer/edit/<lecturer_id>', methods=['GET', 'POST'])
def edit_lecturer(lecturer_id):
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    lecturer = mongo.db.users.find_one({'_id': ObjectId(lecturer_id)})
    
    if not lecturer:
        flash("Lecturer not found.", "danger")
        return redirect(url_for('lec_list'))

    if request.method == 'POST':
        updated_data = {
            'user_id': request.form['user_id'],
            'name': request.form['name'],
            'email': request.form['email'],
            'phone': request.form['phone'],
            'gender': request.form['gender']
        }

        errors = {}
        
        # Check if user ID already exists (excluding current lecturer)
        existing_user = mongo.db.users.find_one({
            'user_id': request.form['user_id'],
            '_id': {'$ne': ObjectId(lecturer_id)}
        }) 
        if existing_user:
            errors['user_id'] = "User ID already exists."

        # Check if email already exists (excluding current lecturer)
        existing_user = mongo.db.users.find_one({
            'email': request.form['email'],
            '_id': {'$ne': ObjectId(lecturer_id)}
        })
        if existing_user:
            errors['email'] = "Email already exists."
        
        # If there are errors, return to form with errors
        if errors:
            return render_template('edit_lecturer.html', lecturer=lecturer, errors=errors)  # Fixed this line
        
        # If no errors, proceed with update
        mongo.db.users.update_one({'_id': ObjectId(lecturer_id)}, {'$set': updated_data})
        flash("Lecturer updated successfully!", "success")
        return redirect(url_for('lec_list'))

    return render_template('edit_lecturer.html', lecturer=lecturer, errors={})

@app.route('/lec/lecturer/delete/<lecturer_id>', methods=['DELETE'])
def delete_lecturer(lecturer_id):
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify(success=False, message="Unauthorized access"), 403

    try:
        lecturer = mongo.db.users.find_one({'_id': ObjectId(lecturer_id), 'role': 'lecturer'})
        if not lecturer:
            return jsonify(success=False, message="Lecturer not found"), 404

        # Check if lecturer is assigned to any collection
        user_id = lecturer.get('user_id')
        has_announcements = mongo.db.announcements.find_one({
            '$or': [{'created_by': user_id}, {'updated_by': user_id}]
        })

        if has_announcements:
            return jsonify(success=False, message="Cannot delete. Lecturer is linked to other records."), 400

        # Proceed with deletion
        mongo.db.users.delete_one({'_id': ObjectId(lecturer_id)})
        return jsonify(success=True, message="Lecturer deleted successfully.")
    
    except Exception as e:
        print("Delete error:", e)
        return jsonify(success=False, message="Failed to delete lecturer. Please try again."), 500

# ============================
# Export Lecturer List to Excel
# ============================
@app.route('/lec/lecturer-list/export')
def export_lecturer_list():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    try:
        # Get filter (if any)
        gender_filter = request.args.get('gender')
        query = {'role': 'lecturer'}
        if gender_filter:
            query['gender'] = gender_filter

        lecturers = list(mongo.db.users.find(query))

        # Excel setup
        wb = Workbook()
        ws = wb.active
        ws.title = "Lecturer List"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = ['No.', 'User ID', 'Full Name', 'Email', 'Phone', 'Gender', 'Registration Date']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        # Fill rows
        for i, lec in enumerate(lecturers, 2):
            ws.cell(row=i, column=1, value=i - 1).border = border
            ws.cell(row=i, column=2, value=lec.get('user_id', '')).border = border
            ws.cell(row=i, column=3, value=lec.get('name', '')).border = border
            ws.cell(row=i, column=4, value=lec.get('email', '')).border = border
            ws.cell(row=i, column=5, value=lec.get('phone', '')).border = border
            ws.cell(row=i, column=6, value=lec.get('gender', '')).border = border

            created = lec.get('created_at')
            reg_date = created.strftime('%d/%m/%Y') if created else 'N/A'
            ws.cell(row=i, column=7, value=reg_date).border = border

        # Adjust width
        widths = [8, 15, 25, 30, 15, 10, 18]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        # Summary
        summary_row = len(lecturers) + 3
        ws.cell(row=summary_row, column=1, value="Total Lecturers:").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=len(lecturers)).font = Font(bold=True)

        if gender_filter:
            ws.cell(row=summary_row + 1, column=1, value="Filter Applied:").font = Font(bold=True)
            ws.cell(row=summary_row + 1, column=2, value=f"Gender = {gender_filter}").font = Font(bold=True)

        ws.cell(row=summary_row + 2, column=1, value="Exported on:").font = Font(bold=True)
        ws.cell(row=summary_row + 2, column=2, value=datetime.now().strftime('%d/%m/%Y %H:%M:%S')).font = Font(bold=True)

        # Output
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Lecturer_List{('_' + gender_filter) if gender_filter else ''}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'

        flash(f"Lecturer list exported successfully! ({len(lecturers)} lecturers)", "success")
        return response

    except Exception as e:
        print(f"Error exporting lecturer list: {e}")
        flash("Failed to export lecturer list. Please try again.", "danger")
        return redirect(url_for('lec_list'))


# ============================
# Course Management
# ============================
@app.route('/course/manage')
def course_manage():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    courses = list(mongo.db.courses.find())  # Fetch all courses
    return render_template('course_manage.html', courses=courses)


@app.route('/lecturer/course/create', methods=['GET', 'POST'])
def course_create():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    errors = {}

    if request.method == 'POST':
        course_code = request.form['course_code'].strip()
        course_name = request.form['course_name'].strip()
        students = request.form.getlist('students')

        existing = mongo.db.courses.find_one({'course_code': course_code})
        if existing:
            errors['course_code'] = "Course code already exists."
            all_students = mongo.db.users.find({'role': 'student'})
            return render_template('course_create.html', errors=errors, all_students=all_students)

        mongo.db.courses.insert_one({
            'course_code': course_code,
            'course_name': course_name,
            'students': students,
            'created at': datetime.now()
        })

        flash("Course created successfully!", "success")
        return redirect(url_for('course_manage'))

    all_students = mongo.db.users.find({'role': 'student'})
    return render_template('course_create.html', errors=errors, all_students=all_students)

@app.route('/lecturer/course/edit/<course_code>', methods=['GET', 'POST'])
def course_edit(course_code):
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    course = mongo.db.courses.find_one({'course_code': course_code})
    if not course:
        flash("Course not found.", "danger")
        return redirect(url_for('course_manage'))

    errors = {}

    if request.method == 'POST':
        new_code = request.form['course_code']
        new_name = request.form['course_name']
        students = request.form.getlist('students')

        if new_code != course_code and mongo.db.courses.find_one({'course_code': new_code}):
            errors['course_code'] = "Course code already exists."
            all_students = mongo.db.users.find({'role': 'student'})
            course['course_code'] = new_code
            course['course_name'] = new_name
            course['students'] = students
            return render_template('course_edit.html', course=course, all_students=all_students, errors=errors)

        mongo.db.courses.update_one(
            {'course_code': course_code},
            {'$set': {
                'course_code': new_code,
                'course_name': new_name,
                'students': students
            }}
        )
        flash("Course updated successfully!", "success")
        return redirect(url_for('course_manage'))

    all_students = mongo.db.users.find({'role': 'student'})
    return render_template('course_edit.html', course=course, all_students=all_students, errors={})

@app.route('/lecturer/course/delete/<course_code>', methods=['DELETE'])
def course_delete(course_code):
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'message': 'Unauthorized access.'}), 403

    # Check if course exists in assessments
    in_assessment = mongo.db.assessments.find_one({'course_codes': course_code})
    if in_assessment:
        return jsonify({'message': 'Cannot delete. This course is linked to assessments.'}), 400

    # Proceed to delete
    result = mongo.db.courses.delete_one({'course_code': course_code})

    if result.deleted_count == 1:
        return jsonify({'message': 'Course deleted successfully!'})
    else:
        return jsonify({'message': 'Course not found or already deleted.'}), 404

# ============================
# Assessment Management
# ============================
@app.route('/lecturer/assessment/manage')
def assessment_manage():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    assessments = mongo.db.assessments.find()
    return render_template('assessment_manage.html', assessments=assessments)

@app.route('/lecturer/assessment/create', methods=['GET', 'POST'])
def assessment_create():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    errors = {}
    if request.method == 'POST':
        code = request.form['assessment_code']
        title = request.form['title']
        course_codes = request.form.getlist('course_codes')
        file = request.files['question_pdf']

        # Check if code exists
        if mongo.db.assessments.find_one({'assessment_code': code}):
            errors['assessment_code'] = "Assessment code already exists."
            courses = mongo.db.courses.find()
            return render_template('assessment_create.html', courses=courses, errors=errors)

        # Save PDF
        if file and file.filename.endswith('.pdf'):
            filename = secure_filename(code + '_' + file.filename)
            path = os.path.join('static/uploads/assessments', filename)
            file.save(path)
        else:
            flash("Only PDF files are allowed.", "danger")
            return redirect(url_for('assessment_create'))

        # Auto-enroll students
        student_ids = []
        for course_code in course_codes:
            course = mongo.db.courses.find_one({'course_code': course_code})
            if course and 'students' in course:
                student_ids.extend(course['students'])
        student_ids = list(set(student_ids))  # remove duplicates

        # Insert into DB
        mongo.db.assessments.insert_one({
            'assessment_code': code,
            'title': title,
            'course_codes': course_codes,
            'students': student_ids,
            'pdf_filename': filename,
            'enrollment_details': {
                'self_enrolled': [],  # Empty initially
                'non_self_enrolled': student_ids  # All auto-enrolled from courses
            }
        })

        flash("Assessment created successfully!", "success")
        return redirect(url_for('assessment_manage'))

    courses = mongo.db.courses.find()
    return render_template('assessment_create.html', courses=courses, errors={})

@app.route('/lecturer/assessment/edit/<assessment_code>', methods=['GET', 'POST'])
def assessment_edit(assessment_code):
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    assessment = mongo.db.assessments.find_one({'assessment_code': assessment_code})
    if not assessment:
        flash("Assessment not found.", "danger")
        return redirect(url_for('assessment_manage'))

    courses = list(mongo.db.courses.find({}, {'course_code': 1, 'course_name': 1, 'students': 1}))
    all_students = list(mongo.db.users.find({'role': 'student'}))
    errors = {}

    if request.method == 'POST':
        new_code = request.form['assessment_code']
        title = request.form['title']
        selected_courses = request.form.getlist('course_codes')
        selected_students = request.form.getlist('students')
        file = request.files.get('question_pdf')

        # Check if new assessment code is already taken by another record
        if new_code != assessment_code and mongo.db.assessments.find_one({'assessment_code': new_code}):
            errors['assessment_code'] = "Assessment code already exists."
            return render_template('assessment_edit.html', assessment=assessment, courses=courses, all_students=all_students, errors=errors)

        # Preserve self-enrolled students and track lecturer changes
        current_self_enrolled = []
        if 'enrollment_details' in assessment:
            current_self_enrolled = assessment['enrollment_details'].get('self_enrolled', [])

        # All students who are not self-enrolled are considered lecturer-managed
        lecturer_managed_students = [s for s in selected_students if s not in current_self_enrolled]

        update_data = {
            'assessment_code': new_code,
            'title': title,
            'course_codes': selected_courses,
            'students': selected_students,
            'enrollment_details': {
                'self_enrolled': current_self_enrolled,  # Preserve existing self-enrolled
                'non_self_enrolled': lecturer_managed_students  # All lecturer-managed students
            }
        }

        # Replace PDF if uploaded
        if file and file.filename.endswith('.pdf'):
            filename = secure_filename(new_code + '_' + file.filename)
            file_path = os.path.join('static/uploads/assessments', filename)
            file.save(file_path)
            update_data['pdf_filename'] = filename

        # Perform the update
        mongo.db.assessments.update_one(
            {'assessment_code': assessment_code},
            {'$set': update_data}
        )

        flash("Assessment updated successfully!", "success")
        return redirect(url_for('assessment_manage'))

    return render_template('assessment_edit.html', assessment=assessment, courses=courses, all_students=all_students, errors={})

@app.route('/lecturer/assessment/delete/<assessment_code>', methods=['DELETE'])
def delete_assessment(assessment_code):
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'message': 'Unauthorized access.'}), 403

    # Check if assessment exists in submissions
    in_submissions = mongo.db.submissions.find_one({'assessment_code': assessment_code})
    in_resubmission_requests = mongo.db.resubmission_requests.find_one({'assessment_code': assessment_code})
    if in_submissions:
        return jsonify({'message': 'Cannot delete. This assessment is linked to submissions.'}), 400
    if in_resubmission_requests:
        return jsonify({'message': 'Cannot delete. This assessment is linked to resubmission requests.'}), 400

    result = mongo.db.assessments.delete_one({'assessment_code': assessment_code})

    if result.deleted_count == 1:
        return jsonify({'message': 'Assessment deleted successfully!'})
    else:
        return jsonify({'message': 'Assessment not found.'}), 404
    
# ============================
# Timetable Management
# ============================
@app.route('/lecturer/exam/timetable')
def exam_timetable():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    exam_timetables = list(mongo.db.exam_timetables.find())
    assessments = list(mongo.db.assessments.find())
    return render_template('exam_timetable.html', exam_timetables=exam_timetables, assessments=assessments)

@app.route('/lecturer/exam/timetable/create', methods=['GET', 'POST'])
def exam_timetable_create():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))
    
    errors = {}

    if request.method == 'POST':
        assessment_code = request.form['assessment_code']
        exam_date = request.form['exam_date']
        start_time = request.form['start_time']
        end_time = request.form['end_time']

        # Check if timetable already exists for this assessment
        if mongo.db.exam_timetables.find_one({'assessment_code': assessment_code}):
            errors['assessment_code'] = "Timetable for this assessment already exists"
            assessments = mongo.db.assessments.find()
            return render_template('exam_timetable_create.html', assessments=assessments, errors=errors)

        # Optional: Validate datetime format
        try:
            datetime.strptime(exam_date, '%Y-%m-%d')
        except ValueError:
            flash("Invalid exam date format.", "danger")
            return redirect(url_for('exam_timetable_create'))
        
        # Convert and validate time
        try:
            start_dt = datetime.strptime(f"{exam_date} {start_time}", '%Y-%m-%d %H:%M')
            end_dt = datetime.strptime(f"{exam_date} {end_time}", '%Y-%m-%d %H:%M')

            if end_dt <= start_dt:
                errors['end_time'] = "End time must be later than start time."
            if start_dt.hour < 8 or end_dt.hour > 19 or end_dt.hour < 8:
                errors['end_time'] = "Exam time must be between 08:00 and 19:00."
        except ValueError:
            errors['end_time'] = "Invalid date or time format."

        # ======================
        # Check for Time Clash for Students
        # ======================
        target_assessment = mongo.db.assessments.find_one({'assessment_code': assessment_code})
        conflict_students = []

        if target_assessment:
            for sid in target_assessment.get('students', []):
                # Find all assessments the student is enrolled in, except current one
                other_assessments = mongo.db.assessments.find({
                    'students': sid,
                    'assessment_code': {'$ne': assessment_code}
                })

                for other in other_assessments:
                    other_tt = mongo.db.exam_timetables.find_one({'assessment_code': other['assessment_code']})
                    if other_tt and other_tt['exam_date'] == exam_date:
                        other_start = datetime.strptime(f"{exam_date} {other_tt['start_time']}", '%Y-%m-%d %H:%M')
                        other_end = datetime.strptime(f"{exam_date} {other_tt['end_time']}", '%Y-%m-%d %H:%M')
                        if start_dt < other_end and end_dt > other_start:
                            conflict_students.append(sid)
                            break

        if conflict_students:
            errors['assessment_code'] = f"Time clash detected with {len(conflict_students)} student(s)."

        if errors:
            used_codes = mongo.db.exam_timetables.distinct('assessment_code')
            assessments = mongo.db.assessments.find({'assessment_code': {'$nin': used_codes}})
            return render_template('exam_timetable_create.html', assessments=assessments, errors=errors)


        mongo.db.exam_timetables.insert_one({
            'assessment_code': assessment_code,
            'exam_date': exam_date,
            'start_time': start_time,
            'end_time': end_time
        })

        flash("Exam timetable created successfully!", "success")
        return redirect(url_for('exam_timetable'))

    # Get assessment codes that already have a timetable
    used_codes = mongo.db.exam_timetables.distinct('assessment_code')
    assessments = mongo.db.assessments.find({'assessment_code': {'$nin': used_codes}})

    return render_template('exam_timetable_create.html', assessments=assessments, errors={})


@app.route('/lecturer/exam/timetable/edit/<assessment_code>', methods=['GET', 'POST'])
def exam_timetable_edit(assessment_code):
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    timetable = mongo.db.exam_timetables.find_one({'assessment_code': assessment_code})
    if not timetable:
        flash("Timetable not found.", "danger")
        return redirect(url_for('exam_timetable'))

    assessments = list(mongo.db.assessments.find())
    errors = {}
    if request.method == 'POST':
        selected_assessment = request.form['assessment_code']
        exam_date = request.form['exam_date']
        start_time = request.form['start_time']
        end_time = request.form['end_time']

        errors = {}

        # ========================
        # Basic Time Validation
        # ========================
        try:
            start_dt = datetime.strptime(f"{exam_date} {start_time}", '%Y-%m-%d %H:%M')
            end_dt = datetime.strptime(f"{exam_date} {end_time}", '%Y-%m-%d %H:%M')

            if end_dt <= start_dt:
                errors['end_time'] = "End time must be after start time."
            elif start_dt.hour < 8 or end_dt.hour > 19:
                errors['end_time'] = "Exam time must be between 08:00 and 19:00."

        except ValueError:
            errors['end_time'] = "Invalid date or time format."

        # ======================================
        # Check for Duplicate Assessment Code
        # ======================================
        if selected_assessment != assessment_code:
            if mongo.db.exam_timetables.find_one({'assessment_code': selected_assessment}):
                errors['assessment_code'] = "Timetable for this assessment already exists."

        # ======================================================
        # Check Student Clash if Time or Assessment Changed
        # ======================================================
        if not errors and (
            selected_assessment != assessment_code or
            exam_date != timetable['exam_date'] or
            start_time != timetable['start_time'] or
            end_time != timetable['end_time']
        ):
            target_assessment = mongo.db.assessments.find_one({'assessment_code': selected_assessment})
            conflict_students = []

            if target_assessment:
                for sid in target_assessment.get('students', []):
                    other_assessments = mongo.db.assessments.find({
                        'students': sid,
                        'assessment_code': {'$ne': selected_assessment}
                    })

                    for other in other_assessments:
                        other_tt = mongo.db.exam_timetables.find_one({'assessment_code': other['assessment_code']})
                        if other_tt and other_tt['exam_date'] == exam_date:
                            other_start = datetime.strptime(f"{exam_date} {other_tt['start_time']}", '%Y-%m-%d %H:%M')
                            other_end = datetime.strptime(f"{exam_date} {other_tt['end_time']}", '%Y-%m-%d %H:%M')
                            if start_dt < other_end and end_dt > other_start:
                                conflict_students.append(sid)
                                break

            if conflict_students:
                errors['assessment_code'] = f"Time clash detected with {len(conflict_students)} student(s)."

        # ======================
        # If No Errors, Update
        # ======================
        if not errors:
            mongo.db.exam_timetables.update_one(
                {'assessment_code': assessment_code},
                {'$set': {
                    'assessment_code': selected_assessment,
                    'exam_date': exam_date,
                    'start_time': start_time,
                    'end_time': end_time
                }}
            )
            flash("Exam timetable updated successfully!", "success")
            return redirect(url_for('exam_timetable'))

        # Update form fields with entered values
        timetable.update({
            'assessment_code': selected_assessment,
            'exam_date': exam_date,
            'start_time': start_time,
            'end_time': end_time
        })

    return render_template(
        'exam_timetable_edit.html',
        timetable=timetable,
        assessments=assessments,
        errors=errors
    )

@app.route('/lecturer/exam/timetable/delete/<assessment_code>', methods=['DELETE'])
def exam_timetable_delete(assessment_code):
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'message': 'Unauthorized access'}), 403

    result = mongo.db.exam_timetables.delete_one({'assessment_code': assessment_code})

    if result.deleted_count:
        return jsonify({'message': 'Exam timetable deleted successfully!'})
    return jsonify({'message': 'Timetable not found'}), 404

# ============================
# Exam Submission Management
# ============================
@app.route('/lecturer/exam/submissions')
def lecturer_submission_overview():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    assessments = list(mongo.db.assessments.find({}))
    return render_template('lec_submission_overview.html', assessments=assessments)

@app.route('/lecturer/exam/submissions/<assessment_code>')
def lecturer_view_submissions(assessment_code):
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    assessment = mongo.db.assessments.find_one({'assessment_code': assessment_code})
    if not assessment:
        flash("Assessment not found.", "danger")
        return redirect(url_for('lecturer_submission_overview'))

    submissions = list(mongo.db.submissions.find({'assessment_code': assessment_code}))
    student_ids = [s['student_id'] for s in submissions]
    students = mongo.db.users.find({'user_id': {'$in': student_ids}})
    student_map = {u['user_id']: u for u in students}

    # Build course mapping from student ID to course_name
    course_codes = assessment.get('course_codes', [])
    relevant_courses = mongo.db.courses.find({'course_code': {'$in': course_codes}})
    student_course_map = {}

    for course in relevant_courses:
        for sid in course.get('students', []):
            if sid not in student_course_map:
                student_course_map[sid] = course['course_name']

    resub_requests = mongo.db.resubmission_requests.find({'assessment_code': assessment_code})
    resub_map = {r['submission_id']: r['status'] for r in resub_requests}

    for s in submissions:
        stu = student_map.get(s['student_id'], {})
        s['student_name'] = stu.get('name', 'N/A')
        s['student_course'] = student_course_map.get(s['student_id'], 'Unknown')
        s['status'] = resub_map.get(str(s['_id']))

        if isinstance(s.get('submitted_at'), datetime):
            s['submitted_at'] = s['submitted_at'].astimezone(timezone('Asia/Kuala_Lumpur'))

    return render_template("lec_exam_submissions.html",
                           assessment=assessment,
                           submissions=submissions)

@app.route('/lecturer/exam/resubmission/<submission_id>', methods=['GET', 'POST'])
def lecturer_view_resubmission(submission_id):
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    # Fetch the resubmission request
    resub = mongo.db.resubmission_requests.find_one({'submission_id': submission_id})
    if not resub:
        flash("Resubmission request not found.", "danger")
        return redirect(url_for('lecturer_submission_overview'))

    if request.method == 'POST':
        action = request.form.get('action')
        if action in ['Approved', 'Rejected']:
            # Update status
            mongo.db.resubmission_requests.update_one(
                {'_id': resub['_id']},
                {'$set': {'status': action}}
            )

            if action == 'Approved':
                # Replace the filename in submissions
                mongo.db.submissions.update_one(
                    {'_id': ObjectId(resub['submission_id'])},
                    {'$set': {'filename': resub['new_filename']}}
                )

                # Delete the old file
                old_path = os.path.join(app.root_path, 'static', 'uploads', 'submissions', resub['old_filename'])
                if os.path.exists(old_path):
                    os.remove(old_path)

            elif action == 'Rejected':
                pass # No file deletion

            flash(f"Resubmission {action.lower()} successfully.", "success")
        return redirect(url_for('lecturer_view_submissions', assessment_code=resub['assessment_code']))

    # Get student details
    student = mongo.db.users.find_one({'user_id': resub['student_id']})
    return render_template("lec_resubmission_detail.html",
                           resub=resub, 
                           student=student) 

# ============================
# Similarity Detection (***Got both Cosine Similarity and Sequence Matching but only using Sequence Matching)
# ============================
@app.route('/lecturer/exam/resubmission/<submission_id>/similarity')
def check_similarity(submission_id):
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    resub = mongo.db.resubmission_requests.find_one({'submission_id': submission_id})
    if not resub:
        flash("Resubmission not found.", "danger")
        return redirect(url_for('lecturer_submission_overview'))

    old_path = os.path.join(app.root_path, 'static', 'uploads', 'submissions', resub['old_filename'])
    new_path = os.path.join(app.root_path, 'static', 'uploads', 'submissions', resub['new_filename'])

    text1 = extract_text(old_path)
    text2 = extract_text(new_path)
    try:
        similarity = calculate_cosine_similarity(text1, text2)
    except Exception as e:
        print("Cosine similarity failed, falling back:", e)
        similarity = calculate_similarity(text1, text2)

    highlighted1, highlighted2 = highlight_matches(text1, text2)

    cosine = calculate_cosine_similarity(text1, text2)
    sequence = calculate_similarity(text1, text2)
    highlighted1, highlighted2 = highlight_matches(text1, text2)

    # Save to DB (update only similarity and timestamp)
    mongo.db.resubmission_requests.update_one(
        {'_id': resub['_id']},
        {'$set': {
            'similarity_score': sequence,
            'similarity_checked_at': datetime.now()
        }}
    )
    
    return render_template("similarity_report.html",
                                resub=resub,
                                similarity=sequence,
                                cosine_score=cosine,
                                original_text=highlighted1,
                                resubmitted_text=highlighted2
                            )

# ============================
# Plagiarism Detection
# ============================

# =======================
# API Token (***Change here if using other account from PlagiarismCheck.org)
# =======================
PLAGIARISM_API_TOKEN = "U-BRbaec9Pcf36p8JWVfYPrsGr2ce2Dl"
# WUURDwIYym3NXwROp0ZxsQG9QF4roldY (1 chance) -- peili.yee01@gmail.com
# TdyW28a9ghkXlvMVHZ5dq8gj7wO9gJ2n (1 chance) -- on94anime@gmail.com


# =======================
# Extract text from .docx
# =======================
def extract_text_from_docx(docx_path):
    try:
        doc = docx.Document(docx_path)
        text_content = "\n".join([para.text for para in doc.paragraphs]).strip()
        print(f"DEBUG: Extracted text length: {len(text_content)}")
        print(f"DEBUG: First 200 chars: {text_content[:200]}...")
        return text_content
    except Exception as e:
        print(f"ERROR: Failed to extract text from DOCX: {e}")
        return None

# =======================
# Submit file for plagiarism check
# =======================
def check_plagiarism(file_path):
    url = "https://plagiarismcheck.org/api/v1/text"
    headers = {'X-API-TOKEN': PLAGIARISM_API_TOKEN}
    print(f"DEBUG: Checking file at path: {file_path}")

    if not os.path.exists(file_path):
        print(f"ERROR: File not found: {file_path}")
        return None

    try:
        if file_path.lower().endswith('.docx'):
            print("DEBUG: File is a DOCX. Extracting...")
            text_content = extract_text_from_docx(file_path)
        elif file_path.lower().endswith('.txt'):
            with open(file_path, 'r', encoding='utf-8') as file:
                text_content = file.read().strip()
        else:
            print("ERROR: Unsupported file type.")
            return None

        if not text_content:
            print("ERROR: No text extracted.")
            return None

        # API expects 'text' and optionally 'language' parameters
        data = {
            'text': text_content,
            'language': 'en'
        }
        
        response = requests.post(url, headers=headers, data=data)
        print(f"DEBUG: Response status: {response.status_code}")
        print(f"DEBUG: Response content: {response.text}")

        # API returns 200 or 201 on success
        if response.status_code in (200, 201):
            response_json = response.json()
            # Check for success flag and data structure as per API docs
            if response_json.get('success') and 'data' in response_json:
                text_id = response_json['data']['text']['id']
                print(f"DEBUG: Text ID: {text_id}")
                return text_id
            else:
                print(f"ERROR: API returned success=false or missing data")
                print(f"ERROR: Response: {response_json}")
        else:
            print(f"ERROR: File upload failed. Status: {response.status_code}")
            print(f"ERROR: Response: {response.text}")
    except Exception as e:
        print(f"ERROR: Exception during plagiarism check: {e}")
        traceback.print_exc()
    return None

# =======================
# Check the status of the plagiarism check
# =======================
def check_plagiarism_status(text_id):
    url = f"https://plagiarismcheck.org/api/v1/text/{text_id}"
    headers = {'X-API-TOKEN': PLAGIARISM_API_TOKEN}
    
    try:
        response = requests.get(url, headers=headers)
        
        if response.status_code == 200:
            response_json = response.json()
            print(f"DEBUG: Full response data: {json.dumps(response_json, indent=2)}")
            
            # According to API docs, the state is in data.state
            state = response_json.get('data', {}).get('state')
            print(f"DEBUG: State of text_id {text_id}: {state}")
            return state
        else:
            print(f"ERROR: Status check failed with status {response.status_code}")
            print(f"ERROR: Response: {response.text}")
    except Exception as e:
        print(f"ERROR: Exception during status check: {e}")
        traceback.print_exc()
    return None

# =======================
# Fetch report from API
# =======================
def fetch_plagiarism_report(text_id):
    url = f"https://plagiarismcheck.org/api/v1/text/report/{text_id}"
    headers = {'X-API-TOKEN': PLAGIARISM_API_TOKEN}
    max_tries = 30

    for attempt in range(1, max_tries + 1):
        try:
            response = requests.get(url, headers=headers)
            print(f"DEBUG: Poll {attempt}/{max_tries} - Status {response.status_code}")
            
            if response.status_code == 200:
                response_json = response.json()
                print(f"DEBUG: Full API response: {json.dumps(response_json, indent=2)}")
                
                data = response_json.get('data', {})

                # Save 'data' to extracted_data.json for debugging
                with open('extracted_data.json', 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)
                print("DEBUG: 'data' saved to extracted_data.json")

                # Check if we have report_data (indicates processing is complete)
                report_data = data.get('report_data', {})
                
                if report_data and 'nodes' in report_data:
                    # Report is ready - extract the data
                    nodes = report_data.get('nodes', [])
                    report = data.get('report', {})
                    sources_list = report_data.get('sources', [])

                    # Extract plagiarism percentage from report.percent
                    percent_str = report.get('percent', '0.00')
                    try:
                        plagiarism_percent = float(percent_str)
                        # Ensure we maintain the same precision as the API
                        plagiarism_percent = round(plagiarism_percent, 2)
                    except (ValueError, TypeError):
                        plagiarism_percent = 0.0

                    # Calculate quotes percentage by text length (correct method)
                    quotes_data = report_data.get('quotes', [])
                    total_document_length = report_data.get('length', 1)  # Avoid division by zero
                    
                    # Sum up the length of all quotes
                    total_quote_length = sum(quote.get('length', 0) for quote in quotes_data)
                    
                    quotes_percentage = round((total_quote_length / total_document_length) * 100, 2) if total_document_length > 0 else 0.0

                    # Extract unique sources that were actually matched
                    # Get all unique source IDs from nodes
                    matched_source_ids = set()
                    for node in nodes:
                        if node.get('enabled', True) and node.get('sources', []):
                            matched_source_ids.update(node.get('sources', []))

                    # Build the source list using only matched sources
                    matched_sources = []
                    for source_id in matched_source_ids:
                        if source_id < len(sources_list):
                            source = sources_list[source_id]
                            # Only include sources with actual matches (plagiarism_percent > 0)
                            if source.get('plagiarism_percent', 0) > 0:
                                matched_sources.append({
                                    'source': source.get('source', ''),
                                    'content_type': source.get('content_type', ''),
                                    'plagiarism_percent': source.get('plagiarism_percent', 0),
                                    'plagiarism_length': source.get('plagiarism_length', 0),
                                    'link': source.get('link', {})
                                })

                    print(f"DEBUG: Extracted - Plagiarism: {plagiarism_percent}%, Quotes: {quotes_percentage}%, Sources count: {len(matched_sources)}")
                    
                    return {
                        'plagiarism_score': plagiarism_percent,
                        'quotes_score': quotes_percentage,
                        'sources': matched_sources,
                        'report_data': report_data  # Include the detailed report data
                    }
                else:
                    # Check if we're still processing by looking at text status
                    print("DEBUG: Report data not available yet, checking text status...")
                    status = check_plagiarism_status(text_id)
                    if status == 5:
                        print("DEBUG: Status is 5 but no report_data - possible API delay, retrying...")
                    elif status == 3:
                        print("DEBUG: Still processing...")
                    elif status == 4:
                        print("DEBUG: API reported error state")
                        break
                    else:
                        print(f"DEBUG: Unknown status: {status}")
            
            elif response.status_code == 404:
                print("ERROR: Text ID not found")
                break
            else:
                print(f"ERROR: Report fetch failed with status {response.status_code}")
                print(f"ERROR: Response: {response.text}")

            time.sleep(10)
        except Exception as e:
            print(f"ERROR: During report fetch: {e}")
            traceback.print_exc()
            break

    print("ERROR: Timed out or failed fetching report")
    return None

# =======================
# Check plagiarism
# =======================
@app.route('/lecturer/plagiarism/<submission_id>', methods=['POST'])
def check_plagiarism_request(submission_id):
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))
    
    try:
        # Fetch submission from the database
        submission = mongo.db.submissions.find_one({'_id': ObjectId(submission_id)})
        if not submission:
            flash("Submission not found.", "danger")
            return redirect(url_for('lecturer_submission_overview'))

        # Construct file path and ensure the file exists
        file_path = os.path.join(app.root_path, 'static', 'uploads', 'submissions', submission['filename'])
        if not os.path.exists(file_path):
            flash("File not found.", "danger")
            return redirect(url_for('lecturer_submission_overview'))

        # Check plagiarism for the file
        text_id = check_plagiarism(file_path)
        if not text_id:
            flash("Plagiarism check failed to submit.", "danger")
            return redirect(url_for('lecturer_submission_overview'))

        # Start fetching the plagiarism report
        flash("Plagiarism check started. Please wait...", "info")
        result = fetch_plagiarism_report(text_id)

        if result:
            plagiarism_score = result['plagiarism_score']
            quotes_score = result['quotes_score']
            sources = result['sources']
            report_data = result.get('report_data', {})  # Get report_data from result
            report_url = f"https://plagiarismcheck.org/profile/check-document-report/{text_id}"

            print(f"DEBUG: Final values - Plagiarism: {plagiarism_score}%, Quotes: {quotes_score}%, Sources: {len(sources)}")

            # Update MongoDB with plagiarism score, quotes score, sources, and report data
            update_data = {
                'plagiarism_score': plagiarism_score,
                'quotes_score': quotes_score,
                'sources': sources,
                'plagiarism_report_url': report_url,
                'plagiarism_text_id': text_id,
                'report_data': report_data  # Store the detailed report data
            }
            
            print(f"DEBUG: Updating MongoDB with: {update_data}")
            
            result_update = mongo.db.submissions.update_one(
                {'_id': ObjectId(submission_id)},
                {'$set': update_data}
            )
            
            if result_update.modified_count > 0:
                flash(f"Plagiarism score: {plagiarism_score}%, Quotes: {quotes_score}%", "success")
            else:
                flash("Plagiarism check completed but database update failed.", "warning")
            
            # Get assessment code for redirect
            assessment_code = submission.get('assessment_code')
            return redirect(url_for('lecturer_view_submissions', assessment_code=assessment_code))

        else:
            flash("Failed to fetch plagiarism result.", "danger")

    except Exception as e:
        print(f"ERROR: check_plagiarism_request: {e}")
        traceback.print_exc()
        flash("Internal error during plagiarism check.", "danger")

    return redirect(url_for('lecturer_view_submissions'))

# =======================
# View detailed plagiarism report
# =======================
@app.route('/lecturer/plagiarism/report/<submission_id>', methods=['GET'])
def lecturer_view_plagiarism_report(submission_id):
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    # Fetch submission data
    submission = mongo.db.submissions.find_one({'_id': ObjectId(submission_id)})
    if not submission:
        flash("Submission not found.", "danger")
        return redirect(url_for('lecturer_submission_overview'))

    # Fetch student info
    student = mongo.db.users.find_one({'user_id': submission['student_id']})
    if not student:
        flash("Student information not found.", "danger")
        return redirect(url_for('lecturer_submission_overview'))

    # Fetch course info
    course = mongo.db.courses.find_one({'students': submission['student_id']})
    course_name = course.get('course_name') if course else "N/A"

    # Add additional fields to student_info
    student['filename'] = submission.get('filename', '')
    student['course'] = course_name

    # Pass everything to template including detailed report data
    return render_template('lec_plagiarism_report.html',
                           plagiarism_report_url=submission.get('plagiarism_report_url'),
                           student_info=student,
                           plagiarism_score=submission.get('plagiarism_score', 0),
                           quotes_score=submission.get('quotes_score', 0),
                           sources=submission.get('sources', []),
                           report_data=submission.get('report_data', {}))

# ============================
# Exam Violation Management
# ============================
@app.route('/lecturer/exam/violations/assessments')
def lecturer_violation_assessments():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    # Only include reviewed violations
    pipeline = [
        {
            '$match': {
                'status': 'reviewed'
            }
        },
        {
            '$group': {
                '_id': '$assessment_code',
                'count': {'$sum': 1}
            }
        },
        {
            '$lookup': {
                'from': 'assessments',
                'localField': '_id',
                'foreignField': 'assessment_code',
                'as': 'assessment_info'
            }
        },
        {
            '$addFields': {
                'title': {
                    '$ifNull': [
                        {'$arrayElemAt': ['$assessment_info.title', 0]},
                        'No Title Available'
                    ]
                }
            }
        },
        {
            '$project': {
                '_id': 1,
                'count': 1,
                'title': 1
            }
        },
        {'$sort': {'_id': 1}}
    ]

    assessments = mongo.db.exam_violations.aggregate(pipeline)
    return render_template('lec_violation_assessments.html', assessments=list(assessments))

@app.route('/lecturer/exam/violations/<assessment_code>')
def lecturer_exam_violations(assessment_code):
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))
    assessment = mongo.db.assessments.find_one({'assessment_code': assessment_code})
    violations = list(mongo.db.exam_violations.find(
        {'assessment_code': assessment_code, 'status': 'reviewed'}
    ).sort("timestamp", -1))

    for v in violations:
        student = mongo.db.users.find_one({'user_id': v['student_id']})
        v['student_name'] = student['name'] if student else "Unknown"

        evidence_docs = list(mongo.db.exam_evidence.find({
            'violation_id': v['_id']
        }))

        merged_files = []
        for doc in evidence_docs:
            merged_files.extend(doc.get('files', []))
        v['evidence'] = {'files': merged_files} if merged_files else None

    return render_template("lec_exam_violation.html", 
                       violations=violations, 
                       assessment_code=assessment_code, 
                       assessment=assessment)

# ============================
# Exam Environment Management
# ============================
@app.route('/lecturer/exam/environment')
def global_exam_environment():
    """Enhanced to load head movement settings"""
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return redirect(url_for('login'))

    try:
        rule = mongo.db.exam_environment_rules.find_one()
        print(f"GET: Loaded rule for display: {rule}")
        return render_template('lec_exam_environment.html', rule=rule)
    except Exception as e:
        print(f"Error loading exam environment: {e}")
        flash("Error loading current settings.", "error")
        return render_template('lec_exam_environment.html', rule=None)

# POST route to save global rules
@app.route('/lecturer/exam/environment', methods=['POST'])
def save_global_environment_rules():
    """Enhanced to save head movement settings"""
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return redirect(url_for('login'))

    try:
        # Browser control settings (your existing logic)
        browser_data = {
            'block_f12': 'block_f12' in request.form,
            'block_ctrl_shift_i': 'block_ctrl_shift_i' in request.form,
            'block_ctrl_u': 'block_ctrl_u' in request.form,
            'block_right_click': 'block_right_click' in request.form,
            'block_copy': 'block_copy' in request.form,
            'block_paste': 'block_paste' in request.form,
        }
        
        # Head movement settings
        violation_duration = float(request.form.get('violation_duration', 2.0))  # seconds
        warning_count = int(request.form.get('warning_count', 3))
        max_yaw = int(request.form.get('max_yaw', 20))
        max_pitch = int(request.form.get('max_pitch', 10))
        
        # Validate input ranges
        if not (0.1 <= violation_duration <= 60):
            flash("Violation duration must be between 0.1 and 60 seconds.", "error")
            return redirect(url_for('global_exam_environment'))
            
        if not (0 <= warning_count <= 10):
            flash("Warning count must be between 0 and 10.", "error")
            return redirect(url_for('global_exam_environment'))
            
        if not (5 <= max_yaw <= 45):
            flash("Head turn angle must be between 5 and 45 degrees.", "error")
            return redirect(url_for('global_exam_environment'))
            
        if not (5 <= max_pitch <= 30):
            flash("Head tilt angle must be between 5 and 30 degrees.", "error")
            return redirect(url_for('global_exam_environment'))
        
        # Combine all settings
        complete_data = {
            **browser_data,  # Include all browser settings
            'head_movement_settings': {
                'violation_duration': int(violation_duration * 1000),  # Convert to milliseconds
                'warning_count': warning_count,
                'max_yaw': max_yaw,
                'max_pitch': max_pitch
            },
            'updated_at': datetime.now()
        }
        
        # Save to database
        result = mongo.db.exam_environment_rules.update_one(
            {},  # Match any document (global settings)
            {'$set': complete_data}, 
            upsert=True
        )
        
        if result.acknowledged:
            print(f"   POST: Saved browser + head movement settings")
            print(f"   Duration: {violation_duration}s, Warnings: {warning_count}")
            print(f"   Angles: ±{max_yaw}° yaw, ±{max_pitch}° pitch")
            flash("Exam environment settings saved successfully!", "success")
        else:
            flash("Failed to save settings. Please try again.", "error")
            
    except ValueError as e:
        print(f"Validation error: {e}")
        flash(f"Invalid input: {str(e)}", "error")
    except Exception as e:
        print(f"Error saving settings: {e}")
        flash("Error saving settings. Please try again.", "error")
    
    return redirect(url_for('global_exam_environment'))

# ============================
# Student Course
# ============================
@app.route('/student/course')
def student_course():
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    user = mongo.db.users.find_one({'user_id': session['user_id']})
    enrolled_courses = []

    if user and 'user_id' in user:
        enrolled_courses = list(mongo.db.courses.find({'students': user['user_id']}))

    return render_template('student_course.html', courses=enrolled_courses)

# ============================
# Student Assessment
# ============================
def time_ranges_overlap(start1, end1, start2, end2):
    return max(start1, start2) < min(end1, end2)

@app.route('/student/assessment', methods=['GET'])
def student_assessment():
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    user_id = session['user_id']

    # Get all assessments student already enrolled in
    enrolled = list(mongo.db.assessments.find({'students': user_id}))
    enrolled_codes = [a['assessment_code'] for a in enrolled]

    # Get the student's course codes
    student_courses = set()
    courses = mongo.db.courses.find({'students': user_id})
    for c in courses:
        codes = c['course_code'] if isinstance(c['course_code'], list) else [c['course_code']]
        student_courses.update(codes)

    # Get all assessments under the student's course(s), excluding enrolled ones
    available = list(mongo.db.assessments.find({
        'course_codes': {'$in': list(student_courses)},
        'students': {'$ne': user_id}
    }))

    return render_template(
        'student_assessment.html',
        enrolled=enrolled,
        available=available
    )

@app.route('/student/assessments/enroll/<code>', methods=['POST'])
def enroll_assessment(code):
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    user_id = session['user_id']
    selected_timetable = mongo.db.exam_timetables.find_one({'assessment_code': code})

    if selected_timetable:
        new_date = selected_timetable['exam_date']
        new_start = selected_timetable['start_time']
        new_end = selected_timetable['end_time']

        # Get all enrolled assessment timetables
        enrolled_assessments = mongo.db.assessments.find({'students': user_id})
        enrolled_codes = [e['assessment_code'] for e in enrolled_assessments]

        existing_timetables = mongo.db.exam_timetables.find({'assessment_code': {'$in': enrolled_codes}})
        for t in existing_timetables:
            if t['exam_date'] == new_date:
                if (new_start < t['end_time']) and (t['start_time'] < new_end):
                    flash(f"Cannot enroll. Time clash with {t['assessment_code']}.", "error")
                    return redirect(url_for('student_assessment'))

    # No conflict → enroll
    mongo.db.assessments.update_one(
        {'assessment_code': code},
        {
            '$addToSet': {
                'students': user_id,
                'enrollment_details.self_enrolled': user_id
            }
        }
    )
    flash("Successfully enrolled in assessment!", "success")
    return redirect(url_for('student_assessment'))

# ============================
# Student Timetable
# ============================
@app.route('/student/exam/timetable')
def student_exam_timetable():
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    student_id = session['user_id']

    # Step 1: Get all assessments where student is enrolled
    assessments = list(mongo.db.assessments.find({'students': {'$in': [student_id]}}))
    assessment_codes = [a['assessment_code'] for a in assessments]
    assessment_map = {a['assessment_code']: a['title'] for a in assessments}

    # Step 2: Get all exam timetables for those assessments
    exam_timetables = list(mongo.db.exam_timetables.find({
        'assessment_code': {'$in': assessment_codes}
    }))

    # Step 3: Process and enrich timetable data
    for e in exam_timetables:
        # Ensure exam_date is a datetime object for template filters
        if isinstance(e['exam_date'], str):
            e['exam_date'] = datetime.strptime(e['exam_date'], '%Y-%m-%d')
        
        # Add fields for rendering
        e['start_hour'] = int(e['start_time'].split(':')[0])
        e['end_hour'] = int(e['end_time'].split(':')[0]) + (1 if int(e['end_time'].split(':')[1]) > 0 else 0)
        
        # Add formatted date string for JavaScript
        e['date_string'] = e['exam_date'].strftime('%Y-%m-%d')

    # Sort by exam_date and start_time
    exam_timetables.sort(key=lambda x: (
        x['exam_date'],
        x['start_time']
    ))

    # Step 4: Create week range (Mon–Sun)
    today = datetime.today()
    start_week = today - timedelta(days=today.weekday())  # Monday
    days = [start_week + timedelta(days=i) for i in range(7)]

    return render_template(
        'student_exam_timetable.html',
        exam_timetables=exam_timetables,
        assessment_map=assessment_map,
        days=days
    )

@app.template_filter('datetimeformat')
def datetimeformat(value, format='%Y-%m-%d'):
    if isinstance(value, str):
        value = datetime.strptime(value, '%Y-%m-%d')
    return value.strftime(format)

# ============================
# Student Join Exam
# ============================
@app.route('/student/exam/join')
def join_exam():
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    student_id = session['user_id']

    # Find all assessments the student is enrolled in
    assessments = list(mongo.db.assessments.find({'students': {'$in': [student_id]}}))
    assessment_codes = [a['assessment_code'] for a in assessments]
    assessment_map = {a['assessment_code']: a['title'] for a in assessments}

    # Get relevant exam schedules
    exam_timetables = list(mongo.db.exam_timetables.find({'assessment_code': {'$in': assessment_codes}}))

    # Sort by date + time
    exam_timetables.sort(key=lambda x: (x['exam_date'], x['start_time']))

    # Get list of assessment_codes the student has already submitted
    submitted_codes = mongo.db.submissions.distinct('assessment_code', {'student_id': student_id})

    today = datetime.today().strftime('%Y-%m-%d')
    now = datetime.now().strftime('%H:%M')

    return render_template("join_exam.html",
                           exam_timetables=exam_timetables,
                           assessment_map=assessment_map,
                           submitted_codes=submitted_codes,  # <-- pass to template
                           today=today,
                           now=now)


# ============================
# Student Exam Interface - Updated with Head Detection
# ============================
@app.route('/student/exam/<assessment_code>/start')
def student_exam_interface(assessment_code):
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    student_id = session['user_id']

    # Fetch the assessment
    assessment = mongo.db.assessments.find_one({'assessment_code': assessment_code})
    if not assessment or student_id not in assessment.get('students', []):
        flash("You are not enrolled in this assessment.", "danger")
        return redirect(url_for('join_exam'))

    # Fetch timetable
    timetable = mongo.db.exam_timetables.find_one({'assessment_code': assessment_code})
    if not timetable:
        flash("No exam scheduled for this assessment.", "danger")
        return redirect(url_for('join_exam'))

    # Timing validation
    now = datetime.now()
    exam_date = datetime.strptime(timetable['exam_date'], '%Y-%m-%d')
    start_time = datetime.strptime(timetable['start_time'], '%H:%M')
    end_time = datetime.strptime(timetable['end_time'], '%H:%M')

    start_dt = datetime.combine(exam_date.date(), start_time.time())
    end_dt = datetime.combine(exam_date.date(), end_time.time())

    if not (start_dt <= now <= end_dt):
        flash("You can only access this exam during the scheduled time.", "danger")
        return redirect(url_for('join_exam'))

    return render_template(
        'student_exam_interface.html',
        assessment=assessment,
        timetable=timetable,
        start_time=start_dt.strftime('%Y-%m-%d %H:%M:%S'),
        end_time=end_dt.strftime('%Y-%m-%d %H:%M:%S'),
        # Add ISO format for JavaScript timeline analysis
        exam_start_iso=start_dt.isoformat(),  # For JavaScript detector
        exam_end_iso=end_dt.isoformat(),      # For JavaScript detector
    )

# ============================
# Student Submit Exam
# ============================
UPLOAD_FOLDER = os.path.join('static', 'uploads', 'submissions')
ALLOWED_EXTENSIONS = {'txt', 'doc', 'docx'} #Only allow student submit files with these extensions

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/student/exam/<assessment_code>/submit', methods=['POST'])
def submit_exam(assessment_code):
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    student_id = session['user_id']
    file = request.files.get('submission_files')
    auto_submit = request.form.get('auto_submit') == 'true'
    now_myt = datetime.now()
    filename = ""

    if file and allowed_file(file.filename):
        filename = secure_filename(
            f"{assessment_code}_{student_id}_{now_myt.strftime('%Y%m%d%H%M%S')}_{file.filename}"
        )
        save_path = os.path.join(UPLOAD_FOLDER, filename)
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        file.save(save_path)

    elif auto_submit:
        filename = "No file submitted"

    else:
        errors = {'submission_files': "Invalid file format. Please upload a txt or doc/docx file."}
        assessment = mongo.db.assessments.find_one({'assessment_code': assessment_code})
        timetable = mongo.db.exam_timetables.find_one({'assessment_code': assessment_code})
        exam_date = datetime.strptime(timetable['exam_date'], '%Y-%m-%d')
        start_time = datetime.strptime(timetable['start_time'], '%H:%M')
        end_time = datetime.strptime(timetable['end_time'], '%H:%M')
        start_dt = datetime.combine(exam_date.date(), start_time.time())
        end_dt = datetime.combine(exam_date.date(), end_time.time())

        return render_template(
            'student_exam_interface.html',
            assessment=assessment,
            timetable=timetable,
            start_time=start_dt.strftime('%Y-%m-%d %H:%M:%S'),
            end_time=end_dt.strftime('%Y-%m-%d %H:%M:%S'),
            exam_start_iso=start_dt.isoformat(),
            exam_end_iso=end_dt.isoformat(),
            errors=errors
        )

    mongo.db.submissions.insert_one({
        "assessment_code": assessment_code,
        "student_id": student_id,
        "filename": filename,
        "submitted_at": now_myt
    })

    flash("Your submission has been uploaded successfully.", "success")
    return redirect(url_for('join_exam'))

# ============================
# Screen Monitoring - Evidence Storage Setup
# ============================
# Evidence folder paths
EVIDENCE_FOLDER = os.path.join('static', 'evidence')
SCREENSHOTS_FOLDER = os.path.join(EVIDENCE_FOLDER, 'screenshots')
VIDEOS_FOLDER = os.path.join(EVIDENCE_FOLDER, 'videos')

# Create evidence folders if they don't exist
os.makedirs(SCREENSHOTS_FOLDER, exist_ok=True)
os.makedirs(VIDEOS_FOLDER, exist_ok=True)

# ============================
# API Route: Enhanced Record Violation with Timeline Support
# ============================
@app.route('/api/record_violation', methods=['POST'])
def record_violation():
    """Enhanced violation recording with timeline and pose data"""
    if 'user_id' not in session or session.get('role') != 'student':
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        data = request.get_json()
        
        # Enhanced violation record with timeline support
        violation_record = {
            'assessment_code': data.get('assessment_code'),
            'student_id': data.get('student_id'),
            'violation_type': data.get('violation_type'),  # Changed from 'type'
            'description': data.get('description'),
            'timestamp': datetime.now(),
            
            # Timeline data for lecturer analysis
            'exam_timeline_position': data.get('exam_timeline_position'),  # exam_start/middle/end
            'exam_start_time': data.get('exam_start_time'),
            'exam_end_time': data.get('exam_end_time'),
            'exam_progress_percent': data.get('exam_progress_percent'),
            'time_remaining_ms': data.get('time_remaining_ms'),
            'elapsed_time_ms': data.get('elapsed_time_ms'),
            
            # Head pose data when violation occurred
            'head_pose_data': data.get('head_pose_data'),
            
            # Detection settings that were active
            'detection_settings': data.get('detection_settings'),

            'status': 'pending',     # All new violations start as pending
            'notes': ''              # Empty notes initially
        }
        
        result = mongo.db.exam_violations.insert_one(violation_record)
        
        # Enhanced logging with timeline info
        timeline_info = violation_record.get('exam_timeline_position', 'unknown')
        progress_info = f"{violation_record.get('exam_progress_percent', 0):.1f}%" if violation_record.get('exam_progress_percent') else 'unknown'
        
        
        if violation_record.get('head_pose_data'):
            pose = violation_record['head_pose_data']
            print(f" Head pose: yaw={pose.get('yaw', 0):.1f}°, pitch={pose.get('pitch', 0):.1f}°, roll={pose.get('roll', 0):.1f}°")
        
        return jsonify({
            'status': 'success',
            'violation_id': str(result.inserted_id),
            'timeline_position': timeline_info,
            'progress_percent': violation_record.get('exam_progress_percent'),
            'message': 'Enhanced violation recorded with timeline data'
        }), 200
        
    except Exception as e:
        print(f"Error recording enhanced violation: {e}")
        return jsonify({'error': 'Failed to record violation', 'details': str(e)}), 500

# ============================
# API Route: HopeNet Head Pose Detection
# ============================
@app.route('/api/stream_head_pose', methods=['POST'])
def stream_head_pose():
    """Process head pose from video stream with enhanced face detection"""
    if 'user_id' not in session or session.get('role') != 'student':
        return jsonify({'error': 'Unauthorized'}), 401

    if head_pose_detector is None:
        return jsonify({'error': 'Head pose detector not available'}), 500

    try:
        data = request.get_json()
        
        if 'image_data' not in data:
            return jsonify({'error': 'No image data provided'}), 400
        
        # Decode base64 image
        try:
            image_data = data['image_data'].split(',')[1]
            image_bytes = base64.b64decode(image_data)
            nparr = np.frombuffer(image_bytes, np.uint8)
            image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        except Exception as e:
            print(f"Failed to decode image: {e}")
            return jsonify({'error': 'Invalid image data'}), 400
        
        if image is None:
            return jsonify({'error': 'Invalid image data'}), 400

        # Detect head pose with enhanced features
        pose_result = head_pose_detector.detect_head_pose_with_landmarks(image)
        
        if pose_result is None:
            return jsonify({
                'status': 'no_face',
                'message': 'No face detected',
                'timestamp': data.get('timestamp'),
                'assessment_code': data.get('assessment_code')
            }), 200

        pose = pose_result['pose']
        face_landmarks = pose_result.get('landmarks', [])
        face_bbox = pose_result.get('bbox')
        nose_point = pose_result.get('nose_point')
        confidence = pose_result.get('confidence', 0.5)

        # Check if looking away
        is_looking_away = head_pose_detector.is_looking_away(pose)
        
        return jsonify({
            'status': 'success',
            'pose': {
                'yaw': round(pose['yaw'], 1),
                'pitch': round(pose['pitch'], 1),
                'roll': round(pose['roll'], 1)
            },
            'face_landmarks': face_landmarks,
            'face_bbox': {
                'x': int(face_bbox['x']),
                'y': int(face_bbox['y']),
                'width': int(face_bbox['width']),
                'height': int(face_bbox['height'])
            },
            'nose_point': {
                'x': int(nose_point['x']),
                'y': int(nose_point['y'])
            },
            'is_looking_away': is_looking_away,
            'confidence': round(confidence, 2),
            'timestamp': data.get('timestamp'),
            'assessment_code': data.get('assessment_code')
        }), 200

    except Exception as e:
        print(f"Error in head pose stream: {e}")
        return jsonify({'error': 'Stream processing failed', 'details': str(e)}), 500

# ============================
# API Route: Enhanced Upload Evidence with Timeline Support
# ============================
@app.route('/api/upload_violation_evidence', methods=['POST'])
def upload_violation_evidence():
    """Enhanced evidence upload with timeline and pose data"""
    if 'user_id' not in session or session.get('role') != 'student':
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        # Get form data
        assessment_code = request.form.get('assessment_code')
        student_id = request.form.get('student_id')
        violation_type = request.form.get('violation_type')
        timestamp_str = request.form.get('timestamp')
        
        # Get enhanced timeline data
        exam_timeline_position = request.form.get('exam_timeline_position')
        exam_start_time = request.form.get('exam_start_time')
        exam_end_time = request.form.get('exam_end_time')
        exam_progress_percent = request.form.get('exam_progress_percent')
        head_pose_data_str = request.form.get('head_pose_data')
        detection_settings_str = request.form.get('detection_settings')
        
        # Parse JSON strings safely
        head_pose_data = None
        detection_settings = None
        
        try:
            if head_pose_data_str:
                head_pose_data = json.loads(head_pose_data_str)
        except json.JSONDecodeError as e:
            print(f"⚠️ Invalid head pose data JSON: {e}")
        
        try:
            if detection_settings_str:
                detection_settings = json.loads(detection_settings_str)
        except json.JSONDecodeError as e:
            print(f"⚠️ Invalid detection settings JSON: {e}")
        
        # Validate required fields
        if not all([assessment_code, student_id, violation_type]):
            return jsonify({'error': 'Missing required fields'}), 400

        uploaded_files = []
        timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:-3]  # Include milliseconds
        
        # Handle screenshot upload
        if 'screenshot' in request.files:
            screenshot = request.files['screenshot']
            if screenshot and screenshot.filename:
                try:
                    # Create secure filename with enhanced naming
                    filename = secure_filename(f"{violation_type}_screenshot_{assessment_code}_{student_id}_{timestamp_str}.png")
                    filepath = os.path.join(SCREENSHOTS_FOLDER, filename)
                    
                    # Ensure directory exists
                    os.makedirs(SCREENSHOTS_FOLDER, exist_ok=True)
                    
                    # Save screenshot
                    screenshot.save(filepath)
                    
                    # Verify file was saved
                    if os.path.exists(filepath):
                        file_size = os.path.getsize(filepath)
                        uploaded_files.append({
                            'type': 'screenshot',
                            'filename': filename,
                            'filepath': filepath,
                            'size': file_size,
                            'url': f'/static/evidence/screenshots/{filename}'
                        })
                        print(f"Screenshot saved: {filename} ({file_size} bytes)")
                    else:
                        print(f"Screenshot save failed: {filename}")
                        
                except Exception as e:
                    print(f"Error saving screenshot: {e}")

        # Handle video upload
        if 'video' in request.files:
            video = request.files['video']
            if video and video.filename:
                try:
                    # Get file extension
                    original_extension = os.path.splitext(video.filename)[1]
                    if not original_extension:
                        original_extension = '.webm'  # Default for browser recordings
                    
                    # Create secure filename with enhanced naming
                    filename = secure_filename(f"{violation_type}_enhanced_{assessment_code}_{student_id}_{timestamp_str}{original_extension}")
                    filepath = os.path.join(VIDEOS_FOLDER, filename)
                    
                    # Ensure directory exists
                    os.makedirs(VIDEOS_FOLDER, exist_ok=True)
                    
                    # Save video
                    video.save(filepath)
                    
                    # Verify file was saved
                    if os.path.exists(filepath):
                        file_size = os.path.getsize(filepath)
                        uploaded_files.append({
                            'type': 'video',
                            'filename': filename,
                            'filepath': filepath,
                            'size': file_size,
                            'duration': None,  # Could add video duration detection later
                            'url': f'/static/evidence/videos/{filename}'
                        })
                        print(f"Video saved: {filename} ({file_size} bytes)")
                    else:
                        print(f"Video save failed: {filename}")
                        
                except Exception as e:
                    print(f"Error saving video: {e}")

        # Create ENHANCED evidence record if files were uploaded
        if uploaded_files:
            violation_id = request.form.get('violation_id')

            evidence_record = {
                'violation_id': ObjectId(violation_id) if violation_id else None,
                'assessment_code': assessment_code,
                'student_id': student_id,
                'violation_type': violation_type,
                'timestamp': datetime.now(),
                
                # Timeline data for lecturer analysis
                'exam_timeline_position': exam_timeline_position,
                'exam_start_time': exam_start_time,
                'exam_end_time': exam_end_time,
                'exam_progress_percent': float(exam_progress_percent) if exam_progress_percent else None,
                
                # Head pose context when evidence was captured
                'head_pose_data': head_pose_data,
                'detection_settings': detection_settings,
                
                # Existing file data
                'files': uploaded_files,
                'total_files': len(uploaded_files),
                'total_size': sum(f.get('size', 0) for f in uploaded_files),
                'uploaded_at': datetime.now(),
                'status': 'uploaded'
            }
            
            # Insert enhanced evidence record
            result = mongo.db.exam_evidence.insert_one(evidence_record)
            
            # Enhanced logging
            timeline_info = exam_timeline_position or 'unknown'
            progress_info = f"{evidence_record.get('exam_progress_percent', 0):.1f}%" if evidence_record.get('exam_progress_percent') else 'unknown'
            
            if head_pose_data:
                print(f"Pose context: yaw={head_pose_data.get('yaw', 0):.1f}°, pitch={head_pose_data.get('pitch', 0):.1f}°")
            
            return jsonify({
                'status': 'success',
                'evidence_id': str(result.inserted_id),
                'files_uploaded': len(uploaded_files),
                'total_size': evidence_record['total_size'],
                'timeline_position': exam_timeline_position,
                'progress_percent': evidence_record.get('exam_progress_percent'),
                'message': f'Successfully uploaded {len(uploaded_files)} file(s) with enhanced timeline data'
            }), 200
        else:
            print(f"⚠️ No files uploaded for {violation_type} evidence")
            return jsonify({
                'status': 'warning',
                'message': 'No files were uploaded',
                'files_uploaded': 0,
                'timeline_position': exam_timeline_position
            }), 200
            
    except Exception as e:
        print(f"Error uploading enhanced evidence: {e}")
        traceback.print_exc()
        return jsonify({
            'error': 'Failed to upload evidence', 
            'details': str(e)
        }), 500

# ============================
# API Route: Get Exam Environment Rules with Head Movement Settings
# ============================
@app.route('/api/exam_environment_rules')
def get_exam_environment_rules():
    """Enhanced to include head movement settings"""
    try:
        rule = mongo.db.exam_environment_rules.find_one({}, {'_id': 0})
        
        # If no rule exists, return defaults
        if not rule:
            rule = {
                'block_f12': True,
                'block_ctrl_shift_i': False,
                'block_ctrl_u': False,
                'block_right_click': True,
                'block_copy': False,
                'block_paste': False,
                'head_movement_settings': {
                    'violation_duration': 2000,  # milliseconds
                    'warning_count': 3,
                    'max_yaw': 20,
                    'max_pitch': 10
                }
            }
        
        # Add head movement settings if they don't exist
        if 'head_movement_settings' not in rule:
            rule['head_movement_settings'] = {
                'violation_duration': 2000,
                'warning_count': 3,
                'max_yaw': 20,
                'max_pitch': 10
            }
        return jsonify(rule)
        
    except Exception as e:
        print(f"Error loading exam environment rules: {e}")
        return jsonify({'error': 'Failed to load rules'}), 500

@app.route('/api/get_head_movement_settings')
def get_head_movement_settings():
    """Enhanced API to get head movement settings for JavaScript"""
    try:
        rule = mongo.db.exam_environment_rules.find_one({}, {'_id': 0})
        
        if rule and 'head_movement_settings' in rule:
            settings = rule['head_movement_settings']
        else:
            # Default settings
            settings = {
                'violation_duration': 2000,  # milliseconds
                'warning_count': 3,
                'max_yaw': 20,
                'max_pitch': 10
            }
        
        return jsonify({
            'status': 'success',
            'settings': settings
        })
        
    except Exception as e:
        print(f"Error getting head movement settings: {e}")
        return jsonify({
            'status': 'error',
            'message': str(e),
            'settings': {
                'violation_duration': 2000,
                'warning_count': 3,
                'max_yaw': 20,
                'max_pitch': 10
            }
        }), 500

# ============================
# Additional API Routes for Enhanced Timeline Analysis
# ============================

@app.route('/api/get_exam_timeline/<assessment_code>')
def get_exam_timeline(assessment_code):
    """Get exam timeline information for a specific assessment"""
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        # Fetch timetable
        timetable = mongo.db.exam_timetables.find_one({'assessment_code': assessment_code})
        if not timetable:
            return jsonify({'error': 'No exam scheduled'}), 404
        
        # Calculate timeline
        exam_date = datetime.strptime(timetable['exam_date'], '%Y-%m-%d')
        start_time = datetime.strptime(timetable['start_time'], '%H:%M')
        end_time = datetime.strptime(timetable['end_time'], '%H:%M')
        
        start_dt = datetime.combine(exam_date.date(), start_time.time())
        end_dt = datetime.combine(exam_date.date(), end_time.time())
        
        return jsonify({
            'status': 'success',
            'exam_start_time': start_dt.isoformat(),
            'exam_end_time': end_dt.isoformat(),
            'duration_minutes': int((end_dt - start_dt).total_seconds() / 60)
        })
        
    except Exception as e:
        print(f"Error getting exam timeline: {e}")
        return jsonify({'error': 'Failed to get timeline'}), 500

@app.route('/api/violations/timeline/<assessment_code>')
def get_violations_by_timeline(assessment_code):
    """Get violations grouped by timeline position for lecturer analysis"""
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        # Get all violations for this assessment
        violations = list(mongo.db.exam_violations.find(
            {'assessment_code': assessment_code},
            {'_id': 0}
        ))
        
        # Group by timeline position
        timeline_analysis = {
            'exam_start': [],
            'exam_middle': [],
            'exam_end': [],
            'unknown': []
        }
        
        for violation in violations:
            position = violation.get('exam_timeline_position', 'unknown')
            if position in timeline_analysis:
                timeline_analysis[position].append(violation)
            else:
                timeline_analysis['unknown'].append(violation)
        
        # Calculate statistics
        stats = {
            'total_violations': len(violations),
            'by_position': {
                position: len(viols) for position, viols in timeline_analysis.items()
            }
        }
        
        return jsonify({
            'status': 'success',
            'assessment_code': assessment_code,
            'timeline_analysis': timeline_analysis,
            'statistics': stats
        })
        
    except Exception as e:
        print(f"Error getting timeline violations: {e}")
        return jsonify({'error': 'Failed to get violations'}), 500
    
# ============================
# Student My Submission
# ============================
@app.route('/student/submissions')
def student_submissions():
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    student_id = session['user_id']
    submissions = list(mongo.db.submissions.find({'student_id': student_id}).sort('submitted_at', -1))

    malaysia = timezone('Asia/Kuala_Lumpur')
    now = datetime.now(malaysia)

    # Fetch status from resubmission_requests
    requests = mongo.db.resubmission_requests.find({'student_id': student_id})
    status_map = {str(r['submission_id']): r['status'] for r in requests}

    for s in submissions:
        if isinstance(s['submitted_at'], datetime):
            s['submitted_at'] = s['submitted_at'].astimezone(malaysia)
            s['can_request_resubmission'] = now - s['submitted_at'] <= timedelta(hours=24) #student can only request resubmission within 24 hours
        else:
            s['can_request_resubmission'] = False
        s['status'] = status_map.get(str(s['_id']), None)

    assessment_codes = list({s['assessment_code'] for s in submissions})
    assessments = mongo.db.assessments.find({'assessment_code': {'$in': assessment_codes}})
    assessment_map = {a['assessment_code']: a['title'] for a in assessments}

    return render_template("student_submission.html",
                           submissions=submissions,
                           assessment_map=assessment_map)

# ============================
# Student Request Resubmit
# ============================
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
@app.route('/student/request-resubmission/<submission_id>', methods=['GET', 'POST'])
def student_request_resubmission(submission_id):
    if 'user_id' not in session or session.get('role') != 'student':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))

    # Get submission
    submission = mongo.db.submissions.find_one({'_id': ObjectId(submission_id)})
    if not submission:
        flash("Submission not found.", "danger")
        return redirect(url_for('student_submissions'))

    student = mongo.db.users.find_one({'user_id': submission['student_id']})
    if not student:
        flash("Student details not found.", "danger")
        return redirect(url_for('student_submissions'))

    # Get course name (via assessment ➝ course_codes[0] ➝ course_name)
    assessment = mongo.db.assessments.find_one({'assessment_code': submission['assessment_code']})
    course_name = ""
    if assessment and assessment.get('course_codes'):
        course = mongo.db.courses.find_one({'course_code': assessment['course_codes'][0]})
        if course:
            course_name = course['course_name']

    # Handle form submission
    if request.method == 'POST':
        new_file = request.files.get('new_file')
        reason = request.form.get('reason')

        if not new_file or new_file.filename == "":
            flash("You must upload a file for resubmission.", "error")
            return redirect(request.url)

        if not allowed_file(new_file.filename):
            flash("Invalid file format. Upload txt or doc/docx.", "error")
            return redirect(request.url)

        filename = secure_filename(f"{submission['assessment_code']}_{student['user_id']}_resub_{datetime.now().strftime('%Y%m%d%H%M%S')}_{new_file.filename}")
        save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        new_file.save(save_path)

        mongo.db.resubmission_requests.insert_one({
            "submission_id": submission_id,
            "student_id": student['user_id'],
            "assessment_code": submission['assessment_code'],
            "old_filename": submission['filename'],
            "submitted_at": submission['submitted_at'],
            "new_filename": filename,
            "reason": reason,
            "status": "Pending",
            "requested_at": datetime.now()
        })

        flash("Resubmission request sent successfully.", "success")
        return redirect(url_for('student_submissions'))

    return render_template('student_request_resubmission.html',
                           student=student,
                           submission=submission,
                           course_name=course_name)

# ============================
# Report module for lecturers
# ============================

# =========================
# USER REGISTRATION REPORT 
# =========================
@app.route('/lecturer/reports/user-registration')
def user_registration_report():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))
    return render_template('reports/user_registration_report.html')


@app.route('/api/reports/user-registration', methods=['GET'])
def get_user_registration_data():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 403

    """
    API endpoint to retrieve user registration data for reports.
    
    Query Parameters:
        type (str): Report type - 'monthly', 'yearly', or 'custom'
        from_date (str): Start date for custom range (YYYY-MM-DD format)
        to_date (str): End date for custom range (YYYY-MM-DD format)
        user_type (str): User type filter - 'all', 'student', or 'lecturer'
    
    Returns:
        JSON response containing:
        - success (bool): Operation status
        - data (dict): Chart data with labels and datasets
        - summary (dict): Summary statistics with growth rates
        - report_type (str): Type of report generated
        - date_range (dict): Date range information
        
    Raises:
        500: Internal server error if database operation fails
    """
    try:
        # Extract query parameters
        report_type = request.args.get('type', 'monthly')
        from_date = request.args.get('from_date')
        to_date = request.args.get('to_date')
        user_type = request.args.get('user_type', 'all')
        
        # Determine date ranges based on report type
        end_date = datetime.now()
        
        if report_type == 'custom' and from_date and to_date:
            start_date = datetime.strptime(from_date, '%Y-%m-%d')
            end_date = datetime.strptime(to_date, '%Y-%m-%d')
            chart_start_date = start_date
        elif report_type == 'monthly':
            # Summary cards: current month only, Chart: full year for context
            start_date = datetime(end_date.year, end_date.month, 1)
            chart_start_date = datetime(end_date.year, 1, 1)
        elif report_type == 'yearly':
            # Summary cards: current year only, Chart: last 3 years for context
            start_date = datetime(end_date.year, 1, 1)
            chart_start_date = datetime(end_date.year - 2, 1, 1)
        else:
            # Default to monthly
            start_date = datetime(end_date.year, end_date.month, 1)
            chart_start_date = datetime(end_date.year, 1, 1)
        
        # Build MongoDB aggregation match criteria
        chart_match_stage = {
            'created_at': {'$gte': chart_start_date, '$lte': end_date}
        }
        if user_type != 'all':
            chart_match_stage['role'] = user_type
            
        # Create aggregation pipeline based on report type
        if report_type == 'yearly':
            pipeline = [
                {'$match': chart_match_stage},
                {
                    '$group': {
                        '_id': {
                            'year': {'$year': '$created_at'},
                            'role': '$role'
                        },
                        'count': {'$sum': 1}
                    }
                },
                {'$sort': {'_id.year': 1}}
            ]
        else:  # monthly or custom
            pipeline = [
                {'$match': chart_match_stage},
                {
                    '$group': {
                        '_id': {
                            'year': {'$year': '$created_at'},
                            'month': {'$month': '$created_at'},
                            'role': '$role'
                        },
                        'count': {'$sum': 1}
                    }
                },
                {'$sort': {'_id.year': 1, '_id.month': 1}}
            ]
        
        # Execute aggregation and process results
        chart_results = list(mongo.db.users.aggregate(pipeline))
        chart_data = process_chart_data(chart_results, chart_start_date, end_date, report_type, user_type)
        summary_stats = get_summary_statistics(start_date, end_date, user_type, report_type)
        
        return jsonify({
            'success': True,
            'data': chart_data,
            'summary': summary_stats,
            'report_type': report_type,
            'date_range': {
                'from': start_date.strftime('%Y-%m-%d'),
                'to': end_date.strftime('%Y-%m-%d'),
                'display': f"{start_date.strftime('%m/%d/%Y')} - {end_date.strftime('%m/%d/%Y')}"
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

def process_chart_data(results, start_date, end_date, report_type, user_type):
    """
    Process MongoDB aggregation results into chart-ready format.
    
    Args:
        results (list): Raw aggregation results from MongoDB
        start_date (datetime): Start date for the chart data range
        end_date (datetime): End date for the chart data range
        report_type (str): Type of report ('monthly', 'yearly', 'custom')
        user_type (str): User type filter applied
    
    Returns:
        dict: Processed chart data containing:
            - labels (list): Chart labels (months/years)
            - students_data (list): Student registration counts per period
            - lecturers_data (list): Lecturer registration counts per period
            - table_data (list): Detailed data for table display with growth rates
    """
    # Organize aggregation results by period and role
    data_dict = {}
    
    for result in results:
        if report_type == 'yearly':
            period = str(result['_id']['year'])
        else:
            year = result['_id']['year']
            month = result['_id']['month']
            period = f"{year}-{month:02d}"
        
        role = result['_id']['role']
        count = result['count']
        
        if period not in data_dict:
            data_dict[period] = {'student': 0, 'lecturer': 0, 'total': 0}
        
        data_dict[period][role] = count
        data_dict[period]['total'] = data_dict[period]['student'] + data_dict[period]['lecturer']
    
    # Generate time series data with proper labels
    labels = []
    students_data = []
    lecturers_data = []
    table_data = []
    
    if report_type == 'yearly':
        # Generate yearly data for the specified range
        current_year = start_date.year
        while current_year <= end_date.year:
            period = str(current_year)
            labels.append(period)
            
            students_count = data_dict.get(period, {}).get('student', 0)
            lecturers_count = data_dict.get(period, {}).get('lecturer', 0)
            total_count = students_count + lecturers_count
            
            students_data.append(students_count)
            lecturers_data.append(lecturers_count)
            
            # Calculate period-over-period growth for table
            growth_rate = 0
            if len(table_data) > 0:
                previous_total = table_data[-1]['total']
                if previous_total > 0:
                    growth_rate = ((total_count - previous_total) / previous_total) * 100
            
            table_data.append({
                'period': period,
                'period_label': period,
                'students': students_count,
                'lecturers': lecturers_count,
                'total': total_count,
                'growth_rate': round(growth_rate, 1)
            })
            
            current_year += 1
    else:
        # Generate monthly data for the specified range
        current = start_date.replace(day=1)
        while current <= end_date:
            period = f"{current.year}-{current.month:02d}"
            period_label = current.strftime('%b %Y')
            labels.append(period_label)
            
            students_count = data_dict.get(period, {}).get('student', 0)
            lecturers_count = data_dict.get(period, {}).get('lecturer', 0)
            total_count = students_count + lecturers_count
            
            students_data.append(students_count)
            lecturers_data.append(lecturers_count)
            
            # Calculate month-over-month growth for table
            growth_rate = 0
            if len(table_data) > 0:
                previous_total = table_data[-1]['total']
                if previous_total > 0:
                    growth_rate = ((total_count - previous_total) / previous_total) * 100
            
            table_data.append({
                'period': period,
                'period_label': period_label,
                'students': students_count,
                'lecturers': lecturers_count,
                'total': total_count,
                'growth_rate': round(growth_rate, 1)
            })
            
            # Advance to next month
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1)
            else:
                current = current.replace(month=current.month + 1)
    
    return {
        'labels': labels,
        'students_data': students_data,
        'lecturers_data': lecturers_data,
        'table_data': table_data
    }

def get_summary_statistics(start_date, end_date, user_type, report_type):
    """
    Calculate summary statistics and growth rates for the specified period.
    
    Args:
        start_date (datetime): Start date of the current period
        end_date (datetime): End date of the current period
        user_type (str): User type filter ('all', 'student', 'lecturer')
        report_type (str): Report type ('monthly', 'yearly', 'custom')
    
    Returns:
        dict: Summary statistics containing:
            - total_users (int): Total users in current period
            - students (int): Student count in current period
            - lecturers (int): Lecturer count in current period
            - total_growth (float): Growth rate compared to previous period
            - students_growth (float): Student growth rate
            - lecturers_growth (float): Lecturer growth rate
            - growth_period_text (str): Description of comparison period
            - has_previous_data (bool): Whether previous period data exists
    """
    # Build match criteria for current period
    match_criteria = {
        'created_at': {'$gte': start_date, '$lte': end_date}
    }
    if user_type != 'all':
        match_criteria['role'] = user_type
    
    # Get current period counts by role
    students_count = mongo.db.users.count_documents({
        **match_criteria,
        'role': 'student'
    })
    
    lecturers_count = mongo.db.users.count_documents({
        **match_criteria,
        'role': 'lecturer'
    })
    
    total_users = students_count + lecturers_count
    
    # Calculate previous period date range based on report type
    growth_period_text = "last period"
    
    if report_type == 'monthly':
        # Compare with previous month
        if start_date.month == 1:
            # January comparison with December of previous year
            previous_year = start_date.year - 1
            previous_month = 12
        else:
            # Other months comparison with previous month
            previous_year = start_date.year
            previous_month = start_date.month - 1
        
        # Calculate full previous month date range
        previous_start = datetime(previous_year, previous_month, 1)
        last_day = calendar.monthrange(previous_year, previous_month)[1]
        previous_end = datetime(previous_year, previous_month, last_day, 23, 59, 59, 999999)
        
        growth_period_text = "last month"
        
    elif report_type == 'yearly':
        # Compare with previous year
        current_year = start_date.year
        previous_start = datetime(current_year - 1, 1, 1)
        previous_end = datetime(current_year - 1, 12, 31, 23, 59, 59, 999999)
        growth_period_text = "last year"
        
    else:  # custom
        # Compare with equivalent period duration before current range
        period_duration = end_date - start_date
        previous_start = start_date - period_duration
        previous_end = start_date - timedelta(microseconds=1)
        
        # Generate appropriate description based on duration
        duration_days = period_duration.days
        if duration_days <= 7:
            growth_period_text = "previous week"
        elif duration_days <= 31:
            growth_period_text = "previous month"
        elif duration_days <= 93:
            growth_period_text = "previous quarter"
        elif duration_days <= 186:
            growth_period_text = "previous 6 months"
        elif duration_days <= 366:
            growth_period_text = "previous year"
        else:
            growth_period_text = f"previous {duration_days} days"
    
    # Query previous period data
    previous_match = {
        'created_at': {'$gte': previous_start, '$lte': previous_end}
    }
    if user_type != 'all':
        previous_match['role'] = user_type
    
    previous_students = mongo.db.users.count_documents({
        **previous_match,
        'role': 'student'
    })
    
    previous_lecturers = mongo.db.users.count_documents({
        **previous_match,
        'role': 'lecturer'
    })
    
    previous_total = previous_students + previous_lecturers
    
    # Calculate growth rates with proper handling of edge cases
    def calculate_growth(current, previous):
        """
        Calculate percentage growth between two values.
        
        Args:
            current (int): Current period value
            previous (int): Previous period value
            
        Returns:
            float: Growth percentage rounded to 2 decimal places
        """
        if previous > 0:
            growth = ((current - previous) / previous) * 100
            return round(growth, 2)
        else:
            return 0 if current == 0 else 100
    
    total_growth = calculate_growth(total_users, previous_total)
    students_growth = calculate_growth(students_count, previous_students)
    lecturers_growth = calculate_growth(lecturers_count, previous_lecturers)
    
    return {
        'total_users': total_users,
        'students': students_count,
        'lecturers': lecturers_count,
        'total_growth': total_growth,
        'students_growth': students_growth,
        'lecturers_growth': lecturers_growth,
        'growth_period_text': growth_period_text,
        'has_previous_data': previous_total > 0
    }

# ============================
# ASSESSMENT ENROLLMENT REPORTS
# ============================
@app.route('/lecturer/reports/assessment-enrollment')
def assessment_enrollment_report():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))
    return render_template('reports/assessment_enrollment_report.html')

@app.route('/api/assessments/list', methods=['GET'])
def get_assessments_list():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 403
    
    try:
        # Get all assessments for the filter dropdown
        assessments = list(mongo.db.assessments.find(
            {}, 
            {'assessment_code': 1, 'title': 1, '_id': 0}
        ).sort('assessment_code', 1))
        
        return jsonify({
            'success': True,
            'assessments': assessments
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reports/assessment-enrollment', methods=['GET'])
def get_assessment_enrollment_data():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 403
    
    try:
        # Get query parameters for filtering
        assessment_filter = request.args.get('assessment', 'all')
        enrollment_type_filter = request.args.get('enrollment_type', 'all')
        
        # Get all analysis data for the dashboard
        coverage_data = get_assessment_coverage_analysis(assessment_filter)
        methods_data = get_assessment_enrollment_methods_analysis(assessment_filter)
        left_out_data = get_assessment_left_out_students_analysis(assessment_filter)
        summary_data = get_assessment_enrollment_summary_stats(assessment_filter)
        self_enrollment_chart_data = get_self_enrollment_by_assessment_chart(assessment_filter)
        
        return jsonify({
            'success': True,
            'data': {
                'coverage': coverage_data,
                'methods': methods_data,
                'left_out': left_out_data,
                'summary': summary_data,
                'self_enrollment_by_assessment': self_enrollment_chart_data
            },
            'filters': {
                'assessment': assessment_filter,
                'enrollment_type': enrollment_type_filter
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

def get_assessment_coverage_analysis(assessment_filter):
    """
    Get coverage analysis for assessments - CLEAN VERSION
    """
    try:
        # Build match stage based on assessment filter
        match_stage = {}
        if assessment_filter != 'all':
            match_stage['assessment_code'] = assessment_filter
        
        # Get assessments directly
        assessments = list(mongo.db.assessments.find(match_stage))
        
        assessments_data = []
        
        for assessment in assessments:
            assessment_code = assessment.get('assessment_code', 'N/A')
            assessment_title = assessment.get('title', 'N/A')
            
            # Get enrollment data
            students = assessment.get('students', [])
            enrollment_details = assessment.get('enrollment_details', {})
            self_enrolled = enrollment_details.get('self_enrolled', [])
            non_self_enrolled = enrollment_details.get('non_self_enrolled', [])
            
            enrolled_students_count = len(students)
            self_enrolled_count = len(self_enrolled)
            lecturer_enrolled_count = len(non_self_enrolled)
            
            # Get course students count - UNIQUE students across all linked courses
            course_codes = assessment.get('course_codes', [])
            if isinstance(course_codes, str):
                course_codes = [course_codes]
            
            all_course_students = set()
            
            if course_codes:
                for course_code in course_codes:
                    course = mongo.db.courses.find_one({'course_code': course_code})
                    if course and course.get('students'):
                        course_students = course['students']
                        all_course_students.update(course_students)
                        
            course_students_count = len(all_course_students)
            
            # Calculate coverage rate
            if course_students_count > 0:
                coverage_rate = round((enrolled_students_count / course_students_count) * 100, 1)
            else:
                coverage_rate = 0
            
            # Calculate students left out - NEW ADDITION
            left_out_students = max(0, course_students_count - enrolled_students_count)
            
            # Keep status for backward compatibility (can remove later if not needed)
            if coverage_rate >= 95:
                status = 'Excellent'
            elif coverage_rate >= 85:
                status = 'Good'
            elif coverage_rate >= 70:
                status = 'Needs Attention'
            else:
                status = 'Critical'
            
            assessments_data.append({
                'assessment_code': assessment_code,
                'assessment_title': assessment_title,
                'course_students': course_students_count,
                'enrolled_students': enrolled_students_count,
                'coverage_rate': coverage_rate,
                'self_enrolled': self_enrolled_count,
                'lecturer_enrolled': lecturer_enrolled_count,
                'left_out_students': left_out_students,
                'status': status  # Keep for compatibility
            })
        
        return {'assessments': assessments_data}
        
    except Exception as e:
        return {'assessments': []}

def get_assessment_enrollment_methods_analysis(assessment_filter):
    """
    Get enrollment methods breakdown by assessment - CLEAN VERSION
    """
    try:
        match_stage = {}
        if assessment_filter != 'all':
            match_stage['assessment_code'] = assessment_filter
        
        # Get assessments directly
        assessments = list(mongo.db.assessments.find(match_stage))
        
        assessments_data = []
        total_self_enrolled = 0
        total_lecturer_enrolled = 0
        
        for assessment in assessments:
            # Get enrollment data
            students = assessment.get('students', [])
            enrollment_details = assessment.get('enrollment_details', {})
            self_enrolled = enrollment_details.get('self_enrolled', [])
            non_self_enrolled = enrollment_details.get('non_self_enrolled', [])
            
            total_enrolled = len(students)
            self_enrolled_count = len(self_enrolled)
            lecturer_enrolled_count = len(non_self_enrolled)
            
            # Calculate self-enrollment rate
            if total_enrolled > 0:
                self_enrollment_rate = round((self_enrolled_count / total_enrolled) * 100, 1)
            else:
                self_enrollment_rate = 0
            
            # Determine effectiveness
            if self_enrollment_rate >= 20:
                effectiveness = 'High'
            elif self_enrollment_rate >= 10:
                effectiveness = 'Medium'
            else:
                effectiveness = 'Low'
            
            assessments_data.append({
                'assessment_code': assessment.get('assessment_code', 'N/A'),
                'assessment_title': assessment.get('title', 'N/A'),
                'total_enrolled': total_enrolled,
                'self_enrolled': self_enrolled_count,
                'lecturer_enrolled': lecturer_enrolled_count,
                'self_enrollment_rate': self_enrollment_rate,
                'effectiveness': effectiveness
            })
            
            # Add to totals
            total_self_enrolled += self_enrolled_count
            total_lecturer_enrolled += lecturer_enrolled_count
        
        return {
            'assessments': assessments_data,
            'totals': {
                'self_enrolled': total_self_enrolled,
                'lecturer_enrolled': total_lecturer_enrolled
            }
        }
        
    except Exception as e:
        return {'assessments': [], 'totals': {'self_enrolled': 0, 'lecturer_enrolled': 0}}

def get_assessment_left_out_students_analysis(assessment_filter):
    """
    Find students who are in courses but not enrolled in assessments - CLEAN VERSION
    """
    try:
        # Build match stage based on assessment filter
        match_stage = {}
        if assessment_filter != 'all':
            match_stage['assessment_code'] = assessment_filter
        
        # Get assessments directly
        assessments = list(mongo.db.assessments.find(match_stage))
        
        left_out_students = []
        affected_courses = set()
        
        for assessment in assessments:
            assessment_code = assessment.get('assessment_code', 'Unknown')
            assessment_title = assessment.get('title', assessment_code)
            
            # Get course codes for this assessment
            course_codes = assessment.get('course_codes', [])
            if isinstance(course_codes, str):
                course_codes = [course_codes]
            
            # Get students enrolled in this assessment
            assessment_students = set(assessment.get('students', []))
            
            # For each linked course, find students not in assessment
            for course_code in course_codes:
                course = mongo.db.courses.find_one({'course_code': course_code})
                if course and course.get('students'):
                    course_students = set(course['students'])
                    course_name = course.get('course_name', course_code)
                    
                    # Find students in course but not in assessment
                    missing_students = course_students - assessment_students
                    
                    if missing_students:
                        affected_courses.add(course_code)
                        
                        for student_id in missing_students:
                            # Get student details
                            student = mongo.db.users.find_one({'user_id': student_id})
                            student_name = student.get('name', 'Unknown') if student else 'Unknown'
                            
                            left_out_students.append({
                                'student_id': student_id,
                                'student_name': student_name,
                                'course_code': course_code,
                                'course_name': course_name,
                                'assessment_code': assessment_code,
                                'assessment_title': assessment_title
                            })
        
        return {
            'students': left_out_students,
            'total_count': len(left_out_students),
            'affected_courses': list(affected_courses)
        }
        
    except Exception as e:
        return {'students': [], 'total_count': 0, 'affected_courses': []}

def get_assessment_enrollment_summary_stats(assessment_filter):
    """
    Calculate summary statistics - CLEAN VERSION (No MongoDB aggregation)
    """
    try:
        # Build match criteria for assessments
        match_criteria = {}
        if assessment_filter != 'all':
            match_criteria['assessment_code'] = assessment_filter
        
        # Get assessments directly
        assessments = list(mongo.db.assessments.find(match_criteria))
        
        if not assessments:
            return {
                'total_assessments': 0,
                'average_coverage_rate': 0,
                'self_enrollment_rate': 0,
                'students_left_out': 0
            }
        
        total_assessments = len(assessments)
        total_enrolled = 0
        total_self_enrolled = 0
        total_coverage_points = 0
        total_left_out = 0
        
        # Process each assessment
        for assessment in assessments:
            # Get enrollment data
            students = assessment.get('students', [])
            enrollment_details = assessment.get('enrollment_details', {})
            self_enrolled = enrollment_details.get('self_enrolled', [])
            
            enrolled_count = len(students)
            self_enrolled_count = len(self_enrolled)
            
            total_enrolled += enrolled_count
            total_self_enrolled += self_enrolled_count
            
            # Get course students count for this assessment
            course_codes = assessment.get('course_codes', [])
            if isinstance(course_codes, str):
                course_codes = [course_codes]
            
            # Calculate unique students across all linked courses
            all_course_students = set()
            
            if course_codes:
                for course_code in course_codes:
                    course = mongo.db.courses.find_one({'course_code': course_code})
                    if course and course.get('students'):
                        course_students = course['students']
                        all_course_students.update(course_students)
            
            total_course_students = len(all_course_students)
            
            # Calculate coverage rate for this assessment
            if total_course_students > 0:
                coverage_rate = (enrolled_count / total_course_students) * 100
                total_coverage_points += coverage_rate
                
                # Calculate left out students for this assessment
                enrolled_set = set(students)
                left_out = len(all_course_students - enrolled_set)
                total_left_out += left_out
        
        # Calculate final rates
        if total_assessments > 0:
            average_coverage_rate = round(total_coverage_points / total_assessments, 1)
        else:
            average_coverage_rate = 0
        
        if total_enrolled > 0:
            self_enrollment_rate = round((total_self_enrolled / total_enrolled) * 100, 1)
        else:
            self_enrollment_rate = 0
        
        return {
            'total_assessments': total_assessments,
            'average_coverage_rate': average_coverage_rate,
            'self_enrollment_rate': self_enrollment_rate,
            'students_left_out': total_left_out
        }
        
    except Exception as e:
        return {
            'total_assessments': 0,
            'average_coverage_rate': 0,
            'self_enrollment_rate': 0,
            'students_left_out': 0
        }

def get_self_enrollment_by_assessment_chart(assessment_filter):
    """
    Get data for bar chart showing self-enrollment count by assessment - CLEAN VERSION
    """
    try:
        match_stage = {}
        if assessment_filter != 'all':
            match_stage['assessment_code'] = assessment_filter
        
        # Get assessments directly
        assessments = list(mongo.db.assessments.find(match_stage))
        
        labels = []
        self_enrolled_counts = []
        
        for assessment in assessments:
            assessment_code = assessment.get('assessment_code', 'Unknown')
            enrollment_details = assessment.get('enrollment_details', {})
            self_enrolled = enrollment_details.get('self_enrolled', [])
            self_enrolled_count = len(self_enrolled)
            
            labels.append(assessment_code)
            self_enrolled_counts.append(self_enrolled_count)
        
        # Sort by assessment code
        sorted_data = sorted(zip(labels, self_enrolled_counts))
        labels, self_enrolled_counts = zip(*sorted_data) if sorted_data else ([], [])
        
        return {
            'labels': list(labels),
            'self_enrolled_counts': list(self_enrolled_counts)
        }
        
    except Exception as e:
        return {'labels': [], 'self_enrolled_counts': []}

# ============================
# VIOLATION ANALYSIS REPORTS
# ============================
@app.route('/lecturer/reports/violation-analysis')
def violation_analysis_report():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))
    return render_template('reports/violation_analysis_report.html')

@app.route('/api/assessments/list-for-violations', methods=['GET'])
def get_assessments_for_violations():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 403
    
    try:
        # Get all assessments that have violations
        assessments_with_violations = list(mongo.db.exam_violations.aggregate([
            {'$group': {'_id': '$assessment_code'}},
            {'$lookup': {
                'from': 'assessments',
                'localField': '_id',
                'foreignField': 'assessment_code',
                'as': 'assessment_info'
            }},
            {'$unwind': {'path': '$assessment_info', 'preserveNullAndEmptyArrays': True}},
            {'$project': {
                'assessment_code': '$_id',
                'title': {'$ifNull': ['$assessment_info.title', 'Unknown Assessment']}
            }},
            {'$sort': {'assessment_code': 1}}
        ]))
        
        return jsonify({
            'success': True,
            'assessments': assessments_with_violations
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reports/violation-analysis', methods=['GET'])
def get_violation_analysis_data():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 403
    
    try:
        # Get query parameters for filtering
        assessment_filter = request.args.get('assessment', 'all')
        violation_type_filter = request.args.get('violation_type', 'all')
        status_filter = request.args.get('status', 'all')
        
        # Get all analysis data - ANALYTICS FOCUSED
        summary_data = get_violation_summary_stats(assessment_filter, violation_type_filter, status_filter)
        chart_data = get_violation_chart_data(assessment_filter, violation_type_filter, status_filter)
        assessment_breakdown = get_assessment_breakdown_data(assessment_filter, violation_type_filter, status_filter)
        head_movement_details = get_head_movement_detailed_data(assessment_filter, violation_type_filter, status_filter)
        
        return jsonify({
            'success': True,
            'data': {
                'summary': summary_data,
                'charts': chart_data,
                'assessment_breakdown': assessment_breakdown,
                'head_movement_details': head_movement_details
            },
            'filters': {
                'assessment': assessment_filter,
                'violation_type': violation_type_filter,
                'status': status_filter
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

def get_violation_summary_stats(assessment_filter, violation_type_filter, status_filter):
    """Get summary statistics for system-wide analytics"""
    try:
        # Build match criteria
        match_criteria = {}
        if assessment_filter != 'all':
            match_criteria['assessment_code'] = assessment_filter
        if violation_type_filter != 'all':
            match_criteria['violation_type'] = violation_type_filter
        if status_filter != 'all':
            match_criteria['status'] = status_filter
        
        # Get all violations matching criteria
        violations = list(mongo.db.exam_violations.find(match_criteria))
        
        total_violations = len(violations)
        
        # Count by status
        status_counts = {'pending': 0, 'reviewed': 0, 'dismissed': 0}
        for violation in violations:
            status = violation.get('status', 'pending')
            if status in status_counts:
                status_counts[status] += 1
        
        # Calculate false positive rate
        false_positive_rate = 0
        if total_violations > 0:
            false_positive_rate = round((status_counts['dismissed'] / total_violations) * 100, 1)
        
        # Count unique assessments and students
        unique_assessments = len(set(v.get('assessment_code') for v in violations if v.get('assessment_code')))
        unique_students = len(set(v.get('student_id') for v in violations if v.get('student_id')))
        
        # Calculate review efficiency
        reviewed_total = status_counts['reviewed'] + status_counts['dismissed']
        review_rate = 0
        if total_violations > 0:
            review_rate = round((reviewed_total / total_violations) * 100, 1)
        
        return {
            'total_violations': total_violations,
            'pending_count': status_counts['pending'],
            'reviewed_count': status_counts['reviewed'],
            'dismissed_count': status_counts['dismissed'],
            'false_positive_rate': false_positive_rate,
            'unique_assessments': unique_assessments,
            'unique_students': unique_students,
            'review_rate': review_rate
        }
        
    except Exception as e:
        return {
            'total_violations': 0,
            'pending_count': 0,
            'reviewed_count': 0,
            'dismissed_count': 0,
            'false_positive_rate': 0,
            'unique_assessments': 0,
            'unique_students': 0,
            'review_rate': 0
        }

def get_violation_chart_data(assessment_filter, violation_type_filter, status_filter):
    """Get chart data for system-wide analytics - FINAL VERSION"""
    try:
        # Build match criteria
        match_criteria = {}
        if assessment_filter != 'all':
            match_criteria['assessment_code'] = assessment_filter
        if violation_type_filter != 'all':
            match_criteria['violation_type'] = violation_type_filter
        if status_filter != 'all':
            match_criteria['status'] = status_filter
        
        # Get violations
        violations = list(mongo.db.exam_violations.find(match_criteria))
        
        # 1. Violation type distribution (for bar chart)
        all_violation_types = ['TAB_SWITCH', 'SCREEN_SHARE_STOPPED', 'WEBCAM_DENIED', 'FORBIDDEN_SHORTCUT', 'looking_away']
        type_counts = {vtype: 0 for vtype in all_violation_types}
        
        for violation in violations:
            v_type = violation.get('violation_type', 'unknown')
            if v_type in type_counts:
                type_counts[v_type] += 1
        
        # Only include types that have violations > 0
        filtered_type_counts = {k: v for k, v in type_counts.items() if v > 0}
        
        # 2. Violations by exams (for horizontal bar chart)
        exam_counts = {}
        for violation in violations:
            exam_code = violation.get('assessment_code', 'unknown')
            exam_counts[exam_code] = exam_counts.get(exam_code, 0) + 1
        
        # Sort exams by violation count (descending) for better visualization
        sorted_exam_items = sorted(exam_counts.items(), key=lambda x: x[1], reverse=True)
        exam_labels = [item[0] for item in sorted_exam_items]
        exam_data = [item[1] for item in sorted_exam_items]
        
        # 3. Timeline distribution (ONLY for head movement / looking_away violations)
        timeline_counts = {'exam_start': 0, 'exam_middle': 0, 'exam_end': 0}
        head_movement_violations = [v for v in violations if v.get('violation_type') == 'looking_away']
        
        for violation in head_movement_violations:
            timeline_position = violation.get('exam_timeline_position', 'unknown')
            if timeline_position in timeline_counts:
                timeline_counts[timeline_position] += 1
        
        # 4. Status distribution
        status_counts = {'pending': 0, 'reviewed': 0, 'dismissed': 0}
        for violation in violations:
            status = violation.get('status', 'pending')
            if status in status_counts:
                status_counts[status] += 1
        
        return {
            'violation_types': {
                'labels': list(filtered_type_counts.keys()),
                'data': list(filtered_type_counts.values())
            },
            'violations_by_exam': {
                'labels': exam_labels,
                'data': exam_data
            },
            'head_movement_timeline': {
                'labels': ['Early Stage', 'Middle Stage', 'Late Stage'],
                'data': [timeline_counts['exam_start'], timeline_counts['exam_middle'], timeline_counts['exam_end']],
                'total_head_movements': len(head_movement_violations)
            },
            'status_distribution': {
                'labels': list(status_counts.keys()),
                'data': list(status_counts.values())
            }
        }
        
    except Exception as e:
        return {
            'violation_types': {'labels': [], 'data': []},
            'violations_by_exam': {'labels': [], 'data': []},
            'head_movement_timeline': {'labels': [], 'data': [], 'total_head_movements': 0},
            'status_distribution': {'labels': [], 'data': []}
        }

def get_assessment_breakdown_data(assessment_filter, violation_type_filter, status_filter):
    """Get violations breakdown by assessment for risk analysis"""
    try:
        # Build match criteria
        match_criteria = {}
        if assessment_filter != 'all':
            match_criteria['assessment_code'] = assessment_filter
        if violation_type_filter != 'all':
            match_criteria['violation_type'] = violation_type_filter
        if status_filter != 'all':
            match_criteria['status'] = status_filter
        
        # Aggregate by assessment
        pipeline = [
            {'$match': match_criteria},
            {'$group': {
                '_id': '$assessment_code',
                'total_violations': {'$sum': 1},
                'pending_count': {'$sum': {'$cond': [{'$eq': ['$status', 'pending']}, 1, 0]}},
                'reviewed_count': {'$sum': {'$cond': [{'$eq': ['$status', 'reviewed']}, 1, 0]}},
                'dismissed_count': {'$sum': {'$cond': [{'$eq': ['$status', 'dismissed']}, 1, 0]}},
                'unique_students': {'$addToSet': '$student_id'}
            }},
            {'$project': {
                'assessment_code': '$_id',
                'total_violations': 1,
                'pending_count': 1,
                'reviewed_count': 1,
                'dismissed_count': 1,
                'unique_students_count': {'$size': '$unique_students'}
            }},
            {'$sort': {'total_violations': -1}}
        ]
        
        results = list(mongo.db.exam_violations.aggregate(pipeline))
        
        # Enrich with assessment titles and calculate metrics
        for result in results:
            assessment_code = result['assessment_code']
            assessment = mongo.db.assessments.find_one({'assessment_code': assessment_code})
            result['assessment_title'] = assessment.get('title', 'Unknown') if assessment else 'Unknown'
            
            # Calculate false positive rate for this assessment
            total = result['total_violations']
            dismissed = result['dismissed_count']
            result['false_positive_rate'] = round((dismissed / total * 100), 1) if total > 0 else 0
            
            # Calculate review progress
            reviewed_total = result['reviewed_count'] + result['dismissed_count']
            result['review_progress'] = round((reviewed_total / total * 100), 1) if total > 0 else 0
        
        return results
        
    except Exception as e:
        return []

def get_head_movement_detailed_data(assessment_filter, violation_type_filter, status_filter, limit=50):
    """Get detailed head movement violations for timeline analysis table"""
    try:
        # Build match criteria - only for looking_away violations
        match_criteria = {'violation_type': 'looking_away'}
        if assessment_filter != 'all':
            match_criteria['assessment_code'] = assessment_filter
        if status_filter != 'all':
            match_criteria['status'] = status_filter
        
        # Get detailed head movement violations
        violations = list(mongo.db.exam_violations.find(
            match_criteria,
            {
                '_id': 1,
                'student_id': 1,
                'assessment_code': 1,
                'exam_timeline_position': 1,
                'exam_progress_percent': 1,
                'head_pose_data': 1,
                'timestamp': 1,
                'status': 1,
                'description': 1
            }
        ).sort('timestamp', -1).limit(limit))
        
        # Enrich with student names and assessment titles
        for violation in violations:
            # Get student name
            student_id = violation.get('student_id')
            if student_id:
                student = mongo.db.users.find_one({'user_id': student_id}, {'name': 1})
                violation['student_name'] = student.get('name', 'Unknown') if student else 'Unknown'
            else:
                violation['student_name'] = 'Unknown'
            
            # Get assessment title
            assessment_code = violation.get('assessment_code')
            if assessment_code:
                assessment = mongo.db.assessments.find_one({'assessment_code': assessment_code}, {'title': 1})
                violation['assessment_title'] = assessment.get('title', 'Unknown') if assessment else 'Unknown'
            else:
                violation['assessment_title'] = 'Unknown'
            
            # Calculate risk level based on timeline position
            timeline_position = violation.get('exam_timeline_position', 'unknown')
            if timeline_position == 'exam_start':
                violation['risk_level'] = 'Low'
                violation['risk_interpretation'] = 'Early stage - Less suspicious'
            elif timeline_position == 'exam_middle':
                violation['risk_level'] = 'Medium'
                violation['risk_interpretation'] = 'Middle stage - Moderate concern'
            elif timeline_position == 'exam_end':
                violation['risk_level'] = 'High'
                violation['risk_interpretation'] = 'Late stage - High concern'
            else:
                violation['risk_level'] = 'Unknown'
                violation['risk_interpretation'] = 'Timeline position unknown'
            
            # Convert ObjectId to string
            violation['_id'] = str(violation['_id'])
        
        return violations
        
    except Exception as e:
        return []


# ============================
# EXAM EVIDENCE REPORT
# ============================
@app.route('/lecturer/reports/exam-evidence')
def exam_evidence_report():
    """Render the exam evidence report page"""
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return redirect(url_for('login'))
    
    return render_template('reports/exam_evidence_report.html')

@app.route('/api/evidence/violations', methods=['GET'])
def get_violations_with_evidence():
    """Get all violations with their evidence files and student info"""
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        # Get filter parameters
        assessment_code = request.args.get('assessment_code', '')
        student_id = request.args.get('student_id', '')
        violation_type = request.args.get('violation_type', '')
        status = request.args.get('status', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))
        
        # Build aggregation pipeline
        match_conditions = {}
        
        if assessment_code:
            match_conditions['assessment_code'] = assessment_code
        if student_id:
            if student_id.isdigit():
                match_conditions['student_id'] = student_id
        if violation_type:
            match_conditions['violation_type'] = violation_type
        if status:
            match_conditions['status'] = status
        
        # Date filtering
        if date_from or date_to:
            date_filter = {}
            if date_from:
                date_filter['$gte'] = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
            if date_to:
                date_filter['$lte'] = datetime.fromisoformat(date_to.replace('Z', '+00:00'))
            match_conditions['timestamp'] = date_filter
        
        # Aggregation pipeline
        pipeline = [
            {'$match': match_conditions},
            {
                '$lookup': {
                    'from': 'exam_evidence',
                    'localField': '_id',
                    'foreignField': 'violation_id',
                    'as': 'evidence_records'
                }
            },
            {
                '$lookup': {
                    'from': 'users',
                    'localField': 'student_id',
                    'foreignField': 'user_id',
                    'as': 'student_info'
                }
            },
            {
                '$lookup': {
                    'from': 'assessments',
                    'localField': 'assessment_code',
                    'foreignField': 'assessment_code',
                    'as': 'assessment_info'
                }
            },
            {
                '$addFields': {
                    'student_name': {'$arrayElemAt': ['$student_info.name', 0]},
                    'assessment_title': {'$arrayElemAt': ['$assessment_info.title', 0]},
                    'evidence_files': {
                        '$reduce': {
                            'input': '$evidence_records.files',
                            'initialValue': [],
                            'in': {'$concatArrays': ['$$value', '$$this']}
                        }
                    },
                    'has_evidence': {'$gt': [{'$size': '$evidence_records'}, 0]}
                }
            }
        ]
        
        # Handle name search if student_id is not numeric
        if student_id and not student_id.isdigit():
            pipeline.append({
                '$match': {
                    'student_name': {'$regex': student_id, '$options': 'i'}
                }
            })
        
        # Add sorting and pagination
        pipeline.extend([
            {'$sort': {'timestamp': -1}},
            {'$skip': (page - 1) * per_page},
            {'$limit': per_page}
        ])
        
        violations = list(mongo.db.exam_violations.aggregate(pipeline))
        
        # Get total count for pagination
        count_pipeline = [
            {'$match': match_conditions},
            {
                '$lookup': {
                    'from': 'users',
                    'localField': 'student_id',
                    'foreignField': 'user_id',
                    'as': 'student_info'
                }
            },
            {
                '$addFields': {
                    'student_name': {'$arrayElemAt': ['$student_info.name', 0]}
                }
            }
        ]
        
        # Handle name search for count
        if student_id and not student_id.isdigit():
            count_pipeline.append({
                '$match': {
                    'student_name': {'$regex': student_id, '$options': 'i'}
                }
            })
        
        count_pipeline.append({'$count': 'total'})
        
        total_count = list(mongo.db.exam_violations.aggregate(count_pipeline))
        total = total_count[0]['total'] if total_count else 0
        
        # Format response
        formatted_violations = []
        for violation in violations:
            formatted_violations.append({
                '_id': str(violation['_id']),
                'assessment_code': violation.get('assessment_code', ''),
                'assessment_title': violation.get('assessment_title', 'Unknown Assessment'),
                'student_id': violation.get('student_id', ''),
                'student_name': violation.get('student_name', 'Unknown Student'),
                'violation_type': violation.get('violation_type', ''),
                'description': violation.get('description', ''),
                'timestamp': violation.get('timestamp', '').isoformat() if violation.get('timestamp') else '',
                'status': violation.get('status', 'pending'),
                'notes': violation.get('notes', ''),
                'has_evidence': violation.get('has_evidence', False),
                'evidence_files': violation.get('evidence_files', []),
                'head_pose_data': violation.get('head_pose_data'),
                'exam_progress_percent': violation.get('exam_progress_percent')
            })
        
        return jsonify({
            'success': True,
            'violations': formatted_violations,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'pages': (total + per_page - 1) // per_page
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/evidence/review', methods=['POST'])
def review_violation():
    """Review a single violation - mark as reviewed or dismissed"""
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        data = request.get_json()
        violation_id = data.get('violation_id')
        status = data.get('status')  # 'pending', 'reviewed', or 'dismissed'
        notes = data.get('notes', '')
        
        if not violation_id or not status:
            return jsonify({'success': False, 'error': 'Missing required fields'}), 400
        
        if status not in ['pending', 'reviewed', 'dismissed']:
            return jsonify({'success': False, 'error': 'Invalid status'}), 400
        
        # Update violation (removed reviewed_by and reviewed_at)
        result = mongo.db.exam_violations.update_one(
            {'_id': ObjectId(violation_id)},
            {
                '$set': {
                    'status': status,
                    'notes': notes
                }
            }
        )
        
        if result.matched_count == 0:
            return jsonify({'success': False, 'error': 'Violation not found'}), 404
        
        return jsonify({'success': True, 'message': 'Violation reviewed successfully'})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/evidence/export', methods=['GET'])
def export_evidence_report():
    """Export evidence report as CSV"""
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        # Get all violations with evidence (apply same filters as main report)
        assessment_code = request.args.get('assessment_code', '')
        student_id = request.args.get('student_id', '')
        violation_type = request.args.get('violation_type', '')
        status = request.args.get('status', '')
        
        match_conditions = {}
        if assessment_code:
            match_conditions['assessment_code'] = assessment_code
        if student_id:
            if student_id.isdigit():
                match_conditions['student_id'] = student_id
        if violation_type:
            match_conditions['violation_type'] = violation_type
        if status:
            match_conditions['status'] = status
        
        # Get violations data
        pipeline = [
            {'$match': match_conditions},
            {
                '$lookup': {
                    'from': 'users',
                    'localField': 'student_id',
                    'foreignField': 'user_id',
                    'as': 'student_info'
                }
            },
            {
                '$lookup': {
                    'from': 'assessments',
                    'localField': 'assessment_code',
                    'foreignField': 'assessment_code',
                    'as': 'assessment_info'
                }
            },
            {
                '$addFields': {
                    'student_name': {'$arrayElemAt': ['$student_info.name', 0]},
                    'assessment_title': {'$arrayElemAt': ['$assessment_info.title', 0]}
                }
            }
        ]
        
        # Handle name search for export
        if student_id and not student_id.isdigit():
            pipeline.append({
                '$match': {
                    'student_name': {'$regex': student_id, '$options': 'i'}
                }
            })
        
        pipeline.append({'$sort': {'timestamp': -1}})
        
        violations = list(mongo.db.exam_violations.aggregate(pipeline))
        
        # Create CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            'Violation ID', 'Assessment Code', 'Assessment Title', 'Student ID', 
            'Student Name', 'Violation Type', 'Description', 'Timestamp', 
            'Status', 'Notes'
        ])
        
        # Write data
        for violation in violations:
            writer.writerow([
                str(violation['_id']),
                violation.get('assessment_code', ''),
                violation.get('assessment_title', ''),
                violation.get('student_id', ''),
                violation.get('student_name', ''),
                violation.get('violation_type', ''),
                violation.get('description', ''),
                violation.get('timestamp', ''),
                violation.get('status', ''),
                violation.get('notes', '')
            ])
        
        # Prepare file for download
        output.seek(0)
        
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'evidence_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/assessments/list-for-violations', methods=['GET'])
def get_assessments_for_evidence_violations():
    """Get list of assessments that have violations for filter dropdown"""
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 403
    
    try:
        # Get all assessments that have violations
        assessments_with_violations = list(mongo.db.exam_violations.aggregate([  
            {'$group': {'_id': '$assessment_code'}},
            {'$lookup': {
                'from': 'assessments',
                'localField': '_id',
                'foreignField': 'assessment_code',
                'as': 'assessment_info'
            }},
            {'$unwind': {'path': '$assessment_info', 'preserveNullAndEmptyArrays': True}},
            {'$project': {
                'assessment_code': '$_id',
                'title': {'$ifNull': ['$assessment_info.title', 'Unknown Assessment']}
            }},
            {'$sort': {'assessment_code': 1}}
        ]))
        
        return jsonify({
            'success': True,
            'assessments': assessments_with_violations
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Serve static files properly
@app.route('/static/evidence/videos/<filename>')
def serve_video(filename):
    """Serve video files with proper MIME type"""
    try:
        response = send_from_directory('static/evidence/videos', filename)
        if filename.endswith('.webm'):
            response.headers['Content-Type'] = 'video/webm'
        elif filename.endswith('.mp4'):
            response.headers['Content-Type'] = 'video/mp4'
        return response
    except Exception as e:
        return jsonify({'error': 'Video file not found'}), 404

@app.route('/static/evidence/screenshots/<filename>')
def serve_screenshot(filename):
    """Serve screenshot files"""
    try:
        return send_from_directory('static/evidence/screenshots', filename)
    except Exception as e:
        return jsonify({'error': 'Screenshot file not found'}), 404


# ============================
# INDIVIDUAL STUDENT BEHAVIOR REPORTS
# ============================
@app.route('/lecturer/reports/individual-student-behavior')
def individual_student_behavior_report():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))
    return render_template('reports/individual_student_behavior_report.html')

@app.route('/api/students/list', methods=['GET'])
def get_students_list():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 403
    
    try:
        # Get all students
        students = list(mongo.db.users.find(
            {'role': 'student'}, 
            {'user_id': 1, 'name': 1, '_id': 0}
        ).sort('name', 1))
        
        return jsonify({
            'success': True,
            'students': [{'student_id': s['user_id'], 'name': s['name']} for s in students]
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reports/student-behavior', methods=['GET'])
def get_student_behavior_data():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 403
    
    try:
        # Get query parameters for filtering
        student_filter = request.args.get('student', 'all')
        violation_type_filter = request.args.get('violation_type', 'all')
        date_range_filter = request.args.get('date_range', 'all')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Get all analysis data
        students_data = get_student_behavior_overview(student_filter, violation_type_filter, date_range_filter, start_date, end_date)
        summary_data = get_behavior_summary_stats(student_filter, violation_type_filter, date_range_filter, start_date, end_date)
        timeline_data = get_violations_timeline_data(student_filter, violation_type_filter, date_range_filter, start_date, end_date)
        types_data = get_violation_types_distribution(student_filter, violation_type_filter, date_range_filter, start_date, end_date)
        assessments_data = get_violations_by_assessment(student_filter, violation_type_filter, date_range_filter, start_date, end_date)
        risk_levels_data = get_risk_level_distribution(student_filter, violation_type_filter, date_range_filter, start_date, end_date)
        high_risk_students = get_high_risk_students(student_filter, violation_type_filter, date_range_filter, start_date, end_date)
        
        return jsonify({
            'success': True,
            'data': {
                'students': students_data,
                'summary': summary_data,
                'timeline': timeline_data,
                'violation_types': types_data,
                'assessments': assessments_data,
                'risk_levels': risk_levels_data,
                'high_risk_students': high_risk_students
            },
            'filters': {
                'student': student_filter,
                'violation_type': violation_type_filter,
                'date_range': date_range_filter,
                'start_date': start_date,
                'end_date': end_date
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reports/student-behavior/individual', methods=['GET'])
def get_individual_student_behavior():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 403
    
    try:
        student_id = request.args.get('student_id')
        if not student_id:
            return jsonify({'success': False, 'error': 'Student ID required'}), 400
        
        # Get individual student detailed data
        individual_data = get_individual_student_analysis(student_id)
        
        return jsonify({
            'success': True,
            'data': individual_data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/evidence/<violation_id>', methods=['GET'])
def get_violation_evidence(violation_id):
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 403
    
    try:
        from bson import ObjectId
        
        # Get evidence for specific violation
        evidence = mongo.db.exam_evidence.find_one({'violation_id': ObjectId(violation_id)})
        
        if not evidence:
            return jsonify({'success': False, 'error': 'Evidence not found'}), 404
        
        return jsonify({
            'success': True,
            'data': {
                'violation_id': str(evidence['violation_id']),
                'files': evidence.get('files', []),
                'total_files': evidence.get('total_files', 0),
                'total_size': evidence.get('total_size', 0),
                'timestamp': evidence.get('timestamp')
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    
def get_date_filter(date_range_filter, start_date, end_date):
    """
    Build date filter based on parameters
    """
    date_filter = {}
    
    if date_range_filter == 'today':
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        tomorrow = today + timedelta(days=1)
        date_filter = {
            '$gte': today,
            '$lt': tomorrow
        }
    elif date_range_filter == 'week':
        week_ago = datetime.now() - timedelta(days=7)
        date_filter = {'$gte': week_ago}
    elif date_range_filter == 'month':
        month_ago = datetime.now() - timedelta(days=30)
        date_filter = {'$gte': month_ago}
    elif date_range_filter == 'custom' and start_date and end_date:
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        date_filter = {
            '$gte': start,
            '$lt': end
        }
    
    return date_filter

def get_student_behavior_overview(student_filter, violation_type_filter, date_range_filter, start_date, end_date):
    """
    Get student behavior overview data
    """
    try:
        # Build match criteria
        match_criteria = {}
        
        if student_filter != 'all':
            match_criteria['student_id'] = student_filter
        
        if violation_type_filter != 'all':
            match_criteria['violation_type'] = violation_type_filter
        
        date_filter = get_date_filter(date_range_filter, start_date, end_date)
        if date_filter:
            match_criteria['timestamp'] = date_filter
        
        # Get all violations with match criteria
        violations = list(mongo.db.exam_violations.find(match_criteria))
        
        # Get all students who have violations or are in assessments
        students_with_data = {}
        
        # Process violations data
        for violation in violations:
            student_id = violation.get('student_id')
            violation_type = violation.get('violation_type', '').lower()
            
            if student_id not in students_with_data:
                # Get student info
                student = mongo.db.users.find_one({'user_id': student_id})
                students_with_data[student_id] = {
                    'student_id': student_id,
                    'student_name': student.get('name', 'Unknown') if student else 'Unknown',
                    'total_violations': 0,
                    'looking_away': 0,
                    'tab_switch': 0,
                    'forbidden_shortcut': 0,
                    'webcam_denied': 0,
                    'screen_share_stopped': 0,
                    'last_violation': None
                }
            
            # Count violations by type
            students_with_data[student_id]['total_violations'] += 1
            
            if 'looking_away' in violation_type:
                students_with_data[student_id]['looking_away'] += 1
            elif 'tab_switch' in violation_type:
                students_with_data[student_id]['tab_switch'] += 1
            elif 'forbidden_shortcut' in violation_type:
                students_with_data[student_id]['forbidden_shortcut'] += 1
            elif 'webcam_denied' in violation_type:
                students_with_data[student_id]['webcam_denied'] += 1
            elif 'screen_share_stopped' in violation_type:
                students_with_data[student_id]['screen_share_stopped'] += 1
            
            # Track last violation
            violation_time = violation.get('timestamp')
            if violation_time:
                if (not students_with_data[student_id]['last_violation'] or 
                    violation_time > students_with_data[student_id]['last_violation']):
                    students_with_data[student_id]['last_violation'] = violation_time
        
        # Convert to list and sort by total violations (descending)
        students_list = list(students_with_data.values())
        students_list.sort(key=lambda x: x['total_violations'], reverse=True)
        
        return students_list
        
    except Exception as e:
        print(f"Error in get_student_behavior_overview: {e}")
        return []

def get_behavior_summary_stats(student_filter, violation_type_filter, date_range_filter, start_date, end_date):
    """
    Calculate behavior summary statistics
    """
    try:
        # Build match criteria
        match_criteria = {}
        
        if student_filter != 'all':
            match_criteria['student_id'] = student_filter
        
        if violation_type_filter != 'all':
            match_criteria['violation_type'] = violation_type_filter
        
        date_filter = get_date_filter(date_range_filter, start_date, end_date)
        if date_filter:
            match_criteria['timestamp'] = date_filter
        
        # Get all violations
        violations = list(mongo.db.exam_violations.find(match_criteria))
        
        # Calculate statistics
        total_violations = len(violations)
        
        # Get unique students from violations
        unique_students = set(v.get('student_id') for v in violations if v.get('student_id'))
        total_students = len(unique_students)
        
        # Calculate average violations per student
        avg_violations = total_violations / total_students if total_students > 0 else 0
        
        # Count high-risk students (>= 10 violations)
        student_violation_counts = {}
        for violation in violations:
            student_id = violation.get('student_id')
            if student_id:
                student_violation_counts[student_id] = student_violation_counts.get(student_id, 0) + 1
        
        high_risk_students = sum(1 for count in student_violation_counts.values() if count >= 10)
        
        return {
            'total_students': total_students,
            'total_violations': total_violations,
            'avg_violations': round(avg_violations, 2),
            'high_risk_students': high_risk_students
        }
        
    except Exception as e:
        print(f"Error in get_behavior_summary_stats: {e}")
        return {
            'total_students': 0,
            'total_violations': 0,
            'avg_violations': 0,
            'high_risk_students': 0
        }

def get_violations_timeline_data(student_filter, violation_type_filter, date_range_filter, start_date, end_date):
    """
    Get violations timeline data for chart - MONTHLY instead of daily
    """
    try:
        
        # Build match criteria
        match_criteria = {}
        
        if student_filter != 'all':
            match_criteria['student_id'] = student_filter
        
        if violation_type_filter != 'all':
            match_criteria['violation_type'] = violation_type_filter
        
        date_filter = get_date_filter(date_range_filter, start_date, end_date)
        if date_filter:
            match_criteria['timestamp'] = date_filter
        
        # Get violations
        violations = list(mongo.db.exam_violations.find(match_criteria))
        
        # Group by month instead of day
        monthly_counts = defaultdict(int)
        
        for violation in violations:
            timestamp = violation.get('timestamp')
            if timestamp:
                # Convert to date string
                if isinstance(timestamp, str):
                    # Try to parse string timestamp
                    try:
                        timestamp = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                    except:
                        continue
                
                month_str = timestamp.strftime('%Y-%m')  # Changed to monthly format
                monthly_counts[month_str] += 1
        
        # Sort months and prepare data
        sorted_months = sorted(monthly_counts.keys())
        
        return {
            'labels': sorted_months,
            'violations': [monthly_counts[month] for month in sorted_months]
        }
        
    except Exception as e:
        print(f"Error in get_violations_timeline_data: {e}")
        return {'labels': [], 'violations': []}

def get_violation_types_distribution(student_filter, violation_type_filter, date_range_filter, start_date, end_date):
    """
    Get violation types distribution for pie chart
    """
    try:
        # Build match criteria
        match_criteria = {}
        
        if student_filter != 'all':
            match_criteria['student_id'] = student_filter
        
        if violation_type_filter != 'all':
            match_criteria['violation_type'] = violation_type_filter
        
        date_filter = get_date_filter(date_range_filter, start_date, end_date)
        if date_filter:
            match_criteria['timestamp'] = date_filter
        
        # Get violations
        violations = list(mongo.db.exam_violations.find(match_criteria))
        
        # Count by type
        type_counts = {
            'looking_away': 0,
            'tab_switch': 0,
            'forbidden_shortcut': 0,
            'webcam_denied': 0,
            'screen_share_stopped': 0
        }
        
        for violation in violations:
            violation_type = violation.get('violation_type', '').lower()
            
            if 'looking_away' in violation_type:
                type_counts['looking_away'] += 1
            elif 'tab_switch' in violation_type:
                type_counts['tab_switch'] += 1
            elif 'forbidden_shortcut' in violation_type:
                type_counts['forbidden_shortcut'] += 1
            elif 'webcam_denied' in violation_type:
                type_counts['webcam_denied'] += 1
            elif 'screen_share_stopped' in violation_type:
                type_counts['screen_share_stopped'] += 1
        
        return type_counts
        
    except Exception as e:
        print(f"Error in get_violation_types_distribution: {e}")
        return {
            'looking_away': 0,
            'tab_switch': 0,
            'forbidden_shortcut': 0,
            'webcam_denied': 0,
            'screen_share_stopped': 0
        }

def get_violations_by_assessment(student_filter, violation_type_filter, date_range_filter, start_date, end_date):
    """
    Get violations grouped by assessment for bar chart
    """
    try:
        
        # Build match criteria
        match_criteria = {}
        
        if student_filter != 'all':
            match_criteria['student_id'] = student_filter
        
        if violation_type_filter != 'all':
            match_criteria['violation_type'] = violation_type_filter
        
        date_filter = get_date_filter(date_range_filter, start_date, end_date)
        if date_filter:
            match_criteria['timestamp'] = date_filter
        
        # Get violations
        violations = list(mongo.db.exam_violations.find(match_criteria))
        
        # Group by assessment
        assessment_counts = defaultdict(int)
        
        for violation in violations:
            assessment_code = violation.get('assessment_code', 'Unknown')
            assessment_counts[assessment_code] += 1
        
        # Sort by count (descending) and take top 10
        sorted_assessments = sorted(assessment_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        
        return {
            'labels': [item[0] for item in sorted_assessments],
            'violations': [item[1] for item in sorted_assessments]
        }
        
    except Exception as e:
        print(f"Error in get_violations_by_assessment: {e}")
        return {'labels': [], 'violations': []}

def get_risk_level_distribution(student_filter, violation_type_filter, date_range_filter, start_date, end_date):
    """
    Get risk level distribution for doughnut chart (replaces time of day)
    """
    try:
        # Get student behavior overview data first
        students_data = get_student_behavior_overview(student_filter, violation_type_filter, date_range_filter, start_date, end_date)
        
        # Count students by risk level
        risk_counts = {
            'low': 0,       # 0-4 violations
            'medium': 0,    # 5-9 violations  
            'high': 0,      # 10-14 violations
            'critical': 0   # 15+ violations
        }
        
        for student in students_data:
            total_violations = student.get('total_violations', 0)
            
            if total_violations >= 15:
                risk_counts['critical'] += 1
            elif total_violations >= 10:
                risk_counts['high'] += 1
            elif total_violations >= 5:
                risk_counts['medium'] += 1
            else:
                risk_counts['low'] += 1
        
        return risk_counts
        
    except Exception as e:
        print(f"Error in get_risk_level_distribution: {e}")
        return {'low': 0, 'medium': 0, 'high': 0, 'critical': 0}

def get_high_risk_students(student_filter, violation_type_filter, date_range_filter, start_date, end_date):
    """
    Get high-risk students (>= 10 violations)
    """
    try:
        # Build match criteria
        match_criteria = {}
        
        if student_filter != 'all':
            match_criteria['student_id'] = student_filter
        
        if violation_type_filter != 'all':
            match_criteria['violation_type'] = violation_type_filter
        
        date_filter = get_date_filter(date_range_filter, start_date, end_date)
        if date_filter:
            match_criteria['timestamp'] = date_filter
        
        # Get violations
        violations = list(mongo.db.exam_violations.find(match_criteria))
        
        # Count violations per student
        student_violation_counts = {}
        for violation in violations:
            student_id = violation.get('student_id')
            if student_id:
                student_violation_counts[student_id] = student_violation_counts.get(student_id, 0) + 1
        
        # Get high-risk students (>= 10 violations)
        high_risk_students = []
        
        for student_id, count in student_violation_counts.items():
            if count >= 10:
                student = mongo.db.users.find_one({'user_id': student_id})
                student_name = student.get('name', 'Unknown') if student else 'Unknown'
                
                high_risk_students.append({
                    'student_id': student_id,
                    'student_name': student_name,
                    'total_violations': count
                })
        
        # Sort by violation count (descending)
        high_risk_students.sort(key=lambda x: x['total_violations'], reverse=True)
        
        return high_risk_students
        
    except Exception as e:
        print(f"Error in get_high_risk_students: {e}")
        return []

def get_individual_student_analysis(student_id):
    """
    Get detailed analysis for individual student
    """
    try:
        # Get student info
        student = mongo.db.users.find_one({'user_id': student_id})
        if not student:
            return {'error': 'Student not found'}
        
        student_name = student.get('name', 'Unknown')
        
        # Get all violations for this student
        violations = list(mongo.db.exam_violations.find(
            {'student_id': student_id}
        ).sort('timestamp', -1))
        
        total_violations = len(violations)
        
        # Get unique assessments (exams taken)
        unique_assessments = set(v.get('assessment_code') for v in violations if v.get('assessment_code'))
        exams_taken = len(unique_assessments)
        
        # Calculate average per exam
        avg_per_exam = total_violations / exams_taken if exams_taken > 0 else 0
        
        # Find most common violation type
        type_counts = {}
        for violation in violations:
            v_type = violation.get('violation_type', 'Unknown')
            type_counts[v_type] = type_counts.get(v_type, 0) + 1
        
        most_common_type = max(type_counts.keys(), key=lambda k: type_counts[k]) if type_counts else 'N/A'
        
        # Prepare timeline data (monthly for individual)
        from datetime import datetime
        from collections import defaultdict
        
        monthly_counts = defaultdict(int)
        
        for violation in violations:
            timestamp = violation.get('timestamp')
            if timestamp:
                if isinstance(timestamp, str):
                    try:
                        timestamp = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                    except:
                        continue
                
                month_str = timestamp.strftime('%Y-%m')
                monthly_counts[month_str] += 1
        
        sorted_months = sorted(monthly_counts.keys())
        timeline_data = {
            'labels': sorted_months,
            'violations': [monthly_counts[month] for month in sorted_months]
        }
        
        # Prepare individual violations data with evidence
        violations_data = []
        
        for violation in violations:
            # Get evidence count
            evidence_count = mongo.db.exam_evidence.count_documents({
                'violation_id': violation.get('_id')
            })
            
            violations_data.append({
                '_id': str(violation.get('_id')),
                'assessment_code': violation.get('assessment_code', 'N/A'),
                'violation_type': violation.get('violation_type', 'Unknown'),
                'description': violation.get('description', 'No description'),
                'timestamp': violation.get('timestamp'),
                'exam_progress_percent': violation.get('exam_progress_percent'),
                'head_pose_data': violation.get('head_pose_data'),
                'status': violation.get('status', 'pending'),
                'evidence_count': evidence_count
            })
        
        return {
            'student_id': student_id,
            'student_name': student_name,
            'total_violations': total_violations,
            'exams_taken': exams_taken,
            'avg_per_exam': round(avg_per_exam, 1),
            'most_common_type': most_common_type,
            'timeline': timeline_data,
            'violations': violations_data
        }
        
    except Exception as e:
        print(f"Error in get_individual_student_analysis: {e}")
        return {'error': str(e)}


# ============================
# PLAGIARISM & SIMILARITY OVERVIEW REPORTS
# ============================

@app.route('/lecturer/reports/plagiarism-overview')
def plagiarism_overview_report():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        flash("Unauthorized access.", "danger")
        return redirect(url_for('login'))
    return render_template('reports/plagiarism_overview_report.html')

@app.route('/api/reports/plagiarism-overview', methods=['GET'])
def get_plagiarism_overview_data():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 403
    
    try:
        # Get query parameters for filtering
        student_filter = request.args.get('student_id', 'all')
        assessment_filter = request.args.get('assessment_code', 'all')
        plagiarism_level = request.args.get('plagiarism_level', 'all')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        
        # Get all analysis data
        plagiarism_data = get_plagiarism_submissions_data(student_filter, assessment_filter, plagiarism_level, date_from, date_to)
        similarity_data = get_similarity_resubmissions_data(student_filter, assessment_filter, date_from, date_to)
        summary_stats = get_plagiarism_summary_stats(student_filter, assessment_filter, plagiarism_level, date_from, date_to)
        chart_data = get_plagiarism_chart_data(student_filter, assessment_filter, plagiarism_level, date_from, date_to)
        alerts = get_plagiarism_alerts(student_filter, assessment_filter, plagiarism_level, date_from, date_to)
        
        return jsonify({
            'success': True,
            'summary_stats': summary_stats,
            'chart_data': chart_data,
            'plagiarism_table': plagiarism_data,
            'similarity_table': similarity_data,
            'alerts': alerts,
            'filters': {
                'student_id': student_filter,
                'assessment_code': assessment_filter,
                'plagiarism_level': plagiarism_level,
                'date_from': date_from,
                'date_to': date_to
            }
        })
        
    except Exception as e:
        print(f"Error in get_plagiarism_overview_data: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reports/export-plagiarism-csv', methods=['GET'])
def export_plagiarism_csv():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 403
    
    try:
        # Get same filters as main report
        student_filter = request.args.get('student_id', 'all')
        assessment_filter = request.args.get('assessment_code', 'all')
        plagiarism_level = request.args.get('plagiarism_level', 'all')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        
        # Get plagiarism data
        plagiarism_data = get_plagiarism_submissions_data(student_filter, assessment_filter, plagiarism_level, date_from, date_to)
        
        # Prepare CSV data
        csv_data = []
        for item in plagiarism_data:
            csv_data.append({
                'Student ID': item['student_id'],
                'Student Name': item['student_name'],
                'Assessment Code': item['assessment_code'],
                'Filename': item['filename'],
                'Submitted Date': item['submitted_at'],
                'Total Plagiarism Score': f"{item['total_plagiarism_score']}%",
                'Base Plagiarism Score': f"{item['base_plagiarism_score']}%",
                'Quotes Score': f"{item['quotes_score']}%",
                'Level': item['plagiarism_level'],
                'Report URL': item['report_url']
            })
        
        return jsonify({
            'success': True,
            'csv_data': csv_data,
            'filename': f'plagiarism_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reports/export-similarity-csv', methods=['GET'])
def export_similarity_csv():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 403
    
    try:
        # Get same filters as main report
        student_filter = request.args.get('student_id', 'all')
        assessment_filter = request.args.get('assessment_code', 'all')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        
        # Get similarity data
        similarity_data = get_similarity_resubmissions_data(student_filter, assessment_filter, date_from, date_to)
        
        # Prepare CSV data
        csv_data = []
        for item in similarity_data:
            csv_data.append({
                'Student ID': item['student_id'],
                'Student Name': item['student_name'],
                'Assessment Code': item['assessment_code'],
                'Original Filename': item['old_filename'],
                'Resubmitted Filename': item['new_filename'],
                'Similarity Score': f"{item['similarity_score']}%",
                'Level': item['similarity_level'],
                'Status': item['status'],
                'Requested Date': item['requested_at']
            })
        
        return jsonify({
            'success': True,
            'csv_data': csv_data,
            'filename': f'similarity_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Add missing API endpoints
@app.route('/api/students/list', methods=['GET'])
def api_get_students_list():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 403
    
    try:
        # Get all students
        students = list(mongo.db.users.find({'role': 'student'}, {'user_id': 1, 'name': 1}))
        
        student_list = []
        for student in students:
            student_list.append({
                'student_id': student['user_id'],
                'name': student.get('name', 'Unknown')
            })
        
        return jsonify({
            'success': True,
            'students': student_list
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/assessments/list', methods=['GET'])
def api_get_assessments_list():
    if 'user_id' not in session or session.get('role') != 'lecturer':
        return jsonify({'success': False, 'error': 'Unauthorized access'}), 403
    
    try:
        # Get all assessments
        assessments = list(mongo.db.assessments.find({}, {'assessment_code': 1}))
        
        assessment_list = []
        for assessment in assessments:
            assessment_list.append({
                'assessment_code': assessment.get('assessment_code', 'Unknown')
            })
        
        return jsonify({
            'success': True,
            'assessments': assessment_list
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

def get_plagiarism_date_filter(date_from, date_to):
    """
    Build date filter for plagiarism queries with proper Malaysia timezone handling
    """
    
    # Use Malaysia timezone
    malaysia_tz = pytz.timezone('Asia/Kuala_Lumpur')
    now = datetime.now(malaysia_tz)
    
    date_filter = {}
    
    if date_from:
        try:
            # Handle special date ranges
            if date_from == 'today':
                # Today from 00:00:00 to 23:59:59
                start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
                end_of_day = now.replace(hour=23, minute=59, second=59, microsecond=999999)
                date_filter = {'$gte': start_of_day, '$lte': end_of_day}
            elif date_from == 'week':
                # Last 7 days from 7 days ago 00:00:00 to now
                week_ago = (now - timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
                date_filter = {'$gte': week_ago, '$lte': now}
            elif date_from == 'month':
                # Last 30 days from 30 days ago 00:00:00 to now
                month_ago = (now - timedelta(days=30)).replace(hour=0, minute=0, second=0, microsecond=0)
                date_filter = {'$gte': month_ago, '$lte': now}
            else:
                # Custom date - parse as date only (no time)
                start_date = datetime.strptime(date_from, '%Y-%m-%d')
                # Set to start of day in Malaysia timezone
                start_date = malaysia_tz.localize(start_date.replace(hour=0, minute=0, second=0, microsecond=0))
                date_filter['$gte'] = start_date
        except Exception as e:
            print(f"Error parsing date_from: {e}")
    
    if date_to and date_from not in ['today', 'week', 'month']:
        try:
            # Custom date - parse as date only (no time)
            end_date = datetime.strptime(date_to, '%Y-%m-%d')
            # Set to END of day in Malaysia timezone (23:59:59.999999)
            end_date = malaysia_tz.localize(end_date.replace(hour=23, minute=59, second=59, microsecond=999999))
            date_filter['$lte'] = end_date
        except Exception as e:
            print(f"Error parsing date_to: {e}")
    
    return date_filter if date_filter else None

def get_plagiarism_submissions_data(student_filter, assessment_filter, plagiarism_level, date_from, date_to):
    """
    Get plagiarism submissions data for table with combined plagiarism + quotes score
    """
    try:
        # Build match criteria
        match_criteria = {}
        
        if student_filter != 'all':
            match_criteria['student_id'] = student_filter
        
        if assessment_filter != 'all':
            match_criteria['assessment_code'] = assessment_filter
        
        # Date filter
        date_filter = get_plagiarism_date_filter(date_from, date_to)
        if date_filter:
            match_criteria['submitted_at'] = date_filter
        
        # Get submissions
        submissions = list(mongo.db.submissions.find(match_criteria).sort('submitted_at', -1))
        
        # Process submissions and apply plagiarism level filter with combined scores
        filtered_submissions = []
        for submission in submissions:
            base_score = submission.get('plagiarism_score', 0)
            quotes_score = submission.get('quotes_score', 0)
            total_score = base_score + quotes_score
            
            # Apply plagiarism level filter based on total score
            if plagiarism_level == 'high' and total_score <= 20:
                continue
            elif plagiarism_level == 'low' and total_score > 20:
                continue
            
            # Add total score to submission for processing
            submission['total_plagiarism_score'] = total_score
            filtered_submissions.append(submission)
        
        # Get student information
        student_ids = list(set([s['student_id'] for s in filtered_submissions]))
        students_dict = {}
        
        if student_ids:
            students_cursor = mongo.db.users.find({'user_id': {'$in': student_ids}})
            students_dict = {s['user_id']: s for s in students_cursor}
        
        # Prepare table data
        plagiarism_table = []
        for submission in filtered_submissions:
            student_info = students_dict.get(submission['student_id'], {})
            
            # Get scores
            base_score = submission.get('plagiarism_score', 0)
            quotes_score = submission.get('quotes_score', 0)
            total_score = submission.get('total_plagiarism_score', 0)
            
            # Determine plagiarism level and color based on total score
            if total_score <= 20:
                level = 'Low'
                color_class = 'success'
            elif total_score <= 50:
                level = 'Medium'
                color_class = 'warning'  # Orange
            else:
                level = 'High'
                color_class = 'danger'   # Red
            
            # Format date
            submitted_at = submission.get('submitted_at')
            formatted_date = ''
            if submitted_at:
                try:
                    if isinstance(submitted_at, str):
                        from datetime import datetime
                        submitted_at = datetime.fromisoformat(submitted_at.replace('Z', '+00:00'))
                    formatted_date = submitted_at.strftime('%Y-%m-%d %H:%M')
                except:
                    formatted_date = str(submitted_at)
            
            # Check if plagiarism report exists
            has_report = bool(submission.get('plagiarism_report_url'))
            
            plagiarism_table.append({
                'id': str(submission['_id']),
                'student_id': submission['student_id'],
                'student_name': student_info.get('name', 'Unknown'),
                'assessment_code': submission.get('assessment_code', 'N/A'),
                'filename': submission.get('filename', 'N/A'),
                'submitted_at': formatted_date,
                'total_plagiarism_score': total_score,  # Combined score for display
                'base_plagiarism_score': base_score,    # Base score
                'quotes_score': quotes_score,           # Quotes score
                'plagiarism_level': level,
                'color_class': color_class,
                'report_url': submission.get('plagiarism_report_url', ''),
                'has_report': has_report
            })
        
        return plagiarism_table
        
    except Exception as e:
        print(f"Error in get_plagiarism_submissions_data: {e}")
        traceback.print_exc()
        return []

def get_similarity_resubmissions_data(student_filter, assessment_filter, date_from, date_to):
    """
    Get similarity resubmissions data for table
    """
    try:
        # Build match criteria
        match_criteria = {}
        
        if student_filter != 'all':
            match_criteria['student_id'] = student_filter
        
        if assessment_filter != 'all':
            match_criteria['assessment_code'] = assessment_filter
        
        # Date filter
        date_filter = get_plagiarism_date_filter(date_from, date_to)
        if date_filter:
            match_criteria['requested_at'] = date_filter
        
        # Get resubmission requests
        resubmissions = list(mongo.db.resubmission_requests.find(match_criteria).sort('requested_at', -1))
        
        # Get student information
        student_ids = list(set([r['student_id'] for r in resubmissions]))
        students_dict = {}
        
        if student_ids:
            students_cursor = mongo.db.users.find({'user_id': {'$in': student_ids}})
            students_dict = {s['user_id']: s for s in students_cursor}
        
        # Prepare table data
        similarity_table = []
        for resubmission in resubmissions:
            student_info = students_dict.get(resubmission['student_id'], {})
            
            # Determine similarity level and color
            score = resubmission.get('similarity_score', 0)
            if score <= 30:
                level = 'Very Different'
                color_class = 'success'
            elif score <= 60:
                level = 'Somewhat Similar'
                color_class = 'warning'
            elif score <= 80:
                level = 'Very Similar'
                color_class = 'danger'
            else:
                level = 'Nearly Identical'
                color_class = 'dark'
            
            # Format date
            requested_at = resubmission.get('requested_at')
            formatted_date = ''
            if requested_at:
                try:
                    if isinstance(requested_at, str):
                        from datetime import datetime
                        requested_at = datetime.fromisoformat(requested_at.replace('Z', '+00:00'))
                    formatted_date = requested_at.strftime('%Y-%m-%d %H:%M')
                except:
                    formatted_date = str(requested_at)
            
            # Check if similarity report exists
            has_similarity_report = bool(resubmission.get('similarity_checked_at'))
            
            similarity_table.append({
                'id': str(resubmission['_id']),
                'submission_id': str(resubmission.get('submission_id', '')),
                'student_id': resubmission['student_id'],
                'student_name': student_info.get('name', 'Unknown'),
                'assessment_code': resubmission.get('assessment_code', 'N/A'),
                'old_filename': resubmission.get('old_filename', 'N/A'),
                'new_filename': resubmission.get('new_filename', 'N/A'),
                'similarity_score': score,
                'similarity_level': level,
                'color_class': color_class,
                'status': resubmission.get('status', 'Pending'),
                'requested_at': formatted_date,
                'has_similarity_report': has_similarity_report
            })
        
        return similarity_table
        
    except Exception as e:
        print(f"Error in get_similarity_resubmissions_data: {e}")
        traceback.print_exc()
        return []

def get_plagiarism_summary_stats(student_filter, assessment_filter, plagiarism_level, date_from, date_to):
    """
    Calculate plagiarism summary statistics using combined scores
    """
    try:
        # Build match criteria for submissions
        submissions_match = {}
        if student_filter != 'all':
            submissions_match['student_id'] = student_filter
        if assessment_filter != 'all':
            submissions_match['assessment_code'] = assessment_filter
        
        date_filter = get_plagiarism_date_filter(date_from, date_to)
        if date_filter:
            submissions_match['submitted_at'] = date_filter
        
        # Get all submissions
        submissions = list(mongo.db.submissions.find(submissions_match))
        
        # Process submissions with combined scores
        filtered_submissions = []
        total_score_sum = 0
        high_plagiarism_count = 0
        
        for submission in submissions:
            base_score = submission.get('plagiarism_score', 0)
            quotes_score = submission.get('quotes_score', 0)
            total_score = base_score + quotes_score
            
            # Apply plagiarism level filter
            if plagiarism_level == 'high' and total_score <= 20:
                continue
            elif plagiarism_level == 'low' and total_score > 20:
                continue
            
            filtered_submissions.append(submission)
            total_score_sum += total_score
            
            if total_score > 20:
                high_plagiarism_count += 1
        
        total_submissions = len(filtered_submissions)
        avg_plagiarism_score = total_score_sum / total_submissions if total_submissions > 0 else 0
        
        # Get resubmission requests count
        resubmissions_match = {}
        if student_filter != 'all':
            resubmissions_match['student_id'] = student_filter
        if assessment_filter != 'all':
            resubmissions_match['assessment_code'] = assessment_filter
        if date_filter:
            resubmissions_match['requested_at'] = date_filter
        
        total_resubmissions = mongo.db.resubmission_requests.count_documents(resubmissions_match)
        
        return {
            'total_submissions': total_submissions,
            'high_plagiarism_count': high_plagiarism_count,
            'avg_plagiarism_score': avg_plagiarism_score,  # No rounding - let frontend handle display
            'total_resubmissions': total_resubmissions
        }
        
    except Exception as e:
        print(f"Error in get_plagiarism_summary_stats: {e}")
        return {
            'total_submissions': 0,
            'high_plagiarism_count': 0,
            'avg_plagiarism_score': 0,
            'total_resubmissions': 0
        }

def get_plagiarism_chart_data(student_filter, assessment_filter, plagiarism_level, date_from, date_to):
    """
    Get chart data for plagiarism overview using combined scores
    """
    try:
        from collections import defaultdict
        from datetime import datetime
        
        # Build match criteria
        submissions_match = {}
        if student_filter != 'all':
            submissions_match['student_id'] = student_filter
        if assessment_filter != 'all':
            submissions_match['assessment_code'] = assessment_filter
        
        date_filter = get_plagiarism_date_filter(date_from, date_to)
        if date_filter:
            submissions_match['submitted_at'] = date_filter
        
        # Get submissions data
        submissions = list(mongo.db.submissions.find(submissions_match))
        
        # Process submissions with combined scores
        processed_submissions = []
        for submission in submissions:
            base_score = submission.get('plagiarism_score', 0)
            quotes_score = submission.get('quotes_score', 0)
            total_score = base_score + quotes_score
            
            # Apply plagiarism level filter
            if plagiarism_level == 'high' and total_score <= 20:
                continue
            elif plagiarism_level == 'low' and total_score > 20:
                continue
            
            # Add total score to submission for processing
            submission_copy = submission.copy()
            submission_copy['total_plagiarism_score'] = total_score
            processed_submissions.append(submission_copy)
        
        # 1. Plagiarism Score Distribution using total scores
        plagiarism_distribution = {
            'Low (0-20%)': 0,
            'Medium (21-50%)': 0,
            'High (51-80%)': 0,
            'Very High (81-100%)': 0
        }
        
        for submission in processed_submissions:
            score = submission.get('total_plagiarism_score', 0)
            if score <= 20:
                plagiarism_distribution['Low (0-20%)'] += 1
            elif score <= 50:
                plagiarism_distribution['Medium (21-50%)'] += 1
            elif score <= 80:
                plagiarism_distribution['High (51-80%)'] += 1
            else:
                plagiarism_distribution['Very High (81-100%)'] += 1
        
        # 2. Submissions Timeline (monthly)
        timeline_data = defaultdict(int)
        for submission in processed_submissions:
            submitted_at = submission.get('submitted_at')
            if submitted_at:
                try:
                    if isinstance(submitted_at, str):
                        submitted_at = datetime.fromisoformat(submitted_at.replace('Z', '+00:00'))
                    month_key = submitted_at.strftime('%Y-%m')
                    timeline_data[month_key] += 1
                except:
                    continue
        
        timeline_chart = {
            'labels': sorted(timeline_data.keys()),
            'data': [timeline_data[month] for month in sorted(timeline_data.keys())]
        }
        
        # 3. Plagiarism by Assessment using total scores
        assessment_plagiarism = defaultdict(list)
        for submission in processed_submissions:
            assessment_code = submission.get('assessment_code', 'Unknown')
            total_score = submission.get('total_plagiarism_score', 0)
            assessment_plagiarism[assessment_code].append(total_score)
        
        assessment_chart = {
            'labels': list(assessment_plagiarism.keys()),
            'data': [sum(scores)/len(scores) if scores else 0 for scores in assessment_plagiarism.values()]
        }
        
        # 4. Similarity Score Distribution
        resubmissions_match = {}
        if student_filter != 'all':
            resubmissions_match['student_id'] = student_filter
        if assessment_filter != 'all':
            resubmissions_match['assessment_code'] = assessment_filter
        if date_filter:
            resubmissions_match['requested_at'] = date_filter
        
        resubmissions = list(mongo.db.resubmission_requests.find(resubmissions_match))
        
        similarity_distribution = {
            'Very Different (0-30%)': 0,
            'Somewhat Similar (31-60%)': 0,
            'Very Similar (61-80%)': 0,
            'Nearly Identical (81-100%)': 0
        }
        
        for resubmission in resubmissions:
            score = resubmission.get('similarity_score', 0)
            if score <= 30:
                similarity_distribution['Very Different (0-30%)'] += 1
            elif score <= 60:
                similarity_distribution['Somewhat Similar (31-60%)'] += 1
            elif score <= 80:
                similarity_distribution['Very Similar (61-80%)'] += 1
            else:
                similarity_distribution['Nearly Identical (81-100%)'] += 1
        
        return {
            'plagiarism_distribution': plagiarism_distribution,
            'timeline': timeline_chart,
            'assessment_plagiarism': assessment_chart,
            'similarity_distribution': similarity_distribution
        }
        
    except Exception as e:
        print(f"Error in get_plagiarism_chart_data: {e}")
        return {
            'plagiarism_distribution': {'Low (0-20%)': 0, 'Medium (21-50%)': 0, 'High (51-80%)': 0, 'Very High (81-100%)': 0},
            'timeline': {'labels': [], 'data': []},
            'assessment_plagiarism': {'labels': [], 'data': []},
            'similarity_distribution': {'Very Different (0-30%)': 0, 'Somewhat Similar (31-60%)': 0, 'Very Similar (61-80%)': 0, 'Nearly Identical (81-100%)': 0}
        }

def get_plagiarism_alerts(student_filter, assessment_filter, plagiarism_level, date_from, date_to):
    """
    Get plagiarism alerts for high total plagiarism submissions
    """
    try:
        # Build match criteria
        match_criteria = {}
        if student_filter != 'all':
            match_criteria['student_id'] = student_filter
        if assessment_filter != 'all':
            match_criteria['assessment_code'] = assessment_filter
        
        date_filter = get_plagiarism_date_filter(date_from, date_to)
        if date_filter:
            match_criteria['submitted_at'] = date_filter
        
        # Get all submissions and calculate total scores
        all_submissions = list(mongo.db.submissions.find(match_criteria))
        high_plagiarism_submissions = []
        
        for submission in all_submissions:
            base_score = submission.get('plagiarism_score', 0)
            quotes_score = submission.get('quotes_score', 0)
            total_score = base_score + quotes_score
            
            if total_score > 20:
                high_plagiarism_submissions.append(submission)
        
        alerts = []
        if high_plagiarism_submissions:
            alerts.append({
                'type': 'warning',
                'message': f'Found {len(high_plagiarism_submissions)} submission(s) with high total plagiarism scores (>20%)'
            })
        
        return alerts
        
    except Exception as e:
        print(f"Error in get_plagiarism_alerts: {e}")
        return []

# ============================
# Custom Error Pages
# ============================

# Error handler for 404 - Page Not Found
@app.errorhandler(404)
def page_not_found(error):
    """Handle 404 errors - Page not found"""
    return render_template('error.html', error_code=404, error_message="Page Not Found"), 404

# Error handler for 500 - Internal Server Error
@app.errorhandler(500)
def internal_server_error(error):
    """Handle 500 errors - Internal server error"""
    return render_template('error.html', error_code=500, error_message="Internal Server Error"), 500

# Error handler for 401 - Unauthorized
@app.errorhandler(401)
def unauthorized(error):
    """Handle 401 errors - Authentication required"""
    return render_template('error.html', error_code=401, error_message="Authentication Required"), 401

# Error handler for 403 - Forbidden
@app.errorhandler(403)
def forbidden(error):
    """Handle 403 errors - Access forbidden"""
    return render_template('error.html', error_code=403, error_message="Access Forbidden"), 403

# ============================
# Logout
# ============================
@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

# Run the Flask app
if __name__ == '__main__':
    app.run(debug=True)
